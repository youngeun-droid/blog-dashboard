from __future__ import annotations

import base64
import calendar
import html
import hashlib
import hmac
import io
import json
import math
import os
import re
import subprocess
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen

import google.generativeai as genai
import pandas as pd
import requests
import streamlit as st
import altair as alt
import streamlit.components.v1 as components
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
try:
    from selenium import webdriver
    from selenium.common.exceptions import TimeoutException, WebDriverException
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


load_dotenv()

_BASE_DIR = Path(__file__).resolve().parent


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/123.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 15
NAVER_SEARCHAD_BASE_URL = os.getenv("NAVER_SEARCHAD_BASE_URL", "https://api.naver.com")
NAVER_SEARCHAD_ACCESS_LICENSE = os.getenv("NAVER_SEARCHAD_ACCESS_LICENSE", "")
NAVER_SEARCHAD_SECRET_KEY = os.getenv("NAVER_SEARCHAD_SECRET_KEY", "")
NAVER_SEARCHAD_CUSTOMER_ID = os.getenv("NAVER_SEARCHAD_CUSTOMER_ID", "")
NAVER_DATALAB_CLIENT_ID = os.getenv("NAVER_DATALAB_CLIENT_ID", "")
NAVER_DATALAB_CLIENT_SECRET = os.getenv("NAVER_DATALAB_CLIENT_SECRET", "")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
APP_AUTH_SECRET = os.getenv("APP_AUTH_SECRET") or GEMINI_API_KEY or NAVER_SEARCHAD_SECRET_KEY or "blog_gacha_internal_auth"
WORKBOOK_CANDIDATE_PATHS = [
    str(_BASE_DIR / "assets" / "downloads" / "[세예] 네이버 키워드 발굴.xlsx"),
    str(Path.cwd() / "[세예] 네이버 키워드 발굴.xlsx"),
]
PROMPT_INFO_ROOT = _BASE_DIR / "assets" / "프롬프트정보"
USER_DATA_ROOT = Path.cwd() / "data" / "users"
USER_ACCOUNT_PATH = USER_DATA_ROOT / "accounts.json"
TRANSLATION_DATA_ROOT = Path.cwd() / "data" / "translation"
TRANSLATION_PROMPTS_PATH = TRANSLATION_DATA_ROOT / "prompts.json"
TRANSLATION_GLOSSARIES_PATH = TRANSLATION_DATA_ROOT / "glossaries.json"
TRANSLATION_PRESET_ROOT = _BASE_DIR / "assets" / "번역프리셋"
JAPANESE_GLOSSARY_SEED_PATH = TRANSLATION_PRESET_ROOT / "일본" / "시술용어집_일본.xlsx"
CHINESE_GLOSSARY_SEED_PATH = TRANSLATION_PRESET_ROOT / "중국어 피부과 시술 용어집  - 용어집.csv"
WEBSITE_CONTENT_LANGUAGES = ["한국어", "영어", "일본어", "대만(번체)", "중국(번체)", "중국(간체)"]
WEBSITE_REFERENCE_SEED_FILES: dict[str, list[Path]] = {
    "한국어": [
        _BASE_DIR / "assets" / "downloads" / "블로그 원본 (1).docx",
        _BASE_DIR / "assets" / "downloads" / "웹콘텐츠 변환 (1).docx",
    ]
}
BRAND_INFO_DOC_PATH = PROMPT_INFO_ROOT / "세예의원 기본 정보 및 강점 (1).docx"
BLOG_REFERENCE_DOC_PATH = PROMPT_INFO_ROOT / "블로그 레퍼런스.docx"
HIGH_PERFORMANCE_EXAMPLE_DOC_PATH = PROMPT_INFO_ROOT / "네이버 상위노출 및 조회수 높은 콘텐츠 예시.docx"
PROCEDURE_DOC_DIRECTORIES = {
    "온다리프팅": PROMPT_INFO_ROOT / "온다리프팅",
    "울쎄라": PROMPT_INFO_ROOT / "울쎄라",
    "올타이트": PROMPT_INFO_ROOT / "올타이트",
    "올타이트리프팅": PROMPT_INFO_ROOT / "올타이트",
    "티타늄": PROMPT_INFO_ROOT / "티타늄",
    "티타늄리프팅": PROMPT_INFO_ROOT / "티타늄",
    "써마지": PROMPT_INFO_ROOT / "써마지",
    "리투오": PROMPT_INFO_ROOT / "리투오, 셀르디엠, 레누바",
    "셀르디엠": PROMPT_INFO_ROOT / "리투오, 셀르디엠, 레누바",
    "레누바": PROMPT_INFO_ROOT / "리투오, 셀르디엠, 레누바",
    "힐로웨이브": PROMPT_INFO_ROOT / "힐로웨이브",
    "BBL": PROMPT_INFO_ROOT / "BBL",
    "라풀렌": PROMPT_INFO_ROOT / "라풀렌:엘란쎄",
    "엘란쎄": PROMPT_INFO_ROOT / "라풀렌:엘란쎄",
}
BASE_BLOG_WRITING_PROMPT = """
설명: 너는 콘텐츠 마케터로 피부과 시술 관련 정보성 블로그 콘텐츠를 주로 작성하는 역할을 수행해야 해.

[목적]
너의 목적은 내가 제시한 시술 주제에 대해 의료법 및 보건복지부 광고 가이드라인을 준수하면서, 피부과 원장이 직접 설명하는 듯한 자연스럽고 신뢰감 있는 블로그 글을 작성하는 거야.
단순한 문법 교정이 아니라, SEO 최적화·시술 정보 전달·병원 전문성 반영이 중심이 되어야 해.
아래의 모든 조건을 필수적으로 준수해서 원고에 반영해야 해.

[기본 작성 조건]
- 내가 원고 제목과 주제, 주력키워드를 제시하면, 이를 기반으로 원고를 작성
- 의료법 제56조·제57조에 위반될 수 있는 표현(과장, 비교, 시술 효과 보장, 전후 사진 유사한 표현 등)은 절대 사용하지 말 것.
- 전체 글자수는 '네이버 글자수 세기' 기준 공백 포함 1,300자 내외로 작성
- 문체는 '~해요'체를 사용하여 병원장이 직접 설명하는 듯이 친근하면서도 신뢰감있는 전문적인 톤을 유지
- 전문 용어는 꼭 필요한 경우에만 사용하고, 반드시 쉽게 풀어서 설명해.
- '블로그 레퍼런스' 파일의 내용은 글 구성과 개요를 참고하는 데에만 활용하고, 시술 정보나 차별점 등 세부 내용에 반영하지 말 것

[키워드 반복 조건]
- 원고 전체 기준으로 주력 키워드는 정확히 6회만 사용하고, 문맥상 자연스럽게 배치
- 되도록 한 문단에 2개 이상 들어가지 않도록 함
- 원고 내의 모든 형태소는 최대 16회까지만 반복 허용

[흐름 및 내용 구성 조건]
- 네이버 SEO 상위 노출 전략에 최적화된 흐름과 구성으로 작성
- 각 문단에 흥미 유도 및 내용 소개의 질문형 문장으로 소제목 작성
- 질문형 소제목 다음으로는 간결하고 명확한 답변 문장이 이어지도록 작성

[원고 개요]
1. 도입부
2. 시술 기본 정보 전달
3. 시술 주제 관련 비교 문단(1) + 비교/정리 표
4. 시술 주제 관련 비교 문단(2) + 비교/정리 표
5. 시술 정보 전달 QnA
6. 의원 전문성 및 차별점 어필 문단
7. 결론 요약

[최종 점검]
- 글자수, 주력 키워드 6회, 반복 형태소 16회 이하, 의료법 준수, 구조 반영 여부를 스스로 점검
""".strip()


@dataclass
class SearchAdKeyword:
    keyword: str
    monthly_pc: int
    monthly_mobile: int
    monthly_total: int
    monthly_docs: int
    competition_index: str
    source_seed: str
    saturation: float
    opportunity_score: float
    recommendation_reason: str
    is_fallback: bool = False
    source_type: str = "api"
    keyword_classification: str = "미분류"


@dataclass
class BlogPostMetrics:
    rank: int
    title: str
    url: str
    author: str
    content: str
    char_count: int
    image_count: int
    video_count: int
    gif_count: int
    top_terms: list[tuple[str, int]]


@dataclass
class KeywordDiagnosis:
    keyword: str
    monthly_pc: int
    monthly_mobile: int
    monthly_total: int
    competition_index: str
    estimated_blog_docs: int
    estimated_saturation: float
    channel_dominance: str
    blog_presence: str
    opportunity_label: str
    content_recommendation: str
    writing_format: str
    expected_month_total: int | None
    expected_month_to_date: int | None
    trend_peak_month: str | None
    trend_change_label: str
    trend_rows: list[dict[str, Any]]


DEFAULT_PROCEDURE_PROFILES: dict[str, dict[str, Any]] = {
    "온다리프팅": {
        "core_description": "극초단파 기반 리프팅 시술로 이중턱, 심부볼, 턱선 정리, 피부 타이트닝 맥락을 함께 설명하는 편이 좋습니다.",
        "must_include": ["이중턱", "심부볼", "턱선", "쿨웨이브", "통증 부담", "개인별 맞춤 핸드피스"],
        "avoid_phrases": ["무조건", "반드시 효과", "즉시 완치", "영구 유지"],
        "preferred_format": "비교형 또는 설명형",
        "cta_style": "상담 전 피부 두께와 지방 분포에 따라 달라질 수 있다는 안내를 넣습니다.",
    },
    "울쎄라": {
        "core_description": "초음파 리프팅 특성과 탄력 개선, 리프팅 축, 통증과 강도, 정품 팁/샷수 맥락이 중요합니다.",
        "must_include": ["초음파", "탄력", "리프팅 축", "정품 팁", "샷수", "개인차"],
        "avoid_phrases": ["완벽한 리프팅", "무통", "한 번으로 끝"],
        "preferred_format": "정보형",
        "cta_style": "상담 시 피부 처짐 정도와 원하는 강도에 따른 맞춤 설명을 권합니다.",
    },
    "스킨부스터": {
        "core_description": "피부결, 수분감, 광채, 반복 시술, 통증/붓기, 제품별 차이를 명확히 구분하는 흐름이 좋습니다.",
        "must_include": ["피부결", "수분감", "광채", "붓기", "주기", "제품별 차이"],
        "avoid_phrases": ["물광 보장", "즉시 완벽", "부작용 없음"],
        "preferred_format": "설명형",
        "cta_style": "현재 피부 컨디션과 원하는 결 개선 방향에 따라 제품을 고르는 문장을 넣습니다.",
    },
}

TRANSLATION_PRESETS: dict[str, dict[str, str]] = {
    "대만(번체)": {
        "locale": "zh-TW",
        "style": "대만 현지 독자가 자연스럽게 읽는 번체 중국어로 번역합니다.",
        "notes": "의료 정보성 블로그 톤을 유지하고, 너무 중국 본토식 어휘는 피합니다.",
    },
    "중국(번체)": {
        "locale": "zh-Hant",
        "style": "전통 중국어권 독자가 읽기 쉬운 번체 중국어로 번역합니다.",
        "notes": "표현은 단정적이지 않게, 의료 정보형 블로그 톤을 유지합니다.",
    },
    "중국(간체)": {
        "locale": "zh-CN",
        "style": "중국 본토 독자가 자연스럽게 읽는 간체 중국어로 번역합니다.",
        "notes": "과장 없이 정보 전달 중심으로 번역합니다.",
    },
    "일본어": {
        "locale": "ja-JP",
        "style": "일본 현지 병원 블로그처럼 정중하고 부드러운 문체로 번역합니다.",
        "notes": "과도한 광고조 표현을 피하고 설명형 흐름을 유지합니다.",
    },
    "영어": {
        "locale": "en-US",
        "style": "자연스럽고 읽기 쉬운 영어 블로그 문체로 번역합니다.",
        "notes": "medical advertising tone 대신 informative clinic blog tone을 유지합니다.",
    },
}

DEFAULT_TRANSLATION_PROMPTS: dict[str, str] = {
    "일본어": """
역할(Role):
사용자가 입력하는 모든 텍스트는 한국어로 작성된 시술 관련 블로그 원문 또는 의료 콘텐츠입니다.
당신의 임무는 먼저 한국어로 번역 의도와 핵심 포인트를 요약한 뒤 일본 Google 및 Yahoo JAPAN 검색에서 상단 노출을 목표로 한 SEO·GEO 최적화 일본어 번역 결과를 출력하는 것입니다.
목표 언어(Target Language): Japanese (日本語)

기본 원칙:
- 제공된 병원 기본정보, 시술 정보, 기기명, 운영 정책, 가격 가이드를 철저히 따릅니다.
- 원문에 없는 정보, 효과, 비교, 추천 문구를 절대 추가하지 않습니다.
- 의학적 확정 표현을 피하고 `個人差があります`, `カウンセリング後に決定されます`의 의미를 유지합니다.
- 일본 의료·미용 블로그에서 통용되는 중립적이고 신뢰 중심의 정보형 문체를 사용합니다.
- 과도한 광고 표현은 금지합니다.

시술명 · 기기명 표기 규칙:
- 시술 및 기기명은 반드시 공식 Full Name으로 표기합니다.
- 축약, 생략, 관용 표현 사용을 금지합니다.

SEYE Clinic 표기 분산 규칙:
- 번역 결과에서는 `SEYEクリニック`, `セイェクリニック` 두 표기를 골고루 분산 사용합니다.
- 한 가지 표기만 반복 사용하지 않습니다.
- 제목(H1)과 본문(H2 이하)에 최소 2종 이상 사용합니다.
- 동일 문단 내 중복 사용은 금지합니다.

SEO · GEO 번역 원칙:
- 의미를 변경하지 않는 범위 내에서 일본 검색 사용자가 실제 사용하는 표현을 선택합니다.
- 아래 키워드는 문맥에 맞게 자연스럽게 반영합니다:
  - 韓国 美容クリニック
  - ソウル 江南（江南エリア）
  - プレミアム アンチエイジング施術
  - 最新医療機器
- 키워드 나열은 금지합니다.

용어 정확도 검증:
- 시술명, 기기명, 의학·피부 관련 용어는 등록된 용어집을 최우선으로 사용합니다.
- 용어집에 없는 경우에는 불확실한 추측 번역을 하지 말고 가장 중립적이고 설명적인 표현을 선택합니다.

출력 원칙:
- 한국어 설명을 추가로 늘어놓지 말고, 앱이 요구하는 JSON 구조에 맞게 번역 본문만 충실히 제공합니다.
- 원문의 문단, 볼드, 표, FAQ 구조를 최대한 유지합니다.
""".strip(),
    "대만(번체)": """
[번역 및 언어 규칙 - 엄격 준수]
- 모든 번역은 대만식 번체자(Traditional Chinese - Taiwan)로 진행합니다.
- 영어와 숫자 앞뒤에 공백을 넣지 않습니다.
- 대만 현지 관습에 맞는 전각 부호(，、。？！：；)를 사용합니다.
- 한국어의 '~의'는 대만 어법에 맞게 `的` 또는 문맥상 자연스러운 형태로 번역합니다.
- 포스팅 마지막에는 반드시 다음 문구를 유지합니다: *本內容遵循醫療法和保健福利部醫療廣告相關法規進行撰寫。*

[SEO 및 검색 최적화]
- 대만(Taiwan) 현지 검색 의도를 우선 반영합니다.
- 현지 사용자가 실제로 검색하는 시술 별명, 고민 키워드, 대만 커뮤니티 어휘를 자연스럽게 반영합니다.
- 원문의 구조, 굵게, 기울임꼴, 표, 링크 형식을 유지합니다.

[브랜드명 고정 번역]
- `세예의원`, `세예클리닉`, `世颜`, `世颜诊所`는 반드시 `SeyeClinic 世顏`으로 번역합니다.
- 브랜드명은 절대 다른 한자로 변형하지 않습니다.

[용어집 우선 원칙]
- 등록된 대만어 용어집을 최우선으로 사용합니다.
- 용어집에 없는 경우에만 기존 고정 용어 사전과 중립적 표현을 사용합니다.

[스타일]
- 신뢰감 있고 전문적인 메디컬 어조를 유지하되, 대만 SNS/검색에서 어색하지 않은 자연스러운 현지 표현을 사용합니다.
- 과장형 광고 카피보다 정보형 의료 콘텐츠 흐름을 우선합니다.
""".strip(),
    "중국(번체)": """
[번역 및 언어 규칙]
- 모든 번역은 번체 중국어로 진행합니다.
- 영어와 숫자 앞뒤의 불필요한 공백은 넣지 않습니다.
- 전각 문장 부호를 사용합니다.
- 원문의 의미를 유지하면서 의료 정보형 블로그 톤으로 번역합니다.

[용어와 브랜드]
- 등록된 용어집을 최우선으로 사용합니다.
- 브랜드명 `세예의원`, `세예클리닉`, `世颜`, `世颜诊所`는 `SeyeClinic 世顏`으로 통일합니다.

[SEO 및 구조]
- 검색 사용자가 읽기 쉬운 표현을 선택하되 과장 광고조 표현은 피합니다.
- 원문의 마크다운 구조, 표, 굵게, 리스트, FAQ 흐름을 유지합니다.
""".strip(),
}

DEFAULT_WEBSITE_PROMPTS: dict[str, str] = {
    "한국어": """
[목적]
나는 연계 피부과(세예의원)와 피부시술에 대한 블로그 콘텐츠를 만드는 콘텐츠 마케터야.
네이버 블로그에 발행해온 글을 재가공해서, 피부과 공식 홈페이지 내 웹 콘텐츠로 올리려고해.

내가 기존 블로그 글을 입력하면, 의료법 및 보건복지부 광고 가이드라인과 네이버 SEO/GEO 전략을 준수하면서
입력한 원본 블로그 내용을 홈페이지용 웹 콘텐츠로 변환해줘.
SEO/GEO 최적화·의료법 준수·시술 정보 전달·병원 전문성 반영이 핵심이야.

[내용 변환 전략]
1. 어뷰징 방지 (유사 문서 필터링 방지)
- 기존 블로그 글의 문장 순서, 단락 구조를 완전히 해체하고 재조합
- 핵심 의미와 정보는 유지하되, 모든 문장을 새롭게 작성하여 유사 문서로 인식되지 않도록 워싱
- 제목도 클릭을 유도하면서 SEO에 최적화된 문장으로 재구성

2. 톤 앤 매너 변경 (블로그 → 전문 웹사이트)
- 친근한 블로그 어투를 병원 공식 웹사이트에 어울리는 전문적이고 신뢰감 있는 어투로 전환
- 후기성 정보가 아니라 의료 정보로서의 전문성을 부각

3. 콘텐츠 구조 최적화
- 명확한 소제목(H2, H3)을 사용
- 소제목 앞에 번호 매기기 필수

4. 키워드 전략 수정
- 주력 키워드는 유지하되 정보성 세부 키워드를 자연스럽게 재배치
- 주력키워드는 동일하게 6회 반복하고, 띄어쓰기 없이 볼드체로 표시
- 특정 형태소가 과하게 반복되지 않도록 Paraphrasing

5. 첫/마지막 문장 고정
- 첫문장: “안녕하세요. 세상의 모든 예쁨, 세예의원입니다.”
- 마지막문장: “본 포스팅의 내용은 의료법과 보건복지부 의료광고 가이드라인을 준수하여 작성되었습니다.”

6. 의료법 제56조·제57조 준수
- 과장, 비교, 효과 보장, 전후 사진 유사 표현 금지

7. 제목/썸네일 문구/클릭 버튼 description
- 제목: SEO/GEO 전략에 맞춰 핵심 내용을 담아 재구성
- 썸네일 문구: 짧고 후킹되게 도출
- Slug: SEO 최적화된 영문 슬러그 도출
- Description: 검색 클릭을 유도하는 설명 문구 도출

[레퍼런스 활용]
- 원본 블로그 원고 파일('블로그 원본')과 웹용으로 변환한 파일('웹콘텐츠 변환')을 레퍼런스로 참고

[최종 점검]
다음 항목을 함께 점검:
1. 의료법 제56조, 57조 위반 표현 여부
2. 틀린 맞춤법 여부
3. 주력키워드 반복 횟수
4. SEO/GEO 최적화 전략 반영 여부
""".strip(),
    "영어": """
Role:
You rewrite Korean clinic content into SEO/GEO-friendly website copy for Google search.

Rules:
- Keep the meaning faithful to the source.
- Do not add medical claims, guarantees, or exaggerated marketing copy.
- Use a polished clinic website tone rather than a casual blog tone.
- Organize the output as a landing/detail page with clear headings and CTA.
""".strip(),
    "일본어": """
役割:
韓国語の原稿を、日本の検索ユーザー向けのクリニックWebサイト用SEO・GEO原稿に再構成します。

原則:
- 原文にない情報や誇張表現は追加しません。
- 医療広告に抵触する恐れのある断定表現を避けます。
- ブログよりも整ったWebサイト情報ページの文体を使用します。
- 見出し構造、FAQ、CTAを自然に整理します。
""".strip(),
    "대만(번체)": """
角色：
你要把原始韓文內容改寫成適合Google SEO／GEO的網站頁面文案。

原則：
- 不新增原文中沒有的療效保證或誇大敘述。
- 維持專業、可信的醫療資訊型語氣。
- 結構要更像網站介紹頁，而不是部落格日記式語氣。
- 以台灣繁體中文輸出。
""".strip(),
    "중국(번체)": """
角色：
你要把原始韓文內容改寫成適合Google SEO／GEO的網站頁面文案。

原則：
- 不新增原文中沒有的療效保證或誇大敘述。
- 維持專業、可信的醫療資訊型語氣。
- 結構偏網站介紹頁、FAQ頁、服務說明頁。
- 使用繁體中文輸出。
""".strip(),
    "중국(간체)": """
角色：
你要把原始韩文内容改写成适合Google SEO／GEO的网站页面文案。

原则：
- 不新增原文中没有的疗效保证或夸大表述。
- 保持专业、可信的信息型医疗语气。
- 输出结构更像网站介绍页、服务页和FAQ页。
- 使用简体中文输出。
""".strip(),
}

DEFAULT_TRANSLATION_GLOSSARIES: dict[str, dict[str, str]] = {
    "대만(번체)": {
        "세예의원": "SeyeClinic 世顏",
        "세예클리닉": "SeyeClinic 世顏",
        "世颜": "SeyeClinic 世顏",
        "世颜诊所": "SeyeClinic 世顏",
        "스킨부스터 유목민": "皮膚遊牧民族",
        "키클리닉": "官方認證皮膚科診所",
        "정품·정량": "正品·足量",
        "올타이트": "Alltite",
        "엘란쎄": "Ellansé少女針",
        "리쥬란": "麗珠蘭",
        "쥬베룩": "Juvelook喬雅露",
        "스킨보톡스": "肉毒水光",
        "더마샤인프로": "德瑪莎水光",
        "스킨부스터": "水光針",
        "셀르디엠": "CellREDM真皮膠原水光",
        "시술": "療程",
        "나이테주사": "年輪針",
        "간편예약상담": "快速諮詢預約",
        "간편 문의하기": "快速諮詢預約",
        "한국 클리닉": "韓國皮膚科診所",
        "한국 피부과": "韓國皮膚科診所",
        "엠보싱": "鼓包",
        "온다리프팅": "ONDA拉提",
        "울쎄라": "Ultherapy",
    },
    "중국(번체)": {
        "세예의원": "SeyeClinic 世顏",
        "세예클리닉": "SeyeClinic 世顏",
        "世颜": "SeyeClinic 世顏",
        "世颜诊所": "SeyeClinic 世顏",
        "스킨부스터 유목민": "皮膚遊牧民族",
        "키클리닉": "官方認證皮膚科診所",
        "정품·정량": "正品·足量",
        "올타이트": "Alltite",
        "엘란쎄": "Ellansé少女針",
        "리쥬란": "麗珠蘭",
        "쥬베룩": "Juvelook喬雅露",
        "스킨보톡스": "肉毒水光",
        "더마샤인프로": "德瑪莎水光",
        "나이테주사": "年輪針",
        "간편예약상담": "快速諮詢預約",
        "간편 문의하기": "快速諮詢預約",
        "한국 클리닉": "韓國皮膚科診所",
        "한국 피부과": "韓國皮膚科診所",
        "엠보싱": "鼓包",
        "온다리프팅": "ONDA拉提",
        "울쎄라": "Ultherapy",
        "스킨부스터": "水光針",
        "셀르디엠": "CellREDM真皮膠原水光",
        "시술": "療程",
    },
    "중국(간체)": {
        "세예의원": "SeyeClinic 世颜",
        "세예클리닉": "SeyeClinic 世颜",
        "世颜": "SeyeClinic 世颜",
        "世颜诊所": "SeyeClinic 世颜",
        "스킨부스터 유목민": "皮肤游牧民族",
        "키클리닉": "官方认证皮肤科诊所",
        "정품·정량": "正品·足量",
        "올타이트": "Alltite",
        "엘란쎄": "Ellansé少女针",
        "리쥬란": "丽珠兰",
        "쥬베룩": "Juvelook乔雅露",
        "스킨보톡스": "肉毒水光",
        "더마샤인프로": "德玛莎水光",
        "셀르디엠": "CellREDM真皮胶原水光",
        "시술": "疗程",
        "나이테주사": "年轮针",
        "간편예약상담": "快速咨询预约",
        "간편 문의하기": "快速咨询预约",
        "한국 클리닉": "韩国皮肤科诊所",
        "한국 피부과": "韩国皮肤科诊所",
        "엠보싱": "鼓包",
        "온다리프팅": "ONDA提升",
        "울쎄라": "Ultherapy",
        "스킨부스터": "水光针",
    },
    "일본어": {
        "세예의원": "セイェ医院",
        "온다리프팅": "ONDAリフティング",
        "울쎄라": "ウルセラ",
        "스킨부스터": "スキンブースター",
    },
    "영어": {
        "세예의원": "Seye Clinic",
        "온다리프팅": "ONDA lifting",
        "울쎄라": "Ultherapy",
        "스킨부스터": "skin booster",
    },
}


def _clean_text_cell(value: Any) -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text


def _load_japanese_glossary_seed() -> dict[str, str]:
    if not JAPANESE_GLOSSARY_SEED_PATH.exists():
        return {}

    try:
        df = pd.read_excel(JAPANESE_GLOSSARY_SEED_PATH, sheet_name="Glossary", header=None)
    except Exception:
        return {}

    source_col = 2
    target_col = 3
    header_row = None
    for row_idx in range(len(df)):
        row_values = [_clean_text_cell(cell) for cell in df.iloc[row_idx].tolist()]
        if "한국어 용어" in row_values:
            source_col = row_values.index("한국어 용어")
            for col_idx, cell in enumerate(row_values):
                if "일본어" in cell:
                    target_col = col_idx
                    header_row = row_idx
                    break
        if header_row is not None:
            break

    parsed: dict[str, str] = {}
    start_row = (header_row + 1) if header_row is not None else 0
    for row_idx in range(start_row, len(df)):
        source = _clean_text_cell(df.iat[row_idx, source_col]) if source_col < df.shape[1] else ""
        target = _clean_text_cell(df.iat[row_idx, target_col]) if target_col < df.shape[1] else ""
        if source and target:
            parsed[source] = target
    return parsed


def _load_chinese_glossary_seed(language_label: str) -> dict[str, str]:
    if language_label not in {"중국(간체)", "중국(번체)"}:
        return {}
    if not CHINESE_GLOSSARY_SEED_PATH.exists():
        return {}

    try:
        df = pd.read_csv(CHINESE_GLOSSARY_SEED_PATH, header=None, encoding="utf-8-sig")
    except Exception:
        try:
            df = pd.read_csv(CHINESE_GLOSSARY_SEED_PATH, header=None, encoding="cp949")
        except Exception:
            return {}

    source_col = None
    target_col = None
    header_row = None
    target_header = "중국어 (Simplified Chinese)" if language_label == "중국(간체)" else "중국어 번체 (Traditional Chinese)"

    for row_idx in range(len(df)):
        row_values = [_clean_text_cell(cell) for cell in df.iloc[row_idx].tolist()]
        if "한국어 용어" in row_values and target_header in row_values:
            source_col = row_values.index("한국어 용어")
            target_col = row_values.index(target_header)
            header_row = row_idx
            break

    if source_col is None or target_col is None:
        return {}

    parsed: dict[str, str] = {}
    for row_idx in range(header_row + 1, len(df)):
        source = _clean_text_cell(df.iat[row_idx, source_col]) if source_col < df.shape[1] else ""
        target = _clean_text_cell(df.iat[row_idx, target_col]) if target_col < df.shape[1] else ""
        if source and target:
            parsed[source] = target
    return parsed


def _default_translation_glossary(language_label: str) -> dict[str, str]:
    glossary = dict(DEFAULT_TRANSLATION_GLOSSARIES.get(language_label, {}))
    if language_label == "일본어":
        glossary.update(_load_japanese_glossary_seed())
        glossary["세예의원"] = "SEYEクリニック"
        glossary["세예클리닉"] = "セイェクリニック"
    elif language_label in {"중국(간체)", "중국(번체)"}:
        glossary.update(_load_chinese_glossary_seed(language_label))
    return glossary


def _default_website_prompt(language_label: str) -> str:
    base = f"""
역할:
당신은 병원 웹사이트 SEO/GEO 랜딩페이지 전문 카피라이터입니다.
입력된 원문을 {language_label} 웹사이트용 콘텐츠로 재가공합니다.

원칙:
- 원문에 없는 시술 효과, 비교 우위, 과장 표현을 새로 추가하지 않습니다.
- 의료 광고 리스크가 있는 표현은 피하고 정보형/상담형 문체를 유지합니다.
- 검색엔진과 지역 검색을 고려해 제목, 소제목, FAQ 흐름을 정리합니다.
- 웹사이트 랜딩페이지에 바로 붙여넣기 쉬운 구조로 작성합니다.
- 굵게, 표, 리스트, FAQ 구조는 가능한 유지합니다.

출력:
- 제목 후보 5개
- 추천 소제목 구조
- 본문 초안
- 체크리스트
""".strip()
    if language_label == "한국어":
        return base + "\n- 한국어 웹사이트 SEO/GEO 페이지 톤으로 작성합니다."
    if language_label == "영어":
        return base + "\n- English clinic landing-page tone, readable and conversion-aware, without sounding overpromotional."
    if language_label == "일본어":
        return base + "\n- 日本の美容クリニック公式サイトで使うような、端正で信頼感のある文体を 유지합니다."
    if language_label == "대만(번체)":
        return base + "\n- 대만 현지 웹사이트/검색 사용자를 고려한 번체 중국어 랜딩페이지 톤으로 작성합니다."
    if language_label == "중국(번체)":
        return base + "\n- 번체 중국어권 사용자가 읽기 쉬운 웹사이트 정보형 톤으로 작성합니다."
    if language_label == "중국(간체)":
        return base + "\n- 간체 중국어권 사용자를 위한 웹사이트 SEO/GEO용 정보형 톤으로 작성합니다."
    return base


def load_user_website_prompts(operator_name: str | None = None) -> dict[str, dict[str, Any]]:
    path = operator_website_prompt_path(operator_name)
    records: dict[str, dict[str, Any]] = {}
    if path.exists():
        try:
            records = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            records = {}
    changed = False
    for language_label in WEBSITE_CONTENT_LANGUAGES:
        if language_label not in records:
            records[language_label] = {
                "text": _default_website_prompt(language_label),
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "updated_by": "system_default",
            }
            changed = True
    if changed:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    return records


def save_user_website_prompt(language_label: str, text: str, operator_name: str) -> None:
    records = load_user_website_prompts(operator_name)
    records[language_label] = {
        "text": text.strip(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": operator_name,
    }
    path = operator_website_prompt_path(operator_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def delete_user_website_prompt(language_label: str, operator_name: str) -> None:
    records = load_user_website_prompts(operator_name)
    records[language_label] = {
        "text": _default_website_prompt(language_label),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": f"{operator_name} (reset)",
    }
    path = operator_website_prompt_path(operator_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def list_user_website_reference_files(language_label: str, operator_name: str | None = None) -> list[Path]:
    directory = operator_website_reference_dir(language_label, operator_name)
    if not directory.exists():
        return []
    return sorted([path for path in directory.iterdir() if path.is_file()])


def save_uploaded_website_reference_files(language_label: str, uploaded_files: list[Any], operator_name: str | None = None) -> None:
    directory = operator_website_reference_dir(language_label, operator_name)
    directory.mkdir(parents=True, exist_ok=True)
    for uploaded in uploaded_files:
        target = directory / uploaded.name
        target.write_bytes(uploaded.getbuffer())


def delete_user_website_reference_file(language_label: str, filename: str, operator_name: str | None = None) -> None:
    target = operator_website_reference_dir(language_label, operator_name) / filename
    if target.exists():
        target.unlink()


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "Referer": "https://search.naver.com/",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        }
    )
    return session


def setup_search_driver(user_agent: str | None = None):
    if not SELENIUM_AVAILABLE:
        raise RuntimeError("Selenium이 설치되지 않아 이 기능을 사용할 수 없습니다. (클라우드 환경)")
    chromedriver_path = os.getenv("CHROMEDRIVER_BINARY", "").strip()
    service = Service(executable_path=chromedriver_path) if chromedriver_path else None
    remote_debugger_addr = os.getenv("REMOTE_CHROME_DEBUGGER_ADDR", "").strip()
    if remote_debugger_addr:
        try:
            pages = json.load(urlopen(f"http://{remote_debugger_addr}/json/list", timeout=3))
            if not pages:
                req = Request(f"http://{remote_debugger_addr}/json/new?about:blank", method="PUT")
                urlopen(req, timeout=3).read()
                time.sleep(0.5)
        except Exception:
            pass
        options = webdriver.ChromeOptions()
        options.page_load_strategy = "none"
        options.add_experimental_option("debuggerAddress", remote_debugger_addr)
        chromium_path = os.getenv("CHROMIUM_BINARY", "/opt/homebrew/bin/chromium")
        if Path(chromium_path).exists():
            options.binary_location = chromium_path
        driver = webdriver.Chrome(options=options, service=service)
        driver.set_page_load_timeout(int(os.getenv("PAGE_LOAD_TIMEOUT", "12")))
        driver.set_script_timeout(int(os.getenv("SCRIPT_TIMEOUT", "8")))
        return driver

    options = webdriver.ChromeOptions()
    options.page_load_strategy = "eager"
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--lang=ko-KR")
    options.add_argument("--window-size=1440,2200")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(f"--user-agent={user_agent or USER_AGENT}")
    chromium_path = os.getenv("CHROMIUM_BINARY", "/opt/homebrew/bin/chromium")
    if Path(chromium_path).exists():
        options.binary_location = chromium_path
    driver = webdriver.Chrome(options=options, service=service)
    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                    Object.defineProperty(navigator, 'platform', {get: () => 'MacIntel'});
                    Object.defineProperty(navigator, 'language', {get: () => 'ko-KR'});
                    Object.defineProperty(navigator, 'languages', {get: () => ['ko-KR', 'ko', 'en-US', 'en']});
                """
            },
        )
    except Exception:
        pass
    driver.set_page_load_timeout(int(os.getenv("PAGE_LOAD_TIMEOUT", "20")))
    driver.set_script_timeout(int(os.getenv("SCRIPT_TIMEOUT", "8")))
    return driver


def safe_int(value: Any) -> int:
    if value in (None, "", "< 10"):
        return 0
    if isinstance(value, (int, float)):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0
    text = str(value).replace(",", "").strip()
    if text.isdigit():
        return int(text)
    try:
        return int(float(text))
    except ValueError:
        return 0


def normalize_seed_keywords(seed_keywords_text: str) -> list[str]:
    return [keyword.strip() for keyword in seed_keywords_text.split(",") if keyword.strip()]


def first_keyword_from_text(text: str) -> str:
    keywords = normalize_seed_keywords(text)
    return keywords[0] if keywords else str(text).strip()


def normalize_keyword_text(text: str) -> str:
    return re.sub(r"\s+", "", str(text).strip().lower())


def keyword_core_tokens(text: str) -> list[str]:
    return re.findall(r"[가-힣A-Za-z]{2,}", text)


def query_focus_tokens(text: str) -> list[str]:
    generic = {
        "효과",
        "가격",
        "후기",
        "통증",
        "주기",
        "유지기간",
        "부작용",
        "관리",
        "비용",
        "원리",
        "전후",
    }
    tokens = keyword_core_tokens(text)
    expanded: list[str] = []
    for token in tokens:
        if token not in expanded:
            expanded.append(token)
        for suffix in generic:
            if token.endswith(suffix) and len(token) > len(suffix):
                stem = token[: -len(suffix)]
                if len(stem) >= 2 and stem not in expanded:
                    expanded.append(stem)
                if suffix not in expanded:
                    expanded.append(suffix)
            if token.startswith(suffix) and len(token) > len(suffix):
                stem = token[len(suffix) :]
                if len(stem) >= 2 and stem not in expanded:
                    expanded.append(stem)
                if suffix not in expanded:
                    expanded.append(suffix)
    focused = [token for token in expanded if token not in generic]
    return focused + [token for token in expanded if token in generic]


def resolve_workbook_path() -> str | None:
    for candidate in WORKBOOK_CANDIDATE_PATHS:
        if Path(candidate).exists():
            return candidate
    return None


def find_matching_sheet_name(seed_keyword: str, sheet_names: list[str]) -> str | None:
    normalized_seed = normalize_keyword_text(seed_keyword)
    exact_matches = [name for name in sheet_names if normalize_keyword_text(name) == normalized_seed]
    if exact_matches:
        return exact_matches[0]

    containing_matches = [name for name in sheet_names if normalized_seed in normalize_keyword_text(name)]
    if containing_matches:
        return containing_matches[0]

    # Allow brand/procedure shorthand matching such as
    # "올타이트리프팅" -> "올타이트", "티타늄리프팅" -> "티타늄".
    trimmed_seed = normalized_seed
    for suffix in ("리프팅", "필러", "부스터", "레이저", "토닝", "셀르디엠"):
        if trimmed_seed.endswith(suffix) and len(trimmed_seed) > len(suffix):
            trimmed_seed = trimmed_seed[: -len(suffix)]
            break

    if trimmed_seed != normalized_seed:
        trimmed_exact = [name for name in sheet_names if normalize_keyword_text(name) == trimmed_seed]
        if trimmed_exact:
            return trimmed_exact[0]

        trimmed_containing = [name for name in sheet_names if trimmed_seed in normalize_keyword_text(name)]
        if trimmed_containing:
            return trimmed_containing[0]

    return None


@st.cache_data(show_spinner=False, ttl=60 * 60)
def load_keyword_workbook_reference() -> dict[str, Any]:
    workbook_path = resolve_workbook_path()
    if not workbook_path:
        return {"path": None, "sheets": {}, "sheet_names": []}

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_payload: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=3, max_row=80, values_only=True):
            if not row or len(row) < 6:
                continue
            keyword = row[2]
            if not keyword:
                continue
            latest_search = None
            latest_blog = None
            latest_saturation = None
            metrics = row[3:]
            for idx in range(0, len(metrics), 3):
                triple = metrics[idx : idx + 3]
                if len(triple) < 3:
                    continue
                search_volume, blog_count, saturation = triple
                if search_volume is not None:
                    latest_search = search_volume
                    latest_blog = blog_count
                    latest_saturation = saturation
            rows.append(
                {
                    "sheet_name": ws.title,
                    "procedure": row[0],
                    "classification": row[1] or "미분류",
                    "keyword": str(keyword).strip(),
                }
            )
        sheet_payload[ws.title] = rows
    return {"path": workbook_path, "sheets": sheet_payload, "sheet_names": list(sheet_payload.keys())}


def extract_modifier(seed_keyword: str, keyword: str) -> str:
    normalized_seed = seed_keyword.strip()
    text = keyword.strip()
    modifier = text.replace(normalized_seed, "")
    return modifier.strip()


def workbook_priority(record: dict[str, Any]) -> tuple[int, int]:
    classification = str(record.get("classification") or "")
    keyword = str(record.get("keyword") or "")
    point = 0
    if classification == "점령":
        point += 5
    preferred_terms = ("가격", "효과", "유지기간", "주기", "후기", "통증", "부작용")
    if any(term in keyword for term in preferred_terms):
        point += 3
    if "병원" in keyword or "피부과" in keyword or "잘하는곳" in keyword:
        point -= 5
    return (point, -len(keyword))


def source_priority_multiplier(source_type: str, challenge_level: str) -> float:
    matrix = {
        "안전": {"workbook": 1.35, "api": 0.9, "fallback": 0.75},
        "균형": {"workbook": 1.15, "api": 1.0, "fallback": 0.8},
        "도전": {"workbook": 1.0, "api": 1.18, "fallback": 0.85},
    }
    return matrix.get(challenge_level, matrix["균형"]).get(source_type, 1.0)


def should_keep_candidate(
    existing: SearchAdKeyword | None,
    candidate: SearchAdKeyword,
    challenge_level: str,
) -> bool:
    if existing is None:
        return True

    source_rank = {
        "안전": {"workbook": 3, "api": 2, "fallback": 1},
        "균형": {"workbook": 3, "api": 3, "fallback": 1},
        "도전": {"api": 3, "workbook": 2, "fallback": 1},
    }
    rank_table = source_rank.get(challenge_level, source_rank["균형"])
    existing_rank = rank_table.get(existing.source_type, 0)
    candidate_rank = rank_table.get(candidate.source_type, 0)
    if candidate_rank != existing_rank:
        return candidate_rank > existing_rank
    return candidate.opportunity_score > existing.opportunity_score


def candidate_sort_key(candidate: SearchAdKeyword, challenge_level: str) -> tuple[int, float, int]:
    source_rank = {
        "안전": {"workbook": 3, "api": 2, "fallback": 1},
        "균형": {"workbook": 3, "api": 3, "fallback": 1},
        "도전": {"api": 3, "workbook": 2, "fallback": 1},
    }
    rank_table = source_rank.get(challenge_level, source_rank["균형"])
    return (
        rank_table.get(candidate.source_type, 0),
        candidate.opportunity_score,
        candidate.monthly_total,
    )


def infer_medical_theme_tokens(seed_keywords: tuple[str, ...]) -> set[str]:
    tokens: set[str] = set()
    for keyword in seed_keywords:
        for token in keyword_core_tokens(keyword):
            tokens.add(token)
            if token.endswith("리프팅"):
                tokens.add("리프팅")
            if token.endswith("필러"):
                tokens.add("필러")
            if token.endswith("부스터"):
                tokens.add("부스터")
    return tokens


def infer_broad_seed_terms(seed_keywords: tuple[str, ...]) -> set[str]:
    broad_terms: set[str] = set()
    for keyword in seed_keywords:
        broad_terms.update(keyword_core_tokens(keyword))
    broad_terms.update({"리프팅", "필러", "부스터"})
    return broad_terms


def is_location_keyword(keyword: str) -> bool:
    district_suffixes = ("동", "읍", "면", "리", "가", "구", "시", "군")
    common_locations = {
        "서울",
        "부산",
        "대구",
        "인천",
        "광주",
        "대전",
        "울산",
        "세종",
        "제주",
        "강남",
        "강남역",
        "신논현",
        "해운대",
        "압구정",
        "잠실",
        "홍대",
        "판교",
        "분당",
        "수원",
        "일산",
        "천안",
        "안산",
        "송도",
        "평촌",
        "부천",
        "동탄",
        "마곡",
        "목동",
        "신사동",
        "청담",
    }
    if any(token in keyword for token in common_locations):
        return True
    tokens = keyword_core_tokens(keyword)
    return any(token.endswith(district_suffixes) and len(token) >= 2 for token in tokens)


def is_too_broad_keyword(keyword: str, seed_keyword: str, broad_terms: set[str]) -> bool:
    normalized = keyword.strip()
    if normalized == seed_keyword.strip():
        return True
    if normalized in broad_terms:
        return True
    if len(normalized) <= 3 and normalized.endswith(("리프팅", "필러", "부스터")):
        return True
    return False


def fallback_keyword_variants(seed_keyword: str) -> list[str]:
    return [
        f"{seed_keyword}효과",
        f"{seed_keyword}유지력",
        f"{seed_keyword}통증",
        f"{seed_keyword}붓기",
        f"{seed_keyword}멍",
        f"{seed_keyword}후관리",
        f"{seed_keyword}추천대상",
        f"{seed_keyword}주기",
        f"{seed_keyword}간격",
    ]


def is_relevant_keyword(
    keyword: str,
    theme_tokens: set[str],
    seed_keyword: str,
    broad_terms: set[str],
    allow_workbook_patterns: bool = False,
) -> bool:
    normalized = keyword.strip()
    if len(normalized) < 2:
        return False

    blocked_contains = {
        "의원",
        "병원",
        "클리닉",
        "피부과",
        "성형외과",
        "한의원",
        "치과",
        "내돈내산",
        "이벤트",
        "추천",
        "맛집",
        "호텔",
        "유명한곳",
        "잘하는곳",
        "붓기",
        "멍",
    }
    if not allow_workbook_patterns:
        blocked_contains.update({"후기", "가격", "비용"})
    if any(token in normalized for token in blocked_contains):
        return False
    if is_location_keyword(normalized):
        return False
    if is_too_broad_keyword(normalized, seed_keyword, broad_terms):
        return False

    if any(token in normalized for token in theme_tokens if len(token) >= 2):
        return True

    allowed_suffixes = ("탄력", "주름", "안티에이징")
    if any(normalized.endswith(suffix) for suffix in allowed_suffixes):
        return True

    allowed_contains = {
        "목주름",
        "이중턱",
        "심부볼",
        "볼처짐",
        "탄력",
        "처짐",
        "잔주름",
        "콜라겐",
        "효과",
        "유지력",
        "통증",
        "멍",
        "붓기",
        "후관리",
    }
    if any(token in normalized for token in allowed_contains):
        return True

    return False


def estimate_document_count_from_competition(monthly_total: int, competition_index: str) -> int:
    estimated_by_competition = {
        "낮음": max(10, int(monthly_total * 0.35)),
        "중간": max(30, int(monthly_total * 0.9)),
        "높음": max(50, int(monthly_total * 1.8)),
        "미확인": max(20, int(monthly_total * 0.8)),
    }
    return estimated_by_competition.get(
        competition_index,
        max(20, int(monthly_total * 0.8)),
    )


def enrich_candidate(row: SearchAdKeyword) -> SearchAdKeyword:
    document_count = fetch_blog_document_count(row.keyword)
    if document_count <= 0:
        document_count = estimate_document_count_from_competition(row.monthly_total, row.competition_index)
    adjusted_saturation = document_count / max(row.monthly_total, 1)
    opportunity_score = (
        math.log1p(row.monthly_total)
        * (1 / (1 + math.log1p(max(document_count, 1)) / 6))
        * to_competition_weight(row.competition_index)
    )
    return SearchAdKeyword(
        keyword=row.keyword,
        monthly_pc=row.monthly_pc,
        monthly_mobile=row.monthly_mobile,
        monthly_total=row.monthly_total,
        monthly_docs=document_count,
        competition_index=row.competition_index,
        source_seed=row.source_seed,
        saturation=adjusted_saturation,
        opportunity_score=opportunity_score,
        recommendation_reason=build_recommendation_reason(
            monthly_total=row.monthly_total,
            saturation=adjusted_saturation,
            competition_index=row.competition_index,
        ),
        is_fallback=row.is_fallback,
        source_type=row.source_type,
        keyword_classification=row.keyword_classification,
    )


def to_competition_weight(raw: str) -> float:
    mapping = {"낮음": 1.15, "중간": 1.0, "높음": 0.85}
    return mapping.get(str(raw).strip(), 1.0)


def competition_label(raw: str) -> str:
    return str(raw).strip() or "미확인"


def saturation_band(value: float) -> str:
    if value <= 0.8:
        return "여유"
    if value <= 3:
        return "적정"
    return "과열"


def recommendation_grade(saturation: float, competition_index: str, is_fallback: bool) -> str:
    if is_fallback:
        return "추정 후보"
    if competition_index == "높음" and saturation >= 1.5:
        return "주의"
    if competition_index == "중간" and saturation >= 1.2:
        return "테스트 추천"
    if competition_index in {"낮음", "중간"} and saturation <= 1.0:
        return "추천"
    if saturation <= 0.8:
        return "추천"
    if saturation <= 1.8:
        return "테스트 추천"
    return "주의"


def build_recommendation_reason(monthly_total: int, saturation: float, competition_index: str) -> str:
    reasons: list[str] = []
    if monthly_total >= 3000:
        reasons.append("검색량이 충분함")
    elif monthly_total >= 800:
        reasons.append("검색량이 안정적임")
    else:
        reasons.append("롱테일 테스트에 적합함")

    band = saturation_band(saturation)
    if band == "여유":
        reasons.append("문서 경쟁이 낮음")
    elif band == "적정":
        reasons.append("포화도가 과하지 않음")
    else:
        reasons.append("경쟁이 높아 제목 설계가 중요함")

    if competition_label(competition_index) == "낮음":
        reasons.append("광고 경쟁도도 낮은 편")
    return " / ".join(reasons)


def generate_searchad_signature(timestamp: str, method: str, uri: str, secret_key: str) -> str:
    message = f"{timestamp}.{method}.{uri}"
    digest = hmac.new(secret_key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).digest()
    return base64.b64encode(digest).decode("utf-8")


def request_searchad_related_keywords(session: requests.Session, hint_keywords: list[str]) -> list[dict[str, Any]]:
    if not (NAVER_SEARCHAD_ACCESS_LICENSE and NAVER_SEARCHAD_SECRET_KEY and NAVER_SEARCHAD_CUSTOMER_ID):
        raise RuntimeError(
            "네이버 검색광고 API 인증 정보가 없습니다. "
            ".env 에 NAVER_SEARCHAD_ACCESS_LICENSE, NAVER_SEARCHAD_SECRET_KEY, "
            "NAVER_SEARCHAD_CUSTOMER_ID 를 설정해 주세요."
        )

    uri = "/keywordstool"
    method = "GET"
    timestamp = str(int(time.time() * 1000))
    signature = generate_searchad_signature(timestamp, method, uri, NAVER_SEARCHAD_SECRET_KEY)
    last_error: requests.HTTPError | None = None
    for attempt in range(3):
        response = session.get(
            f"{NAVER_SEARCHAD_BASE_URL.rstrip('/')}{uri}",
            params={"hintKeywords": ",".join(hint_keywords), "showDetail": 1},
            headers={
                "X-Timestamp": timestamp,
                "X-API-KEY": NAVER_SEARCHAD_ACCESS_LICENSE,
                "X-Customer": NAVER_SEARCHAD_CUSTOMER_ID,
                "X-Signature": signature,
            },
            timeout=REQUEST_TIMEOUT,
        )
        if response.status_code != 429:
            response.raise_for_status()
            data = response.json()
            return data.get("keywordList", [])
        time.sleep(1.2 * (attempt + 1))
        last_error = requests.HTTPError(
            "네이버 검색광고 API 호출이 잠시 제한되었습니다. 잠시 후 다시 시도해 주세요.",
            response=response,
        )

    if last_error:
        raise last_error
    raise RuntimeError("네이버 검색광고 API 호출에 실패했습니다.")


@st.cache_data(show_spinner=False, ttl=60 * 60)
def fetch_searchad_metrics_for_keywords(seed_keyword: str) -> dict[str, dict[str, Any]]:
    session = make_session()
    rows = request_searchad_related_keywords(session, [seed_keyword])
    metrics: dict[str, dict[str, Any]] = {}
    for row in rows:
        keyword = str(row.get("relKeyword", "")).strip()
        if not keyword:
            continue
        monthly_pc = safe_int(row.get("monthlyPcQcCnt"))
        monthly_mobile = safe_int(row.get("monthlyMobileQcCnt"))
        metrics[keyword] = {
            "monthly_pc": monthly_pc,
            "monthly_mobile": monthly_mobile,
            "monthly_total": monthly_pc + monthly_mobile,
            "competition_index": competition_label(row.get("compIdx", "")),
        }
    return metrics


@st.cache_data(show_spinner=False, ttl=60 * 60)
def fetch_exact_searchad_metric(keyword: str) -> dict[str, Any] | None:
    session = make_session()
    rows = request_searchad_related_keywords(session, [keyword])
    normalized_target = normalize_keyword_text(keyword)
    for row in rows:
        rel_keyword = str(row.get("relKeyword", "")).strip()
        if normalize_keyword_text(rel_keyword) != normalized_target:
            continue
        monthly_pc = safe_int(row.get("monthlyPcQcCnt"))
        monthly_mobile = safe_int(row.get("monthlyMobileQcCnt"))
        return {
            "monthly_pc": monthly_pc,
            "monthly_mobile": monthly_mobile,
            "monthly_total": monthly_pc + monthly_mobile,
            "competition_index": competition_label(row.get("compIdx", "")),
        }
    return None


def fallback_searchad_metric(keyword: str) -> dict[str, Any] | None:
    core_keyword = first_keyword_from_text(keyword)
    tokens = query_focus_tokens(core_keyword)
    if not tokens:
        tokens = keyword_core_tokens(core_keyword)
    candidates = [core_keyword] + tokens
    seen: set[str] = set()
    for candidate in candidates:
        normalized = normalize_keyword_text(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        try:
            metric = fetch_exact_searchad_metric(candidate)
            if metric and metric["monthly_total"] > 0:
                return metric
        except Exception:
            continue
    return None


@st.cache_data(show_spinner=False, ttl=60 * 60 * 4)
def fetch_blog_document_count(keyword: str) -> int:
    session = make_session()
    response = session.get(
        "https://search.naver.com/search.naver",
        params={"where": "view", "query": keyword},
        timeout=REQUEST_TIMEOUT,
    )
    if response.status_code == 403:
        return 0
    response.raise_for_status()
    html = response.text
    patterns = [
        r'{"total":"?([\d,]+)"?',
        r'"title":"VIEW".{0,300}"total":"?([\d,]+)"?',
        r'검색결과\s*약\s*([\d,]+)건',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.DOTALL)
        if match:
            return safe_int(match.group(1))
    return 0


@st.cache_data(show_spinner=False, ttl=60 * 60 * 12)
def fetch_datalab_monthly_trend(keyword: str, months: int = 12) -> list[dict[str, Any]]:
    if not (NAVER_DATALAB_CLIENT_ID and NAVER_DATALAB_CLIENT_SECRET):
        return []

    today = datetime.now()
    start_month_index = today.month - months + 1
    start_year = today.year
    while start_month_index <= 0:
        start_month_index += 12
        start_year -= 1
    start_date = datetime(start_year, start_month_index, 1).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    response = requests.post(
        "https://openapi.naver.com/v1/datalab/search",
        headers={
            "X-Naver-Client-Id": NAVER_DATALAB_CLIENT_ID,
            "X-Naver-Client-Secret": NAVER_DATALAB_CLIENT_SECRET,
            "Content-Type": "application/json",
        },
        json={
            "startDate": start_date,
            "endDate": end_date,
            "timeUnit": "month",
            "keywordGroups": [{"groupName": keyword, "keywords": [keyword]}],
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    payload = response.json()
    results = payload.get("results", [])
    if not results:
        return []
    rows = []
    for item in results[0].get("data", []):
        period = str(item.get("period", ""))
        ratio = float(item.get("ratio", 0.0))
        month_label = period[5:7].lstrip("0") + "월" if len(period) >= 7 else period
        rows.append({"period": period, "month": month_label, "ratio": ratio})
    return rows


def describe_channel_dominance(channel_counts: dict[str, int]) -> str:
    if not channel_counts:
        return "메인 1면 채널 데이터 없음"
    channel_map = {"blog": "블로그", "cafe": "카페", "kin": "지식인", "site": "웹사이트", "other": "기타"}
    dominant_key, dominant_count = max(channel_counts.items(), key=lambda item: item[1])
    label = channel_map.get(dominant_key, dominant_key)
    if dominant_key == "blog":
        return f"블로그 우세 ({dominant_count}건)"
    return f"{label} 우세 ({dominant_count}건)"


def infer_writing_format(posts: list[dict[str, Any]], top_terms: list[tuple[str, int]], keyword: str) -> str:
    titles = " ".join(post.get("title", "") for post in posts)
    normalized_titles = normalize_keyword_text(titles)
    if any(token in normalized_titles for token in ["비교", "vs", "차이"]):
        return "비교형"
    if any(token in normalized_titles for token in ["후기", "전후", "솔직"]):
        return "후기형"
    if any(term in dict(top_terms) for term in ["통증", "효과", "유지기간", "주기"]):
        return "정보형"
    if "효과" in keyword or "가격" in keyword:
        return "정보형"
    return "설명형"


def classify_opportunity(
    competition_index: str,
    saturation: float,
    channel_counts: dict[str, int],
    blog_post_count: int,
) -> tuple[str, str]:
    dominant = max(channel_counts.items(), key=lambda item: item[1])[0] if channel_counts else "other"
    if blog_post_count == 0:
        return "블로그 비우세", "메인 1면에 블로그보다 다른 채널이 강합니다. 블로그보다 카페/지식인 전략을 먼저 검토하세요."
    if dominant == "blog" and competition_index != "높음" and saturation <= 1.2:
        return "진입 가능성 높음", "메인 1면에서 블로그 노출이 확인되고 포화도도 과열 수준은 아닙니다."
    if saturation <= 1.8:
        return "테스트 추천", "상위 블로그는 존재하지만 제목 설계와 발행 완성도가 중요합니다."
    return "경쟁 강함", "상위 블로그 경쟁이 강한 편이라 롱테일 조합이나 다른 채널 병행이 필요합니다."


def build_keyword_diagnosis(summary: dict[str, Any]) -> KeywordDiagnosis:
    keyword = summary["keyword"]
    metric_source_label = "SearchAd 실측"
    try:
        metric = fetch_exact_searchad_metric(keyword) or {
            "monthly_pc": 0,
            "monthly_mobile": 0,
            "monthly_total": 0,
            "competition_index": "미확인",
        }
    except Exception:
        metric = {
            "monthly_pc": 0,
            "monthly_mobile": 0,
            "monthly_total": 0,
            "competition_index": "미확인",
        }
    if metric["monthly_total"] <= 0:
        fallback_metric = fallback_searchad_metric(keyword)
        if fallback_metric:
            metric = fallback_metric
            metric_source_label = "SearchAd 보정 실측"
    monthly_pc = metric["monthly_pc"]
    monthly_mobile = metric["monthly_mobile"]
    monthly_total = metric["monthly_total"]
    competition_index = metric["competition_index"]
    estimated_blog_docs = fetch_blog_document_count(keyword)
    if estimated_blog_docs <= 0:
        estimated_blog_docs = estimate_document_count_from_competition(monthly_total, competition_index)
    estimated_saturation = estimated_blog_docs / max(monthly_total, 1)

    trend_rows = fetch_datalab_monthly_trend(keyword, months=12)
    expected_month_total: int | None = None
    expected_month_to_date: int | None = None
    trend_peak_month: str | None = None
    trend_change_label = "DataLab 미연동"
    if trend_rows:
        ratios = [row["ratio"] for row in trend_rows if row["ratio"] > 0]
        if ratios:
            average_ratio = sum(ratios) / len(ratios)
            current_ratio = trend_rows[-1]["ratio"]
            if monthly_total > 0:
                seasonality_factor = current_ratio / average_ratio if average_ratio else 1.0
                expected_month_total = max(0, int(round(monthly_total * seasonality_factor)))
                today = datetime.now()
                month_last_day = calendar.monthrange(today.year, today.month)[1]
                progress = today.day / month_last_day
                expected_month_to_date = max(0, int(round(expected_month_total * progress)))
            peak_row = max(trend_rows, key=lambda row: row["ratio"])
            trend_peak_month = peak_row["month"]
            prev_ratio = trend_rows[-2]["ratio"] if len(trend_rows) >= 2 else current_ratio
            if current_ratio >= prev_ratio * 1.1:
                trend_change_label = "상승세"
            elif current_ratio <= prev_ratio * 0.9:
                trend_change_label = "하락세"
            else:
                trend_change_label = "보합세"

    channel_counts = summary.get("channel_counts", {})
    channel_dominance = describe_channel_dominance(channel_counts)
    blog_presence = "메인 1면 블로그 확인" if summary.get("post_count", 0) > 0 else "메인 1면 블로그 없음"
    opportunity_label, content_recommendation = classify_opportunity(
        competition_index=competition_index,
        saturation=estimated_saturation,
        channel_counts=channel_counts,
        blog_post_count=summary.get("post_count", 0),
    )
    writing_format = infer_writing_format(summary.get("posts", []), summary.get("top_terms", []), keyword)

    return KeywordDiagnosis(
        keyword=keyword,
        monthly_pc=monthly_pc,
        monthly_mobile=monthly_mobile,
        monthly_total=monthly_total,
        competition_index=competition_index,
        estimated_blog_docs=estimated_blog_docs,
        estimated_saturation=estimated_saturation,
        channel_dominance=channel_dominance,
        blog_presence=blog_presence,
        opportunity_label=opportunity_label,
        content_recommendation=content_recommendation,
        writing_format=writing_format,
        expected_month_total=expected_month_total,
        expected_month_to_date=expected_month_to_date,
        trend_peak_month=trend_peak_month,
        trend_change_label=trend_change_label,
        trend_rows=trend_rows,
    )


@st.cache_data(show_spinner=False, ttl=60 * 30)
def build_keyword_candidates(
    seed_keywords: tuple[str, ...],
    min_search_volume: int,
    max_search_volume: int | None,
    max_keyword_count: int,
    challenge_level: str,
) -> dict[str, list[dict[str, Any]]]:
    session = make_session()
    results_by_seed: dict[str, list[dict[str, Any]]] = {}
    broad_terms = infer_broad_seed_terms(seed_keywords)
    workbook_reference = load_keyword_workbook_reference()

    for seed_keyword in seed_keywords:
        matching_sheet = find_matching_sheet_name(seed_keyword, workbook_reference["sheet_names"])
        theme_seed_sources = [seed_keyword]
        if matching_sheet and matching_sheet != seed_keyword:
            theme_seed_sources.append(matching_sheet)
        theme_tokens = infer_medical_theme_tokens(tuple(theme_seed_sources))
        unique_rows: dict[str, SearchAdKeyword] = {}
        preferred_modifiers: set[str] = set()
        sheet_rows = workbook_reference["sheets"].get(matching_sheet, []) if matching_sheet else []
        sheet_rows = sorted(sheet_rows, key=workbook_priority, reverse=True)
        api_metrics = fetch_searchad_metrics_for_keywords(seed_keyword)
        if matching_sheet and matching_sheet != seed_keyword:
            sheet_metrics = fetch_searchad_metrics_for_keywords(matching_sheet)
            for keyword, metric in sheet_metrics.items():
                api_metrics.setdefault(keyword, metric)
        workbook_seeded_count = 0
        exact_lookup_budget = 4 if challenge_level == "안전" else 2

        for record in sheet_rows:
            keyword = record["keyword"]
            if not is_relevant_keyword(
                keyword,
                theme_tokens,
                seed_keyword,
                broad_terms,
                allow_workbook_patterns=True,
            ):
                continue
            modifier = extract_modifier(seed_keyword, keyword)
            if modifier:
                preferred_modifiers.add(modifier)
            metric = api_metrics.get(keyword)
            if metric is None and exact_lookup_budget > 0:
                metric = fetch_exact_searchad_metric(keyword)
                exact_lookup_budget -= 1
            if not metric:
                continue
            if metric["monthly_total"] < min_search_volume:
                continue
            if max_search_volume is not None and metric["monthly_total"] > max_search_volume:
                continue
            unique_rows[keyword] = SearchAdKeyword(
                keyword=keyword,
                monthly_pc=metric["monthly_pc"],
                monthly_mobile=metric["monthly_mobile"],
                monthly_total=metric["monthly_total"],
                monthly_docs=0,
                competition_index=metric["competition_index"],
                source_seed=seed_keyword,
                saturation=0.0,
                opportunity_score=math.log1p(max(metric["monthly_total"], 1))
                * 1.1
                * source_priority_multiplier("workbook", challenge_level),
                recommendation_reason=f"내부 시트 {record['classification']} 키워드",
                is_fallback=False,
                source_type="workbook",
                keyword_classification=str(record["classification"] or "미분류"),
            )
            workbook_seeded_count += 1

        should_fetch_api = not (challenge_level == "안전" and workbook_seeded_count >= max_keyword_count)
        if should_fetch_api:
            for keyword, metric in api_metrics.items():
                if not keyword:
                    continue
                if not is_relevant_keyword(keyword, theme_tokens, seed_keyword, broad_terms):
                    continue
                modifier = extract_modifier(seed_keyword, keyword)
                if challenge_level != "도전" and preferred_modifiers and modifier and modifier not in preferred_modifiers:
                    allowed_generic_modifiers = {"효과", "가격", "주기", "통증", "후기", "유지기간", "붓기", "부작용", "관리", "주의사항", "원리", "전후"}
                    if modifier not in allowed_generic_modifiers:
                        continue

                monthly_pc = metric["monthly_pc"]
                monthly_mobile = metric["monthly_mobile"]
                monthly_total = metric["monthly_total"]
                if monthly_total < min_search_volume:
                    continue
                if max_search_volume is not None and monthly_total > max_search_volume:
                    continue

                competition_index = metric["competition_index"]
                candidate = SearchAdKeyword(
                    keyword=keyword,
                    monthly_pc=monthly_pc,
                    monthly_mobile=monthly_mobile,
                    monthly_total=monthly_total,
                    monthly_docs=0,
                    competition_index=competition_index,
                    source_seed=seed_keyword,
                    saturation=0.0,
                    opportunity_score=(
                        math.log1p(monthly_total)
                        * to_competition_weight(competition_index)
                        * source_priority_multiplier("api", challenge_level)
                    ),
                    recommendation_reason="",
                    is_fallback=False,
                    source_type="api",
                    keyword_classification="API 추천",
                )
                existing = unique_rows.get(keyword)
                if should_keep_candidate(existing, candidate, challenge_level):
                    unique_rows[keyword] = candidate

        rough_candidates = list(unique_rows.values())
        rough_candidates.sort(key=lambda item: candidate_sort_key(item, challenge_level), reverse=True)

        enriched: list[SearchAdKeyword] = []
        for row in rough_candidates[: max(max_keyword_count * 3, 12)]:
            enriched.append(enrich_candidate(row))

        fallback_rows: list[SearchAdKeyword] = []
        existing_keywords = {item.keyword for item in enriched}
        for idx, fallback_keyword in enumerate(fallback_keyword_variants(seed_keyword), start=1):
            if fallback_keyword in existing_keywords:
                continue
            fallback_rows.append(
                enrich_candidate(
                    SearchAdKeyword(
                        keyword=fallback_keyword,
                        monthly_pc=max(0, 30 - idx * 2),
                        monthly_mobile=max(0, 120 - idx * 5),
                        monthly_total=max(80, 220 - idx * 10),
                        monthly_docs=0,
                        competition_index="미확인",
                        source_seed=seed_keyword,
                        saturation=0.0,
                        opportunity_score=0.0,
                        recommendation_reason="",
                        is_fallback=True,
                        source_type="fallback",
                        keyword_classification="보조 추천",
                    )
                )
            )
        if len(enriched) < max_keyword_count:
            enriched.extend(fallback_rows[: max_keyword_count - len(enriched)])

        enriched.sort(key=lambda item: candidate_sort_key(item, challenge_level), reverse=True)
        results_by_seed[seed_keyword] = [item.__dict__ for item in enriched[:max_keyword_count]]

    return results_by_seed


def normalize_naver_url(url: str) -> str:
    if not url:
        return url
    if url.startswith("//"):
        url = f"https:{url}"

    cleaned = html.unescape(unquote(url)).replace("\\u002F", "/").replace("\\/", "/")
    cleaned = cleaned.split("&quot;")[0].split('"')[0].split("'")[0]
    cleaned = re.sub(r"(https?://(?:m\.)?blog\.naver\.com/[^\s<>\"]+).*", r"\1", cleaned)
    parsed = urlparse(cleaned)

    # Naver search cards often wrap the real destination in query params.
    query = parse_qs(parsed.query)
    for key in ("url", "u", "target", "targetUrl", "src"):
        values = query.get(key)
        if not values:
            continue
        candidate = values[0]
        candidate = html.unescape(unquote(candidate)).replace("\\u002F", "/").replace("\\/", "/")
        candidate = candidate.split("&quot;")[0].split('"')[0].split("'")[0]
        if candidate.startswith("//"):
            candidate = f"https:{candidate}"
        if candidate.startswith("http://") or candidate.startswith("https://"):
            return candidate

    return cleaned


def is_valid_naver_blog_url(url: str) -> bool:
    if not url:
        return False
    parsed = urlparse(url)
    if parsed.netloc not in {"blog.naver.com", "m.blog.naver.com"}:
        return False
    lowered = url.lower()
    if "&quot;" in lowered or "source=" in lowered:
        return False
    path = parsed.path.strip("/")
    if not path or path.lower().startswith("myblog.naver"):
        return False
    if path.lower().endswith(".naver"):
        return False
    if path.lower().startswith(("postview.naver", "postlist.naver", "prologue", "blogid")):
        return True
    segments = [segment for segment in path.split("/") if segment]
    return len(segments) >= 2


def classify_main_result_channel(title_href: str, block_hrefs: list[str], block_text: str) -> tuple[str, str]:
    normalized_candidates: list[str] = []
    for raw_href in [title_href, *block_hrefs]:
        href = normalize_naver_url(str(raw_href).strip())
        if href and href not in normalized_candidates:
            normalized_candidates.append(href)

    for href in normalized_candidates:
        if is_valid_naver_blog_url(href):
            return "blog", href

    lowered_text = block_text.lower()
    if "blog.naver.com" in lowered_text or "m.blog.naver.com" in lowered_text:
        for href in normalized_candidates:
            if href:
                return "blog", href
        return "blog", title_href

    for href in normalized_candidates:
        netloc = urlparse(href).netloc.lower()
        if "cafe.naver.com" in netloc:
            return "cafe", href
        if "kin.naver.com" in netloc or "kin" in netloc:
            return "kin", href
        if netloc and "naver.com" not in netloc:
            return "site", href

    if "지식in" in lowered_text or "지식인" in lowered_text:
        return "kin", title_href
    if "카페" in block_text:
        return "cafe", title_href
    return "other", title_href


def extract_blog_candidates_from_html(page_html: str, limit: int = 5) -> list[dict[str, str]]:
    patterns = [
        r'https?://blog\.naver\.com/[^\s"\'<>]+',
        r'https?://m\.blog\.naver\.com/[^\s"\'<>]+',
        r'https%3A%2F%2Fblog\.naver\.com%2F[^"\']+',
        r'https%3A%2F%2Fm\.blog\.naver\.com%2F[^"\']+',
    ]
    seen: set[str] = set()
    links: list[dict[str, str]] = []
    for pattern in patterns:
        for match in re.findall(pattern, page_html):
            cleaned = unquote(match).replace("\\u002F", "/").replace("\\/", "/")
            cleaned = html.unescape(cleaned).rstrip('",\'<>')
            cleaned = normalize_naver_url(cleaned)
            if not is_valid_naver_blog_url(cleaned):
                continue
            if cleaned in seen:
                continue
            seen.add(cleaned)
            links.append({"title": cleaned, "url": cleaned})
            if len(links) >= limit:
                return links
    return links


@st.cache_data(show_spinner=False, ttl=60 * 60)
def extract_main_page_results(keyword: str, limit: int = 40) -> list[dict[str, Any]]:
    search_url = f"https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query={requests.utils.quote(keyword)}"
    driver = None
    try:
        driver = setup_search_driver()
        driver.get(search_url)
        wait = WebDriverWait(driver, 12)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#main_pack, #content")))
        time.sleep(1.0)
        try:
            driver.execute_script("window.stop();")
        except Exception:
            pass

        data = driver.execute_script(
            """
            const normalize = (text) => (text || '').replace(/\\s+/g, ' ').trim();
            const titleSelector = 'a.title_link[href], a.api_txt_lines.total_tit[href], a.total_tit[href], a.link_tit[href], a[class*="title"][href]';
            let titleAnchors = Array.from(document.querySelectorAll(`#main_pack ${titleSelector}, #content ${titleSelector}`)).slice(0, 120);
            if (!titleAnchors.length) {
              titleAnchors = Array.from(document.querySelectorAll('#main_pack a[href], #content a[href]'))
                .filter((a) => normalize(a.textContent || '').length >= 8)
                .slice(0, 180);
            }
            const seen = new Set();
            const blocks = [];

            titleAnchors.forEach((a) => {
              const block = a.closest('li, article, div[class*="item"], div[class*="bx"], div[class*="total"], div[class*="detail"], section, div');
              if (!block || seen.has(block)) {
                return;
              }
              seen.add(block);
              const rect = block.getBoundingClientRect();
              const title = normalize(a.textContent || '').slice(0, 120);
              const blockText = normalize(block.textContent || '').slice(0, 220);
              const blockClass = normalize(block.className || '').slice(0, 120);
              const headerNode = block.querySelector('h2, h3, strong, [class*="title"]');
              const headerText = normalize(headerNode ? (headerNode.textContent || '') : '').slice(0, 80);
              const hrefs = Array.from(block.querySelectorAll('a[href]'))
                .slice(0, 12)
                .map((link) => link.href || '')
                .filter(Boolean);
              blocks.push({
                titleHref: a.href || '',
                title,
                hrefs,
                y: Number.isFinite(rect.top) ? rect.top : 999999,
                x: Number.isFinite(rect.left) ? rect.left : 999999,
                blockText,
                blockClass,
                headerText,
              });
            });

            return blocks;
            """
        )

        results: list[dict[str, Any]] = []
        seen_keys: set[tuple[str, str]] = set()
        for item in data:
            title = str(item.get("title", "")).strip()
            if len(title) < 6:
                continue
            y = float(item.get("y", 999999))
            if y < 0 or y > 5200:
                continue
            channel, resolved_url = classify_main_result_channel(
                str(item.get("titleHref", "")).strip(),
                list(item.get("hrefs", [])),
                str(item.get("blockText", "")),
            )
            if channel == "other":
                continue
            dedupe_value = resolved_url or title
            key = (channel, dedupe_value)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            results.append(
                {
                    "title": title,
                    "url": resolved_url,
                    "channel": channel,
                    "y": y,
                    "header": str(item.get("headerText", ""))[:80],
                    "hint": str(item.get("blockText", ""))[:180],
                }
            )

        results.sort(key=lambda row: row["y"])
        if not results:
            html_links = extract_blog_candidates_from_html(driver.page_source, limit=max(limit, 10))
            if html_links:
                return [
                    {
                        "title": item["title"],
                        "url": item["url"],
                        "channel": "blog",
                        "y": 9999,
                        "header": "HTML fallback",
                        "hint": "",
                    }
                    for item in html_links[:limit]
                ]
        return results[:limit]
    except Exception as exc:
        if "Selenium" in str(exc):
            raise
        raise RuntimeError(
            f"셀레니움 브라우저를 시작하지 못했습니다. {exc.__class__.__name__}: {exc}"
        ) from exc
        raise RuntimeError("네이버 통합검색 페이지 로딩이 지연되었습니다. 잠시 후 다시 시도해 주세요.") from exc
    finally:
        if driver is not None:
            driver.quit()


@st.cache_data(show_spinner=False, ttl=60 * 30)
def debug_extract_blog_links(keyword: str, limit: int = 15) -> list[dict[str, str]]:
    rows = extract_main_page_results(keyword, limit=limit)
    return [
        {
            "title": row["title"],
            "url": row["url"],
            "채널": row["channel"],
            "헤더": row["header"],
            "블록 힌트": row["hint"],
            "Y": row["y"],
        }
        for row in rows
    ]


def extract_text_from_selectors(soup: BeautifulSoup, selectors: list[str]) -> str:
    for selector in selectors:
        node = soup.select_one(selector)
        if node:
            text = node.get_text("\n", strip=True)
            if text:
                return text
    return ""


def extract_top_terms(text: str, limit: int = 12) -> list[tuple[str, int]]:
    tokens = re.findall(r"[가-힣A-Za-z]{2,}", text)
    stopwords = {
        "그리고",
        "하지만",
        "때문에",
        "정말",
        "너무",
        "이번",
        "통해",
        "대한",
        "관련",
        "시술",
        "병원",
        "의원",
        "클리닉",
        "효과",
        "관리",
        "추천",
        "진행",
        "고객",
        "안내",
        "후기",
        "경우",
        "있어요",
        "합니다",
        "였습니다",
        "입니다",
        "가능",
        "부분",
        "위해",
        "네이버",
        "블로그",
    }
    filtered = [token for token in tokens if token not in stopwords and len(token) >= 2]
    counts = Counter(filtered)
    return counts.most_common(limit)


def count_blog_media(content_root: BeautifulSoup | None, soup: BeautifulSoup) -> tuple[int, int, int]:
    root = content_root or soup

    def normalize_media_url(raw_url: Any) -> str:
        if not raw_url:
            return ""
        url = html.unescape(str(raw_url)).strip()
        if not url or url.startswith("data:"):
            return ""
        url = url.split("#", 1)[0]
        url = url.split("?", 1)[0]
        return url

    def has_excluded_context(node: Any) -> bool:
        for parent in [node, *list(getattr(node, "parents", []))[:6]]:
            classes = " ".join(parent.get("class", [])) if getattr(parent, "get", None) else ""
            lowered = classes.lower()
            if any(token in lowered for token in ["comment", "profile", "reaction", "badge", "sympathy", "nclicks"]):
                return True
        return False

    def nearest_media_block(img: Any) -> str:
        block_selectors = [
            ".se-imageStrip-item",
            ".se-gallery-image-item",
            ".se-component.se-image",
            ".se-module-image",
            ".se-section-image",
            ".se-image-resource",
            ".se-media-image",
        ]
        for parent in [img, *list(getattr(img, "parents", []))[:8]]:
            if not getattr(parent, "select_one", None):
                continue
            for selector in block_selectors:
                try:
                    if parent.select_one(":scope") is not None and parent.select_one(selector) is parent:
                        return f"block:{id(parent)}"
                except Exception:
                    pass
            parent_classes = " ".join(parent.get("class", [])) if parent.get("class") else ""
            if any(
                token in parent_classes
                for token in [
                    "se-imageStrip-item",
                    "se-gallery-image-item",
                    "se-component",
                    "se-module-image",
                    "se-section-image",
                    "se-image-resource",
                ]
            ):
                return f"block:{id(parent)}"
        return ""

    image_nodes = root.select("img")
    media_blocks: set[str] = set()
    gif_blocks: set[str] = set()
    excluded_tokens = ["profile", "icon", "sp_common", "emoji", "sticker", "avatar", "blank.gif"]

    for img in image_nodes:
        if has_excluded_context(img):
            continue

        candidates = [
            img.get("data-lazy-src"),
            img.get("data-src"),
            img.get("data-gif-url"),
            img.get("src"),
        ]
        src = ""
        for candidate in candidates:
            normalized = normalize_media_url(candidate)
            if normalized:
                src = normalized
                break
        if not src:
            continue

        lowered = src.lower()
        if any(token in lowered for token in excluded_tokens):
            continue

        width = img.get("width") or img.get("data-width")
        height = img.get("height") or img.get("data-height")
        try:
            if width and height and int(width) <= 80 and int(height) <= 80:
                continue
        except (TypeError, ValueError):
            pass

        block_key = nearest_media_block(img) or f"url:{src}"
        if ".gif" in lowered or "type=g" in lowered or "animated" in lowered:
            gif_blocks.add(block_key)
        media_blocks.add(block_key)

    video_nodes = root.select(
        "video, iframe[src*='tv.naver.com'], iframe[src*='youtube.com'], iframe[src*='youtu.be']"
    )
    video_urls = {
        normalize_media_url(node.get("src") or node.get("poster") or node.get("data-video-url"))
        for node in video_nodes
    }
    video_urls = {url for url in video_urls if url}
    return len(media_blocks), len(video_urls), len(gif_blocks)


def blog_post_keyword_relevance(post: BlogPostMetrics, keyword: str) -> int:
    focus_tokens = query_focus_tokens(keyword)
    haystack = normalize_keyword_text(f"{post.title} {post.content[:2500]}")
    if not focus_tokens:
        return 0
    score = 0
    for token in focus_tokens:
        if token in normalize_keyword_text(post.title):
            score += 3
        if token in haystack:
            score += 1
    return score


def fetch_naver_blog_post(session: requests.Session, item: dict[str, str], rank: int) -> BlogPostMetrics:
    response = session.get(item["url"], timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    iframe = soup.select_one("iframe#mainFrame")
    if iframe and iframe.get("src"):
        iframe_url = f"https://blog.naver.com{iframe['src']}"
        response = session.get(iframe_url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

    title = extract_text_from_selectors(
        soup,
        [".se-title-text span", ".pcol1 .title", ".title_1"],
    )
    if not title:
        meta_title = soup.select_one("meta[property='og:title']")
        title = meta_title.get("content", "").strip() if meta_title else item["title"]

    content = extract_text_from_selectors(
        soup,
        [".se-main-container", "#postViewArea", ".post-view", ".contents_style", ".post_ct"],
    )

    if not content:
        scripts = soup.find_all("script")
        joined_scripts = "\n".join(script.get_text(" ", strip=True) for script in scripts)
        matches = re.findall(r'"(?:content|summary|description)"\s*:\s*"(.+?)"', joined_scripts)
        decoded = " ".join(bytes(match, "utf-8").decode("unicode_escape", "ignore") for match in matches[:20])
        content = re.sub(r"<[^>]+>", " ", decoded)

    content = re.sub(r"\s+", " ", content).strip()
    content_root = None
    for selector in [".se-main-container", "#postViewArea", ".post-view", ".contents_style", ".post_ct"]:
        content_root = soup.select_one(selector)
        if content_root:
            break
    try:
        image_count, video_count, gif_count = count_blog_media(content_root, soup)
    except Exception:
        image_count, video_count, gif_count = 0, 0, 0
    author_node = soup.select_one(".nick, .blog2_series .ell, .se_author .nick")
    author = author_node.get_text(" ", strip=True) if author_node else "작성자 미상"

    return BlogPostMetrics(
        rank=rank,
        title=title or item["title"],
        url=item["url"],
        author=author,
        content=content,
        char_count=len(content),
        image_count=image_count,
        video_count=video_count,
        gif_count=gif_count,
        top_terms=extract_top_terms(content),
    )


@st.cache_data(show_spinner=False, ttl=60 * 60)
def analyze_keyword_competition(keyword: str, view_top_n: int) -> dict[str, Any]:
    session = make_session()
    main_page_results = extract_main_page_results(keyword, limit=40)
    channel_counts = Counter(row["channel"] for row in main_page_results)
    blog_links = [row for row in main_page_results if row["channel"] == "blog"]
    posts: list[BlogPostMetrics] = []

    if not blog_links:
        dominant_channel = channel_counts.most_common(1)[0][0] if channel_counts else "other"
        channel_label = {
            "cafe": "카페",
            "kin": "지식인",
            "site": "웹사이트",
            "blog": "블로그",
            "other": "기타",
        }.get(dominant_channel, dominant_channel)
        return {
            "keyword": keyword,
            "post_count": 0,
            "avg_chars": 0,
            "median_chars": 0,
            "avg_images": 0,
            "top_terms": [],
            "recommended_chars": 0,
            "recommended_images": 0,
            "writing_angle": "메인 1면에 블로그 노출이 확인되지 않았습니다.",
            "posts": [],
            "main_page_results": main_page_results,
            "channel_counts": dict(channel_counts),
            "channel_insight": f"메인 1면에는 블로그 노출이 확인되지 않았습니다. 현재는 {channel_label} 채널이 더 강하게 보입니다.",
        }

    for idx, item in enumerate(blog_links, start=1):
        try:
            posts.append(fetch_naver_blog_post(session, item, rank=idx))
        except Exception:
            continue

    if not posts:
        raise RuntimeError("메인 1면의 블로그 링크는 확인됐지만 본문을 읽지 못했습니다. 다시 시도해 주세요.")

    relevant_posts = [post for post in posts if blog_post_keyword_relevance(post, keyword) > 0]
    if relevant_posts:
        posts = relevant_posts

    posts.sort(key=lambda post: post.rank)
    posts = posts[:view_top_n]
    for idx, post in enumerate(posts, start=1):
        post.rank = idx

    avg_chars = round(sum(post.char_count for post in posts) / len(posts))
    avg_images = round(sum(post.image_count for post in posts) / len(posts), 1)
    avg_videos = round(sum(post.video_count for post in posts) / len(posts), 1)
    avg_gifs = round(sum(post.gif_count for post in posts) / len(posts), 1)
    median_chars = int(pd.Series([post.char_count for post in posts]).median())
    term_counter: Counter[str] = Counter()
    for post in posts:
        term_counter.update(dict(post.top_terms))

    top_terms = term_counter.most_common(12)
    summary = {
        "keyword": keyword,
        "post_count": len(posts),
        "avg_chars": avg_chars,
        "median_chars": median_chars,
        "avg_images": avg_images,
        "avg_videos": avg_videos,
        "avg_gifs": avg_gifs,
        "top_terms": top_terms,
        "recommended_chars": max(1400, median_chars + 150),
        "recommended_images": max(4, math.ceil(avg_images)),
        "writing_angle": build_writing_angle(top_terms),
        "posts": [post.__dict__ for post in posts],
        "main_page_results": main_page_results,
        "channel_counts": dict(channel_counts),
        "channel_insight": "",
    }
    return summary


def build_writing_angle(top_terms: list[tuple[str, int]]) -> str:
    if not top_terms:
        return "시술 설명 + 상담 포인트 중심으로 구성"
    terms = [term for term, _ in top_terms[:4]]
    return f"{', '.join(terms)} 중심의 설명형 콘텐츠가 유리"


def build_post_detail_table(posts: list[dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "순위": post["rank"],
                "제목": post["title"],
                "작성자": post["author"],
                "글자 수": post["char_count"],
                "이미지 수": post["image_count"],
                "원문 링크": post["url"],
            }
            for post in posts
        ]
    )


def render_keyword_diagnosis_section(summary: dict[str, Any]) -> None:
    diagnosis = build_keyword_diagnosis(summary)
    st.markdown("**키워드 종합진단**")

    dominant_text = diagnosis.channel_dominance
    has_blog = summary.get("post_count", 0) > 0
    if "우세" in dominant_text and has_blog and not dominant_text.startswith("블로그"):
        mixed_signal = f"{dominant_text}, 다만 블로그도 메인 1면에 일부 노출됩니다."
    else:
        mixed_signal = dominant_text

    if diagnosis.opportunity_label == "진입 가능성 높음":
        hero_badge_class = "good"
    elif diagnosis.opportunity_label == "테스트 추천":
        hero_badge_class = "warn"
    else:
        hero_badge_class = "cold"

    hero_message = (
        f"이 키워드는 `{mixed_signal}` 흐름입니다. "
        f"블로그 관점에서는 `{diagnosis.opportunity_label}` 단계로 보고, "
        f"`{diagnosis.writing_format}` 포맷으로 접근하는 편이 좋습니다."
    )

    st.markdown(
        f"""
        <div class="diagnosis-hero">
            <div class="diagnosis-hero-title">작성 판단 한 줄 결론
                <span class="diagnosis-badge {hero_badge_class}">{diagnosis.opportunity_label}</span>
            </div>
            <div class="diagnosis-hero-main">{hero_message}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    top_cols = st.columns(5)
    top_cols[0].metric("키워드명", diagnosis.keyword)
    top_cols[1].metric("PC", f"{diagnosis.monthly_pc:,}")
    top_cols[2].metric("Mobile", f"{diagnosis.monthly_mobile:,}")
    top_cols[3].metric("Total", f"{diagnosis.monthly_total:,}")
    top_cols[4].metric("광고 경쟁도", diagnosis.competition_index)

    insight_cols = st.columns(3)
    with insight_cols[0]:
        st.markdown(
            f"""
            <div class="diagnosis-card">
                <div class="diagnosis-label">실측</div>
                <ul class="diagnosis-list">
                    <li>메인 1면 블로그: <b>{diagnosis.blog_presence}</b></li>
                    <li>메인 1면 채널 해석: <b>{mixed_signal}</b></li>
                    <li>블로그 문서 수(추정): <b>{diagnosis.estimated_blog_docs:,}건</b></li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with insight_cols[1]:
        est_lines = [
            f"<li>블로그 포화도: <b>{diagnosis.estimated_saturation:.2f}</b></li>",
            f"<li>월별 추이 해석: <b>{diagnosis.trend_change_label}</b></li>",
        ]
        if diagnosis.expected_month_to_date is not None and diagnosis.expected_month_total is not None:
            est_lines.insert(1, f"<li>이번 달 현재 추정 검색량: <b>{diagnosis.expected_month_to_date:,}</b></li>")
            est_lines.insert(2, f"<li>이번 달 월말 예상 검색량: <b>{diagnosis.expected_month_total:,}</b></li>")
        else:
            est_lines.insert(1, "<li>월말 예상 검색량은 SearchAd 실측 검색량이 잡힐 때 표시합니다.</li>")
        st.markdown(
            f"""
            <div class="diagnosis-card">
                <div class="diagnosis-label">추정</div>
                <ul class="diagnosis-list">
                    {''.join(est_lines)}
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with insight_cols[2]:
        st.markdown(
            f"""
            <div class="diagnosis-card">
                <div class="diagnosis-label">추천</div>
                <ul class="diagnosis-list">
                    <li>블로그 진입 판단: <b>{diagnosis.opportunity_label}</b></li>
                    <li>추천 작성 형식: <b>{diagnosis.writing_format}</b></li>
                    <li>{diagnosis.content_recommendation}</li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )

    if diagnosis.trend_rows:
        trend_df = pd.DataFrame(diagnosis.trend_rows)
        peak_ratio = trend_df["ratio"].max()
        min_ratio = trend_df["ratio"].min()
        trend_df["bar_type"] = trend_df["ratio"].apply(
            lambda value: "피크 월" if value == peak_ratio else "최저 월" if value == min_ratio else "일반 월"
        )
        st.markdown("**최근 12개월 월별 검색 트렌드**")
        trend_chart = (
            alt.Chart(trend_df)
            .mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6)
            .encode(
                x=alt.X("month:N", sort=list(trend_df["month"]), axis=alt.Axis(labelAngle=0, labelPadding=10, title=None)),
                y=alt.Y("ratio:Q", title=None),
                color=alt.Color(
                    "bar_type:N",
                    scale=alt.Scale(
                        domain=["피크 월", "최저 월", "일반 월"],
                        range=["#1d4ed8", "#ef4444", "#8ec5ff"],
                    ),
                    legend=alt.Legend(title=None, orient="top"),
                ),
                tooltip=[
                    alt.Tooltip("month:N", title="월"),
                    alt.Tooltip("ratio:Q", title="상대 추이", format=".1f"),
                    alt.Tooltip("bar_type:N", title="구분"),
                ],
            )
            .properties(height=320)
        )
        st.altair_chart(trend_chart, use_container_width=True)
        if diagnosis.trend_peak_month:
            st.caption(
                f"DataLab 기준 피크 월은 `{diagnosis.trend_peak_month}`이며 현재는 `{diagnosis.trend_change_label}` 흐름입니다. "
                "월별 추이는 상대 비율이고, 월말 예상 검색량은 SearchAd 월간 검색수에 계절성 비율을 반영한 추정치입니다."
            )
    else:
        st.caption("DataLab API 키가 없으면 월별 검색 트렌드와 월말 예상 검색량은 표시되지 않습니다.")


def build_channel_content_ideas(channel_counts: dict[str, int]) -> list[str]:
    if not channel_counts:
        return ["메인 1면 채널 구성이 확인되지 않아 추가 판단이 필요합니다."]

    dominant = max(channel_counts.items(), key=lambda item: item[1])[0]
    ideas = []
    if dominant == "cafe":
        ideas.append("카페형 질문/후기 문체가 강한 키워드입니다. 실제 고민형 제목과 댓글 유도형 구성이 유리합니다.")
        ideas.append("비교 질문형 콘텐츠를 먼저 검토해 보세요. 예: 효과, 유지기간, 통증, 원장별 차이.")
    elif dominant == "kin":
        ideas.append("지식인형 정보 수요가 강한 키워드입니다. 짧고 명확한 Q&A 구조가 유리합니다.")
        ideas.append("블로그도 작성하되, FAQ형 소제목과 답변형 문장을 더 강화하는 편이 좋습니다.")
    elif dominant == "site":
        ideas.append("병원/브랜드 웹페이지 성격의 노출이 강합니다. 시술 개요, 장점, 적응증을 명확히 정리한 소개형 콘텐츠가 유리합니다.")
        ideas.append("블로그를 쓰더라도 랜딩형 제목보다 정보 정리형 제목이 더 적합할 수 있습니다.")
    else:
        ideas.append("채널 우세가 뚜렷하지 않습니다. 블로그와 카페형 소재를 병행 테스트해 보세요.")
    return ideas


def resolve_procedure_profile(seed_keyword: str, custom_description: str, operator_name: str | None = None) -> dict[str, Any]:
    combined_profiles = {**DEFAULT_PROCEDURE_PROFILES, **load_user_procedure_profiles(operator_name)}
    for profile_keyword, profile in combined_profiles.items():
        if normalize_keyword_text(profile_keyword) in normalize_keyword_text(seed_keyword) or normalize_keyword_text(seed_keyword) in normalize_keyword_text(profile_keyword):
            return {
                "profile_name": profile_keyword,
                "source": "preset",
                **profile,
            }
    description = custom_description.strip()
    if description:
        return {
            "profile_name": seed_keyword,
            "source": "custom",
            "core_description": description,
            "must_include": [],
            "avoid_phrases": ["무조건", "반드시 효과", "완치", "영구 유지"],
            "preferred_format": "설명형",
            "cta_style": "개인별 피부 상태와 고민에 따라 적합성이 다를 수 있다는 안내를 넣습니다.",
        }
    return {
        "profile_name": seed_keyword,
        "source": "generic",
        "core_description": "시술 원리, 기대 효과, 통증/붓기, 유지기간, 적합 대상, 주의사항을 균형 있게 설명합니다.",
        "must_include": ["효과", "통증", "유지기간", "적합 대상", "주의사항"],
        "avoid_phrases": ["무조건", "반드시 효과", "완치", "영구 유지"],
        "preferred_format": "설명형",
        "cta_style": "상담 전 현재 고민과 피부 상태에 맞는지 확인하는 문장을 넣습니다.",
    }


def load_user_procedure_profiles(operator_name: str | None = None) -> dict[str, dict[str, Any]]:
    profile_path = operator_profile_path(operator_name)
    if not profile_path.exists():
        return {}
    try:
        return json.loads(profile_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_user_procedure_profile(profile_name: str, profile_data: dict[str, Any], operator_name: str | None = None) -> None:
    profile_path = operator_profile_path(operator_name)
    profile_path.parent.mkdir(parents=True, exist_ok=True)
    profiles = load_user_procedure_profiles(operator_name)
    profiles[profile_name] = profile_data
    profile_path.write_text(
        json.dumps(profiles, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def delete_user_procedure_profile(profile_name: str, operator_name: str | None = None) -> None:
    profiles = load_user_procedure_profiles(operator_name)
    if profile_name in profiles:
        del profiles[profile_name]
        profile_path = operator_profile_path(operator_name)
        profile_path.parent.mkdir(parents=True, exist_ok=True)
        profile_path.write_text(
            json.dumps(profiles, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def procedure_source_dir(procedure_name: str, operator_name: str | None = None) -> Path:
    normalized = normalize_keyword_text(procedure_name) or "default"
    return operator_source_dir(operator_name) / normalized


def list_user_reference_files(procedure_name: str, operator_name: str | None = None) -> list[Path]:
    directory = procedure_source_dir(procedure_name, operator_name)
    if not directory.exists():
        return []
    return sorted([path for path in directory.iterdir() if path.is_file()])


def save_uploaded_reference_files(procedure_name: str, uploaded_files: list[Any], operator_name: str | None = None) -> None:
    directory = procedure_source_dir(procedure_name, operator_name)
    directory.mkdir(parents=True, exist_ok=True)
    for uploaded in uploaded_files:
        target = directory / uploaded.name
        target.write_bytes(uploaded.getbuffer())


def delete_user_reference_file(procedure_name: str, filename: str, operator_name: str | None = None) -> None:
    target = procedure_source_dir(procedure_name, operator_name) / filename
    if target.exists():
        target.unlink()


def list_builtin_reference_files(procedure_name: str) -> list[Path]:
    normalized = normalize_keyword_text(procedure_name)
    target_dir = None
    for key, path in PROCEDURE_DOC_DIRECTORIES.items():
        if normalize_keyword_text(key) in normalized or normalized in normalize_keyword_text(key):
            target_dir = path
            break
    if not target_dir or not target_dir.exists():
        return []
    return sorted(
        [
            path
            for path in target_dir.iterdir()
            if path.is_file() and path.suffix.lower() in {".docx", ".doc", ".txt", ".pdf"}
        ]
    )


def import_builtin_reference_file(procedure_name: str, source_file: Path, operator_name: str | None = None) -> None:
    target_dir = procedure_source_dir(procedure_name, operator_name)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / source_file.name
    target_path.write_bytes(source_file.read_bytes())


def _default_translation_prompt(language_label: str) -> str:
    if language_label in DEFAULT_TRANSLATION_PROMPTS:
        return DEFAULT_TRANSLATION_PROMPTS[language_label]
    preset = TRANSLATION_PRESETS[language_label]
    return (
        f"{language_label}({preset['locale']}) 번역 기본 프롬프트\n"
        f"- {preset['style']}\n"
        f"- {preset['notes']}\n"
        "- 병원 정보성 블로그 톤 유지\n"
        "- 과장 광고 문구 금지\n"
        "- 표, 볼드, 소제목 구조 유지"
    )


def ensure_translation_data_seeded() -> None:
    TRANSLATION_DATA_ROOT.mkdir(parents=True, exist_ok=True)

    if TRANSLATION_PROMPTS_PATH.exists():
        try:
            prompts = json.loads(TRANSLATION_PROMPTS_PATH.read_text(encoding="utf-8"))
        except Exception:
            prompts = {}
    else:
        prompts = {}

    if TRANSLATION_GLOSSARIES_PATH.exists():
        try:
            glossaries = json.loads(TRANSLATION_GLOSSARIES_PATH.read_text(encoding="utf-8"))
        except Exception:
            glossaries = {}
    else:
        glossaries = {}

    changed = False
    for language_label in TRANSLATION_PRESETS:
        default_prompt = _default_translation_prompt(language_label)
        default_glossary = _default_translation_glossary(language_label)

        if language_label not in prompts or prompts.get(language_label, {}).get("updated_by") == "system_seed":
            prompts[language_label] = {
                "text": default_prompt,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "updated_by": "system_seed",
            }
            changed = True

        default_filename = "system_seed"
        if language_label == "일본어" and JAPANESE_GLOSSARY_SEED_PATH.exists():
            default_filename = JAPANESE_GLOSSARY_SEED_PATH.name

        if language_label not in glossaries or glossaries.get(language_label, {}).get("updated_by") == "system_seed":
            glossaries[language_label] = {
                "terms": default_glossary,
                "filename": default_filename,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "updated_by": "system_seed",
            }
            changed = True

    if changed:
        TRANSLATION_PROMPTS_PATH.write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")
        TRANSLATION_GLOSSARIES_PATH.write_text(json.dumps(glossaries, ensure_ascii=False, indent=2), encoding="utf-8")


def load_translation_prompts() -> dict[str, dict[str, Any]]:
    ensure_translation_data_seeded()
    try:
        return json.loads(TRANSLATION_PROMPTS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_translation_prompt(language_label: str, text: str, username: str) -> None:
    prompts = load_translation_prompts()
    prompts[language_label] = {
        "text": text.strip(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": username,
    }
    TRANSLATION_PROMPTS_PATH.write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")


def delete_translation_prompt(language_label: str, username: str) -> None:
    prompts = load_translation_prompts()
    prompts[language_label] = {
        "text": _default_translation_prompt(language_label),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": f"{username} (reset)",
    }
    TRANSLATION_PROMPTS_PATH.write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")


def load_translation_glossaries() -> dict[str, dict[str, Any]]:
    ensure_translation_data_seeded()
    try:
        return json.loads(TRANSLATION_GLOSSARIES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_translation_glossary(language_label: str, terms: dict[str, str], filename: str, username: str) -> None:
    glossaries = load_translation_glossaries()
    glossaries[language_label] = {
        "terms": terms,
        "filename": filename,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": username,
    }
    TRANSLATION_GLOSSARIES_PATH.write_text(json.dumps(glossaries, ensure_ascii=False, indent=2), encoding="utf-8")


def delete_translation_glossary(language_label: str, username: str) -> None:
    glossaries = load_translation_glossaries()
    default_filename = "system_seed"
    if language_label == "일본어" and JAPANESE_GLOSSARY_SEED_PATH.exists():
        default_filename = JAPANESE_GLOSSARY_SEED_PATH.name
    glossaries[language_label] = {
        "terms": _default_translation_glossary(language_label),
        "filename": default_filename,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": f"{username} (reset)",
    }
    TRANSLATION_GLOSSARIES_PATH.write_text(json.dumps(glossaries, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_glossary_upload(uploaded_file: Any) -> dict[str, str]:
    suffix = Path(uploaded_file.name).suffix.lower()
    raw_bytes = uploaded_file.getvalue()
    if suffix == ".json":
        payload = json.loads(raw_bytes.decode("utf-8"))
        if isinstance(payload, dict):
            return {str(k).strip(): str(v).strip() for k, v in payload.items() if str(k).strip() and str(v).strip()}
        if isinstance(payload, list):
            parsed = {}
            for row in payload:
                if isinstance(row, dict):
                    keys = list(row.keys())
                    if len(keys) >= 2:
                        source = str(row.get(keys[0], "")).strip()
                        target = str(row.get(keys[1], "")).strip()
                        if source and target:
                            parsed[source] = target
            return parsed
        raise RuntimeError("JSON 용어집 형식을 이해하지 못했습니다.")

    if suffix == ".csv":
        df = pd.read_csv(io.BytesIO(raw_bytes))
    elif suffix == ".xlsx":
        df = pd.read_excel(io.BytesIO(raw_bytes))
    else:
        raise RuntimeError("용어집은 .json, .csv, .xlsx 형식만 업로드할 수 있습니다.")

    if df.shape[1] < 2:
        raise RuntimeError("용어집 파일에는 최소 2개 컬럼이 필요합니다.")
    first_col, second_col = df.columns[:2]
    parsed = {}
    for _, row in df.iterrows():
        source = str(row[first_col]).strip()
        target = str(row[second_col]).strip()
        if source and source.lower() != "nan" and target and target.lower() != "nan":
            parsed[source] = target
    return parsed


def compress_prompt_text(text: str, max_chars: int = 3500) -> str:
    cleaned = re.sub(r"\s+", " ", str(text)).strip()
    return cleaned[:max_chars]


@st.cache_data(show_spinner=False, ttl=60 * 60 * 12)
def read_rich_text_document(path_str: str) -> str:
    path = Path(path_str)
    if not path.exists():
        return ""
    if path.suffix.lower() == ".pdf":
        return ""
    if path.suffix.lower() in {".txt", ".md"}:
        try:
            return compress_prompt_text(path.read_text(encoding="utf-8"), max_chars=5000)
        except Exception:
            try:
                return compress_prompt_text(path.read_text(encoding="utf-8-sig"), max_chars=5000)
            except Exception:
                return ""
    try:
        result = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", str(path)],
            capture_output=True,
            text=True,
            check=True,
        )
        return compress_prompt_text(result.stdout, max_chars=5000)
    except Exception:
        return ""


@st.cache_data(show_spinner=False, ttl=60 * 60 * 12)
def load_prompt_knowledge(seed_keyword: str, operator_name: str) -> dict[str, Any]:
    brand_info = read_rich_text_document(str(BRAND_INFO_DOC_PATH))
    blog_reference = read_rich_text_document(str(BLOG_REFERENCE_DOC_PATH))
    high_performance_examples = read_rich_text_document(str(HIGH_PERFORMANCE_EXAMPLE_DOC_PATH))

    procedure_dir = None
    normalized_seed = normalize_keyword_text(seed_keyword)
    for key, path in PROCEDURE_DOC_DIRECTORIES.items():
        if normalize_keyword_text(key) in normalized_seed or normalized_seed in normalize_keyword_text(key):
            procedure_dir = path
            break

    procedure_docs: list[dict[str, str]] = []
    if procedure_dir and procedure_dir.exists():
        for file_path in sorted(procedure_dir.iterdir()):
            if not file_path.is_file() or file_path.suffix.lower() not in {".docx", ".doc", ".txt"}:
                continue
            content = read_rich_text_document(str(file_path))
            if not content:
                continue
            procedure_docs.append(
                {
                    "name": file_path.stem,
                    "content": compress_prompt_text(content, max_chars=3000),
                }
            )

    for file_path in list_user_reference_files(seed_keyword, operator_name):
        if file_path.suffix.lower() not in {".docx", ".doc", ".txt"}:
            continue
        content = read_rich_text_document(str(file_path))
        if not content:
            continue
        procedure_docs.append(
            {
                "name": f"사용자 업로드 - {file_path.stem}",
                "content": compress_prompt_text(content, max_chars=3000),
            }
        )

    return {
        "brand_info": compress_prompt_text(brand_info, max_chars=5000),
        "blog_reference": compress_prompt_text(blog_reference, max_chars=3500),
        "high_performance_examples": compress_prompt_text(high_performance_examples, max_chars=3500),
        "procedure_docs": procedure_docs,
    }


@st.cache_data(show_spinner=False, ttl=60 * 30)
def fetch_reference_url_summary(url: str) -> dict[str, str]:
    response = make_session().get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    title = (
        extract_text_from_selectors(soup, ["title", "meta[property='og:title']"])
        or soup.title.get_text(" ", strip=True)
        if soup.title
        else url
    )
    text = extract_text_from_selectors(
        soup,
        ["article", "main", ".se-main-container", "#postViewArea", ".post-view", ".entry-content", ".elementor-widget-container", "body"],
    )
    summary = compress_prompt_text(text, max_chars=2500)
    return {"title": title or url, "summary": summary, "url": url}


def generate_gemini_json_response(model: Any, prompt: str, response_label: str) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent",
                params={"key": GEMINI_API_KEY},
                json={
                    "contents": [
                        {
                            "role": "user",
                            "parts": [{"text": prompt}],
                        }
                    ],
                    "generationConfig": {
                        "temperature": 0.8,
                        "topP": 0.95,
                    },
                },
                timeout=(10, 90),
            )
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                try:
                    wait_seconds = float(retry_after) if retry_after else (2.5 * (attempt + 1))
                except ValueError:
                    wait_seconds = 2.5 * (attempt + 1)
                if attempt < 2:
                    time.sleep(wait_seconds)
                    continue
                raise RuntimeError(
                    "Gemini API 호출이 잠시 제한되었습니다(429). 잠시 후 다시 시도해 주세요."
                )
            response.raise_for_status()
            payload_json = response.json()
            candidates = payload_json.get("candidates", [])
            if not candidates:
                raise RuntimeError(f"{response_label} 응답에 후보가 없습니다.")
            parts = candidates[0].get("content", {}).get("parts", [])
            text = "\n".join(str(part.get("text", "")) for part in parts if part.get("text")).strip()
            return parse_llm_json_response(text, response_label)
        except Exception as exc:
            last_error = exc
            error_text = str(exc)
            is_network_issue = any(
                token in error_text
                for token in [
                    "DNS resolution failed",
                    "Could not contact DNS servers",
                    "503",
                    "Timeout",
                    "timed out",
                    "ServiceUnavailable",
                ]
            )
            is_rate_limited = "429" in error_text or "Too Many Requests" in error_text or "잠시 제한" in error_text
            if attempt < 2 and is_rate_limited:
                time.sleep(2.5 * (attempt + 1))
                continue
            if is_rate_limited:
                raise RuntimeError(
                    "Gemini API 호출이 잠시 제한되었습니다(429). 잠시 후 다시 시도해 주세요."
                ) from exc
            if attempt < 2 and is_network_issue:
                time.sleep(1.5 * (attempt + 1))
                continue
            if is_network_issue:
                raise RuntimeError(
                    "Gemini API 네트워크 연결이 잠시 불안정합니다. 잠시 후 다시 시도해 주세요. "
                    "사내 DNS 또는 인터넷 연결 상태에 따라 일시적으로 발생할 수 있습니다."
                ) from exc
            raise
    raise RuntimeError(f"{response_label} 호출에 실패했습니다: {last_error}")


def generate_blog_draft_with_gemini(
    summary: dict[str, Any],
    clinic_name: str,
    tone_style: str,
    include_cta: bool,
    seed_keyword: str,
    procedure_profile: dict[str, Any],
    reference_urls_text: str,
    operator_name: str,
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")

    top_terms = ", ".join(term for term, _ in summary["top_terms"][:8])
    competitor_titles = "\n".join(f"- {post['title']}" for post in summary["posts"][:5])
    cta_rule = "마지막 문단에 상담 유도 문장을 2문장 이내로 추가합니다." if include_cta else "과한 상담 유도 문장은 넣지 않습니다."
    must_include = ", ".join(procedure_profile.get("must_include", [])) or "핵심 시술 특징"
    avoid_phrases = ", ".join(procedure_profile.get("avoid_phrases", [])) or "과장 표현"
    prompt_knowledge = load_prompt_knowledge(seed_keyword, operator_name)
    reference_urls = [line.strip() for line in reference_urls_text.splitlines() if line.strip()][:3]
    reference_summaries = []
    for url in reference_urls:
        try:
            reference_summaries.append(fetch_reference_url_summary(url))
        except Exception:
            continue
    reference_summary_text = "\n".join(
        f"- {item['title']} ({item['url']}): {item['summary']}"
        for item in reference_summaries
    ) or "추가 참고 URL 없음"
    procedure_doc_text = "\n".join(
        f"- {item['name']}: {item['content']}"
        for item in prompt_knowledge.get("procedure_docs", [])
    ) or "시술별 첨부 문서 없음"
    profile_source = {
        "preset": "저장된 시술 프리셋",
        "custom": "사용자 입력 시술 설명",
        "generic": "기본 공통 프롬프트",
    }.get(procedure_profile.get("source", "generic"), "기본 공통 프롬프트")

    prompt = f"""
당신은 병원 마케팅 블로그 전문 카피라이터입니다.
아래 분석 데이터를 바탕으로 네이버 블로그용 원고를 작성하세요.

[공통 작성 프롬프트]
{BASE_BLOG_WRITING_PROMPT}

[브랜드]
- 병원명: {clinic_name}
- 원하는 문체: {tone_style}
- 주력 시술명: {seed_keyword}

[세예의원 기본 정보 및 강점]
{prompt_knowledge.get("brand_info", "브랜드 정보 없음")}

[타겟 키워드]
{summary["keyword"]}

[시술 프로필]
- 적용 소스: {profile_source}
- 프로필 기준 시술명: {procedure_profile.get("profile_name", seed_keyword)}
- 시술 설명: {procedure_profile.get("core_description", "")}
- 반드시 녹일 포인트: {must_include}
- 피해야 할 표현: {avoid_phrases}
- 선호 포맷: {procedure_profile.get("preferred_format", "설명형")}
- CTA 가이드: {procedure_profile.get("cta_style", "")}

[시술별 첨부 문서]
{procedure_doc_text}

[블로그 레퍼런스 - 구조 참고 전용]
{prompt_knowledge.get("blog_reference", "블로그 레퍼런스 없음")}

[상위노출/조회수 높은 콘텐츠 예시 - 흐름 참고 전용]
{prompt_knowledge.get("high_performance_examples", "상위노출 예시 없음")}

[추가 참고 URL 요약]
{reference_summary_text}

[경쟁 문서 인사이트]
- 분석 문서 수: {summary["post_count"]}개
- 권장 글자 수: 약 {summary["recommended_chars"]}자 이상
- 권장 이미지 수: {summary["recommended_images"]}장
- 자주 쓰인 핵심 단어: {top_terms}
- 추천 작성 각도: {summary["writing_angle"]}
- 상위 제목 참고:
{competitor_titles}

[작성 원칙]
1. 제목은 클릭을 유도하되 과장 표현은 피합니다.
2. 본문은 도입, 문제 공감, 시술 설명, 기대 효과, 주의사항, 마무리 순서로 씁니다.
3. 타겟 키워드는 자연스럽게 5회 내외 포함합니다.
4. 의료 광고 문구처럼 단정적으로 표현하지 말고, 상담/개인차 문장을 적절히 섞습니다.
5. 실제 블로그에 바로 붙여넣기 좋게 일반 텍스트 중심으로 작성합니다.
6. {cta_rule}
7. [시술 프로필]에 적힌 반드시 녹일 포인트는 가능한 자연스럽게 본문에 반영합니다.
8. [시술 프로필]의 피해야 할 표현은 사용하지 않습니다.
9. 추천 작성 형식은 {procedure_profile.get("preferred_format", "설명형")}을 우선으로 따릅니다.
10. [세예의원 기본 정보 및 강점]의 내용은 광고처럼 보이지 않게 자연스럽게 녹입니다.
11. [블로그 레퍼런스]와 [상위노출/조회수 높은 콘텐츠 예시]는 문장 복제 금지, 구조와 흐름 참고용으로만 사용합니다.
12. [시술별 첨부 문서]와 [추가 참고 URL 요약]에 있는 팩트는 우선 반영하되, 문장은 새롭게 재구성합니다.

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "titles": ["제목1", "제목2", "제목3", "제목4", "제목5"],
  "body": "본문 초안",
  "outline": ["소제목1", "소제목2", "소제목3", "소제목4"],
  "checklist": [
    "글자 수: 준수 여부 / 실제 글자수",
    "주력 키워드 6회: 준수 여부 / 실제 횟수",
    "반복 형태소 점검: 준수 여부 / 주요 반복 단어",
    "의료법 준수 여부: 준수 여부 / 위반 표현 여부",
    "구조 반영 여부: 준수 여부"
  ]
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, "Gemini")
    titles = payload.get("titles", [])
    body = payload.get("body", "")
    outline = payload.get("outline", [])
    checklist = payload.get("checklist", [])
    if not titles or not body:
        raise RuntimeError("Gemini 응답에 제목 또는 본문이 비어 있습니다.")
    return {"titles": titles[:5], "body": body, "outline": outline, "checklist": checklist}


def revise_blog_draft_with_gemini(
    original_draft: dict[str, Any],
    revision_request: str,
    summary: dict[str, Any],
    clinic_name: str,
    tone_style: str,
    include_cta: bool,
    seed_keyword: str,
    procedure_profile: dict[str, Any],
    reference_urls_text: str,
    operator_name: str,
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")
    if not revision_request.strip():
        raise RuntimeError("추가 수정 요청을 입력해 주세요.")

    prompt_knowledge = load_prompt_knowledge(seed_keyword, operator_name)
    reference_urls = [line.strip() for line in reference_urls_text.splitlines() if line.strip()][:3]
    reference_summaries = []
    for url in reference_urls:
        try:
            reference_summaries.append(fetch_reference_url_summary(url))
        except Exception:
            continue
    reference_summary_text = "\n".join(
        f"- {item['title']} ({item['url']}): {item['summary']}"
        for item in reference_summaries
    ) or "추가 참고 URL 없음"

    prompt = f"""
당신은 병원 마케팅 블로그 전문 카피라이터입니다.
아래 초안은 이미 작성된 1차 원고입니다. 이 초안을 버리지 말고, 사용자의 추가 수정 요청을 반영해 2차 초안으로 재작성하세요.

[공통 작성 프롬프트]
{BASE_BLOG_WRITING_PROMPT}

[브랜드]
- 병원명: {clinic_name}
- 원하는 문체: {tone_style}
- 주력 시술명: {seed_keyword}

[세예의원 기본 정보 및 강점]
{prompt_knowledge.get("brand_info", "브랜드 정보 없음")}

[시술 프로필]
- 시술 설명: {procedure_profile.get("core_description", "")}
- 반드시 녹일 포인트: {", ".join(procedure_profile.get("must_include", []))}
- 피해야 할 표현: {", ".join(procedure_profile.get("avoid_phrases", []))}
- 선호 포맷: {procedure_profile.get("preferred_format", "설명형")}

[추가 참고 URL 요약]
{reference_summary_text}

[경쟁 문서 인사이트]
- 타겟 키워드: {summary["keyword"]}
- 권장 글자 수: 약 {summary["recommended_chars"]}자 이상
- 권장 이미지 수: {summary["recommended_images"]}장
- 추천 작성 각도: {summary["writing_angle"]}

[기존 1차 초안]
제목 후보:
{chr(10).join(f"- {title}" for title in original_draft.get("titles", []))}

소제목 구조:
{chr(10).join(f"- {item}" for item in original_draft.get("outline", []))}

본문:
{original_draft.get("body", "")}

[사용자 추가 수정 요청]
{revision_request}

[수정 원칙]
1. 사용자의 수정 요청을 최우선 반영합니다.
2. 기존 초안의 장점을 살리되, 요구사항에 맞게 문장을 재구성합니다.
3. 의료광고 위반 소지 표현은 제거합니다.
4. 주력 키워드 6회 사용, 1,300자 내외, 질문형 소제목 구조를 유지합니다.
5. 마지막에 체크리스트도 함께 제출합니다.

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "titles": ["제목1", "제목2", "제목3", "제목4", "제목5"],
  "body": "본문 초안",
  "outline": ["소제목1", "소제목2", "소제목3", "소제목4"],
  "checklist": [
    "글자 수: 준수 여부 / 실제 글자수",
    "주력 키워드 6회: 준수 여부 / 실제 횟수",
    "반복 형태소 점검: 준수 여부 / 주요 반복 단어",
    "의료법 준수 여부: 준수 여부 / 위반 표현 여부",
    "구조 반영 여부: 준수 여부"
  ]
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, "Gemini 수정")
    titles = payload.get("titles", [])
    body = payload.get("body", "")
    outline = payload.get("outline", [])
    checklist = payload.get("checklist", [])
    if not titles or not body:
        raise RuntimeError("Gemini 수정 응답에 제목 또는 본문이 비어 있습니다.")
    return {"titles": titles[:5], "body": body, "outline": outline, "checklist": checklist}


def generate_translation_with_gemini(
    korean_text: str,
    target_language_label: str,
    additional_request: str,
    seed_keyword: str,
    clinic_name: str,
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")
    if not korean_text.strip():
        raise RuntimeError("번역할 한국어 원고를 입력해 주세요.")

    preset = TRANSLATION_PRESETS[target_language_label]
    prompt_records = load_translation_prompts()
    glossary_records = load_translation_glossaries()
    language_prompt = prompt_records.get(target_language_label, {}).get("text", _default_translation_prompt(target_language_label))
    glossary_terms = glossary_records.get(target_language_label, {}).get("terms", {})
    glossary_text = "\n".join(f"- {src} => {dst}" for src, dst in glossary_terms.items()) or "등록된 용어집 없음"
    prompt = f"""
당신은 병원 블로그 콘텐츠 현지화 전문 번역가입니다.
아래 한국어 원고를 {target_language_label}({preset["locale"]})로 번역하세요.

[번역 목표]
- {preset["style"]}
- {preset["notes"]}
- 병원명 `{clinic_name}` 과 시술명 `{seed_keyword}` 맥락은 유지합니다.
- 정보성 블로그 톤을 유지하되 과장 광고처럼 보이지 않게 번역합니다.

[언어별 등록 프롬프트]
{language_prompt}

[언어별 용어집]
{glossary_text}

[형식 유지 규칙]
1. 원문의 문단 구조를 최대한 유지합니다.
2. `**볼드**` 표시는 그대로 유지합니다.
3. 마크다운 표는 표 형식을 유지합니다.
4. 리스트, 소제목(`##`, `###`)도 유지합니다.
5. 번역문만 출력하고 불필요한 설명은 넣지 않습니다.

[추가 요청]
{additional_request.strip() or "추가 요청 없음"}

[한국어 원문]
{korean_text}

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "translated_body": "번역된 본문",
  "summary_note": "번역 톤 요약 한 줄"
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, f"{target_language_label} 번역")
    translated_body = payload.get("translated_body", "").strip()
    if not translated_body:
        raise RuntimeError("번역 결과가 비어 있습니다.")
    return {
        "translated_body": translated_body,
        "summary_note": payload.get("summary_note", "").strip(),
        "language_label": target_language_label,
        "locale": preset["locale"],
    }


def revise_translation_with_gemini(
    original_translation: str,
    target_language_label: str,
    revision_request: str,
    seed_keyword: str,
    clinic_name: str,
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")
    if not revision_request.strip():
        raise RuntimeError("추가 수정 요청을 입력해 주세요.")

    preset = TRANSLATION_PRESETS[target_language_label]
    prompt_records = load_translation_prompts()
    glossary_records = load_translation_glossaries()
    language_prompt = prompt_records.get(target_language_label, {}).get("text", _default_translation_prompt(target_language_label))
    glossary_terms = glossary_records.get(target_language_label, {}).get("terms", {})
    glossary_text = "\n".join(f"- {src} => {dst}" for src, dst in glossary_terms.items()) or "등록된 용어집 없음"
    prompt = f"""
당신은 병원 블로그 콘텐츠 현지화 전문 번역가입니다.
아래 번역 초안을 {target_language_label}({preset["locale"]}) 기준으로 더 다듬어 주세요.

[문체 기준]
- {preset["style"]}
- {preset["notes"]}
- 병원명 `{clinic_name}` 과 시술명 `{seed_keyword}` 맥락 유지

[언어별 등록 프롬프트]
{language_prompt}

[언어별 용어집]
{glossary_text}

[유지 규칙]
1. `**볼드**`, 표, 리스트, 소제목 구조는 유지합니다.
2. 불필요한 설명은 넣지 말고 번역문만 출력합니다.

[기존 번역 초안]
{original_translation}

[추가 수정 요청]
{revision_request}

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "translated_body": "수정된 번역문",
  "summary_note": "수정 요약 한 줄"
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, f"{target_language_label} 번역 수정")
    translated_body = payload.get("translated_body", "").strip()
    if not translated_body:
        raise RuntimeError("수정된 번역 결과가 비어 있습니다.")
    return {
        "translated_body": translated_body,
        "summary_note": payload.get("summary_note", "").strip(),
        "language_label": target_language_label,
        "locale": preset["locale"],
    }


def generate_website_copy_with_gemini(
    source_text: str,
    target_language_label: str,
    additional_request: str,
    clinic_name: str,
    operator_name: str,
    seed_keyword: str = "",
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")
    if not source_text.strip():
        raise RuntimeError("재가공할 원문을 입력해 주세요.")

    prompt_records = load_user_website_prompts(operator_name)
    language_prompt = prompt_records.get(target_language_label, {}).get("text", _default_website_prompt(target_language_label))
    reference_docs = load_website_reference_knowledge(target_language_label, operator_name)
    reference_doc_text = "\n\n".join(
        f"[참고자료: {item['name']}]\n{item['content']}" for item in reference_docs
    ) or "등록된 참고 자료 없음"
    locale = TRANSLATION_PRESETS.get(target_language_label, {}).get("locale", "ko-KR" if target_language_label == "한국어" else "")

    prompt = f"""
당신은 클리닉 웹사이트 SEO/GEO 콘텐츠 전문 카피라이터입니다.
아래 원문을 바탕으로 `{target_language_label}` 웹사이트 원고로 재가공하세요.

[목표]
- 블로그 원문을 웹사이트 소개/상세/FAQ형 원고로 재구성합니다.
- 구글 SEO/GEO를 고려해 제목과 섹션 구조를 더 또렷하게 정리합니다.
- 의료광고 위반 소지가 있는 과장 표현, 확정 표현은 넣지 않습니다.
- 원문에 없는 정보는 함부로 추가하지 않습니다.

[언어/출력 대상]
- 언어: {target_language_label}
- 로케일: {locale or "기본"}
- 병원명: {clinic_name}
- 연관 시술/키워드: {seed_keyword or "미지정"}

[언어별 웹사이트 프롬프트]
{language_prompt}

[언어별 참고 자료]
{reference_doc_text}

[원문]
{source_text}

[추가 수정 요청]
{additional_request.strip() or "추가 요청 없음"}

[작성 규칙]
1. 제목은 반드시 웹사이트 제목처럼 자연스럽고 완결된 문장으로 씁니다.
2. 제목 후보는 각각 명확한 검색 의도와 핵심 주제를 담아야 하며, 단순 문장 파편이나 소제목처럼 쓰지 않습니다.
3. 제목은 가능한 한 "시술명 + 고민/비교/가격/원리/대상" 구조를 반영합니다.
4. 썸네일 문구는 1문장, 짧고 후킹되게 만듭니다.
5. slug는 영문 소문자와 하이픈만 사용해 3~6단어 수준으로 생성합니다.
6. description은 검색 결과에서 클릭을 유도하는 1~2문장 설명으로 씁니다.
7. 본문은 H1 소개 → 핵심 포인트 → 적합 대상/특징 → FAQ → 안내/CTA 흐름을 기본으로 합니다.
8. 리스트, 표, 볼드가 필요하면 유지하거나 더 보기 좋게 재정리합니다.
9. 복붙해서 바로 사용할 수 있도록 일반 텍스트/마크다운 친화적으로 씁니다.
10. FAQ는 3개 이상 포함해도 좋습니다.
11. CTA는 과장 없이 상담/안내형으로 정리합니다.
12. 한국어 웹사이트 원고일 경우 주력 키워드는 가능하면 정확히 6회 반복하고 **볼드체**로 표시합니다.
13. 첫 문장과 마지막 문장 고정 규칙이 프롬프트에 있으면 우선 적용합니다.

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "titles": ["제목1", "제목2", "제목3", "제목4", "제목5"],
  "thumbnail_text": "썸네일 문구",
  "slug": "seo-friendly-slug",
  "description": "클릭 버튼/메타용 설명 문구",
  "body": "웹사이트 원고 본문",
  "outline": ["섹션1", "섹션2", "섹션3", "섹션4"],
  "checklist": [
    "언어 적합성: 준수 여부 / 적용 언어",
    "SEO/GEO 구조: 준수 여부 / 핵심 섹션",
    "의료법 준수 여부: 준수 여부 / 위반 표현 여부",
    "참고 자료 반영 여부: 준수 여부 / 반영 자료",
    "복붙 사용성: 준수 여부 / 표·볼드·목록 정리 상태"
  ]
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, "웹사이트 원고")
    titles = payload.get("titles", [])
    thumbnail_text = payload.get("thumbnail_text", "").strip()
    slug = payload.get("slug", "").strip()
    description = payload.get("description", "").strip()
    body = payload.get("body", "")
    outline = payload.get("outline", [])
    checklist = payload.get("checklist", [])
    if not titles or not body:
        raise RuntimeError("웹사이트 원고 응답에 제목 또는 본문이 비어 있습니다.")
    return {
        "titles": titles[:5],
        "thumbnail_text": thumbnail_text,
        "slug": slug,
        "description": description,
        "body": body,
        "outline": outline,
        "checklist": checklist,
        "language_label": target_language_label,
        "locale": locale,
    }


def revise_website_copy_with_gemini(
    original_draft: dict[str, Any],
    revision_request: str,
    target_language_label: str,
    clinic_name: str,
    operator_name: str,
    seed_keyword: str = "",
) -> dict[str, Any]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY 가 없습니다. .env 파일에 설정해 주세요.")
    if not revision_request.strip():
        raise RuntimeError("추가 수정 요청을 입력해 주세요.")

    prompt_records = load_user_website_prompts(operator_name)
    language_prompt = prompt_records.get(target_language_label, {}).get("text", _default_website_prompt(target_language_label))
    reference_docs = load_website_reference_knowledge(target_language_label, operator_name)
    reference_doc_text = "\n\n".join(
        f"[참고자료: {item['name']}]\n{item['content']}" for item in reference_docs
    ) or "등록된 참고 자료 없음"
    locale = TRANSLATION_PRESETS.get(target_language_label, {}).get("locale", "ko-KR" if target_language_label == "한국어" else "")

    prompt = f"""
당신은 클리닉 웹사이트 SEO/GEO 콘텐츠 전문 카피라이터입니다.
아래 1차 웹사이트 원고를 바탕으로 추가 수정 요청을 반영한 2차 원고를 작성하세요.

[언어]
- 언어: {target_language_label}
- 로케일: {locale or "기본"}
- 병원명: {clinic_name}
- 연관 시술/키워드: {seed_keyword or "미지정"}

[언어별 웹사이트 프롬프트]
{language_prompt}

[언어별 참고 자료]
{reference_doc_text}

[기존 제목 후보]
{chr(10).join(f"- {title}" for title in original_draft.get("titles", []))}

[기존 섹션 구조]
{chr(10).join(f"- {item}" for item in original_draft.get("outline", []))}

[기존 본문]
{original_draft.get("body", "")}

[추가 수정 요청]
{revision_request}

[수정 규칙]
1. 사용자의 요청을 최우선 반영합니다.
2. 웹사이트 소개/상세 페이지 톤을 유지합니다.
3. 구조와 가독성, 검색 친화성을 더 좋게 정리합니다.
4. 의료광고 위반 소지가 있는 표현은 제거합니다.
5. 제목 후보 5개, 썸네일 문구, slug, description도 함께 다시 다듬습니다.

반드시 아래 JSON 형식으로만 답변하세요.
{{
  "titles": ["제목1", "제목2", "제목3", "제목4", "제목5"],
  "thumbnail_text": "썸네일 문구",
  "slug": "seo-friendly-slug",
  "description": "클릭 버튼/메타용 설명 문구",
  "body": "웹사이트 원고 본문",
  "outline": ["섹션1", "섹션2", "섹션3", "섹션4"],
  "checklist": [
    "언어 적합성: 준수 여부 / 적용 언어",
    "SEO/GEO 구조: 준수 여부 / 핵심 섹션",
    "의료법 준수 여부: 준수 여부 / 위반 표현 여부",
    "참고 자료 반영 여부: 준수 여부 / 반영 자료",
    "복붙 사용성: 준수 여부 / 표·볼드·목록 정리 상태"
  ]
}}
""".strip()

    payload = generate_gemini_json_response(None, prompt, "웹사이트 원고 수정")
    titles = payload.get("titles", [])
    thumbnail_text = payload.get("thumbnail_text", "").strip()
    slug = payload.get("slug", "").strip()
    description = payload.get("description", "").strip()
    body = payload.get("body", "")
    outline = payload.get("outline", [])
    checklist = payload.get("checklist", [])
    if not titles or not body:
        raise RuntimeError("웹사이트 수정 원고 응답에 제목 또는 본문이 비어 있습니다.")
    return {
        "titles": titles[:5],
        "thumbnail_text": thumbnail_text,
        "slug": slug,
        "description": description,
        "body": body,
        "outline": outline,
        "checklist": checklist,
        "language_label": target_language_label,
        "locale": locale,
    }


def render_keyword_table(candidate_dicts: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for item in candidate_dicts:
        row = SearchAdKeyword(**item)
        rows.append(
            {
                "키워드": row.keyword,
                "모수 키워드": row.source_seed,
                "월간 검색수": row.monthly_total,
                "PC": row.monthly_pc,
                "모바일": row.monthly_mobile,
                "문서 수(검색결과 추정)": row.monthly_docs,
                "포화도": round(row.saturation, 2),
                "추천 등급": recommendation_grade(
                    saturation=row.saturation,
                    competition_index=row.competition_index,
                    is_fallback=row.is_fallback,
                ),
                "경쟁 강도": row.competition_index,
                "기회 점수": round(row.opportunity_score, 2),
                "데이터 구분": (
                    "내부 시트"
                    if row.source_type == "workbook"
                    else "추정 후보"
                    if row.is_fallback
                    else "API 기반"
                ),
                "키워드 분류": row.keyword_classification,
                "추천 이유": row.recommendation_reason,
            }
        )
    return pd.DataFrame(rows)


def style_keyword_table(df: pd.DataFrame):
    def grade_color(value: Any) -> str:
        text = str(value).strip()
        if text == "추천":
            return "color: #d62828; font-weight: 700;"
        if text in {"비추천", "주의"}:
            return "color: #1d4ed8; font-weight: 700;"
        if text == "테스트 추천":
            return "color: #b45309; font-weight: 700;"
        if text == "추정 후보":
            return "color: #6b7280; font-weight: 600;"
        return ""

    if "추천 등급" in df.columns:
        return df.style.map(grade_color, subset=["추천 등급"])
    return df.style


def flatten_grouped_candidates(grouped_candidates: dict[str, list[dict[str, Any]]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for seed_keyword, items in grouped_candidates.items():
        if not items:
            continue
        df = render_keyword_table(items)
        df.insert(0, "주력 시술", seed_keyword)
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def ensure_session_defaults() -> None:
    st.session_state.setdefault("recommended_keywords_df", None)
    st.session_state.setdefault("grouped_candidates", None)
    st.session_state.setdefault("analysis_summary", None)
    st.session_state.setdefault("generated_draft", None)
    st.session_state.setdefault("generated_draft_context", None)
    st.session_state.setdefault("current_page", "home")
    st.session_state.setdefault("phase3_direct_seed_override", "")
    st.session_state.setdefault("current_operator", "young")
    st.session_state.setdefault("is_authenticated", False)
    st.session_state.setdefault("auth_user", "")
    st.session_state.setdefault("home_auth_mode", "")
    st.session_state.setdefault("post_login_page", "home")
    st.session_state.setdefault("translated_draft", None)
    st.session_state.setdefault("translated_draft_context", None)
    st.session_state.setdefault("website_generated_draft", None)
    st.session_state.setdefault("website_generated_context", None)


def go_to_page(page_name: str) -> None:
    st.session_state.current_page = page_name


def current_operator_name() -> str:
    raw = str(st.session_state.get("current_operator", "young")).strip()
    return raw or "young"


def parse_llm_json_response(text: str, response_label: str) -> dict[str, Any]:
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise RuntimeError(f"{response_label} 응답을 JSON으로 해석하지 못했습니다.")

    raw_json = match.group(0).strip()
    try:
        return json.loads(raw_json, strict=False)
    except json.JSONDecodeError:
        cleaned = raw_json.replace("\r\n", "\n").replace("\r", "\n")
        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", cleaned)
        return json.loads(cleaned, strict=False)


def operator_slug(operator_name: str | None = None) -> str:
    text = str(operator_name or current_operator_name()).strip().lower()
    slug = re.sub(r"[^0-9a-zA-Z가-힣_-]+", "_", text)
    slug = slug.strip("_")
    return slug or "young"


def operator_data_dir(operator_name: str | None = None) -> Path:
    return USER_DATA_ROOT / operator_slug(operator_name)


def operator_profile_path(operator_name: str | None = None) -> Path:
    return operator_data_dir(operator_name) / "procedure_profiles.json"


def operator_source_dir(operator_name: str | None = None) -> Path:
    return operator_data_dir(operator_name) / "procedure_sources"


def storage_slug(text: str) -> str:
    slug = re.sub(r"[^0-9a-zA-Z가-힣_-]+", "_", str(text).strip().lower())
    slug = slug.strip("_")
    return slug or "default"


def operator_website_prompt_path(operator_name: str | None = None) -> Path:
    return operator_data_dir(operator_name) / "website_prompts.json"


def operator_website_reference_dir(language_label: str, operator_name: str | None = None) -> Path:
    return operator_data_dir(operator_name) / "website_reference_files" / storage_slug(language_label)


def _default_website_prompt(language_label: str) -> str:
    if language_label in DEFAULT_WEBSITE_PROMPTS:
        return DEFAULT_WEBSITE_PROMPTS[language_label]
    if language_label in TRANSLATION_PRESETS:
        preset = TRANSLATION_PRESETS[language_label]
        return (
            f"{language_label}({preset['locale']}) 웹사이트 원고 프롬프트\n"
            f"- {preset['style']}\n"
            "- 블로그보다 정리된 웹사이트 소개 문체\n"
            "- 과장 광고 문구 금지\n"
            "- H1/H2/H3 구조와 FAQ 반영"
        )
    return "웹사이트 SEO/GEO 원고를 위한 기본 프롬프트"


def load_user_website_prompts(operator_name: str | None = None) -> dict[str, dict[str, Any]]:
    prompt_path = operator_website_prompt_path(operator_name)
    if not prompt_path.exists():
        return {}
    try:
        return json.loads(prompt_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_user_website_prompt(language_label: str, text: str, operator_name: str | None = None) -> None:
    prompt_path = operator_website_prompt_path(operator_name)
    prompt_path.parent.mkdir(parents=True, exist_ok=True)
    prompts = load_user_website_prompts(operator_name)
    prompts[language_label] = {
        "text": text.strip(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": operator_name or current_operator_name(),
    }
    prompt_path.write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")


def delete_user_website_prompt(language_label: str, operator_name: str | None = None) -> None:
    prompts = load_user_website_prompts(operator_name)
    if language_label in prompts:
        del prompts[language_label]
        prompt_path = operator_website_prompt_path(operator_name)
        prompt_path.parent.mkdir(parents=True, exist_ok=True)
        prompt_path.write_text(json.dumps(prompts, ensure_ascii=False, indent=2), encoding="utf-8")


def list_user_website_reference_files(language_label: str, operator_name: str | None = None) -> list[Path]:
    directory = operator_website_reference_dir(language_label, operator_name)
    if not directory.exists():
        return []
    return sorted([path for path in directory.iterdir() if path.is_file()])


def save_uploaded_website_reference_files(language_label: str, uploaded_files: list[Any], operator_name: str | None = None) -> None:
    directory = operator_website_reference_dir(language_label, operator_name)
    directory.mkdir(parents=True, exist_ok=True)
    for uploaded in uploaded_files:
        target = directory / uploaded.name
        target.write_bytes(uploaded.getbuffer())


def delete_user_website_reference_file(language_label: str, filename: str, operator_name: str | None = None) -> None:
    target = operator_website_reference_dir(language_label, operator_name) / filename
    if target.exists():
        target.unlink()


def load_website_reference_knowledge(language_label: str, operator_name: str | None = None) -> list[dict[str, str]]:
    reference_docs: list[dict[str, str]] = []
    for seed_path in WEBSITE_REFERENCE_SEED_FILES.get(language_label, []):
        content = read_rich_text_document(str(seed_path))
        if not content:
            continue
        reference_docs.append(
            {
                "name": f"기본 자료 - {seed_path.stem}",
                "content": compress_prompt_text(content, max_chars=3500),
            }
        )
    for file_path in list_user_website_reference_files(language_label, operator_name):
        if file_path.suffix.lower() not in {".docx", ".doc", ".txt", ".md"}:
            continue
        content = read_rich_text_document(str(file_path))
        if not content:
            continue
        reference_docs.append(
            {
                "name": file_path.stem,
                "content": compress_prompt_text(content, max_chars=3500),
            }
        )
    return reference_docs


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _auth_signature(username: str) -> str:
    return hmac.new(APP_AUTH_SECRET.encode("utf-8"), username.encode("utf-8"), hashlib.sha256).hexdigest()


def build_auth_token(username: str) -> str:
    payload = f"{username}|{_auth_signature(username)}"
    return base64.urlsafe_b64encode(payload.encode("utf-8")).decode("utf-8")


def decode_auth_token(token: str) -> str:
    try:
        raw = base64.urlsafe_b64decode(token.encode("utf-8")).decode("utf-8")
        username, signature = raw.split("|", 1)
    except Exception:
        return ""
    if not username or signature != _auth_signature(username):
        return ""
    if username not in load_user_accounts():
        return ""
    return username


def persist_login(username: str) -> None:
    st.session_state.is_authenticated = True
    st.session_state.auth_user = username
    st.session_state.current_operator = username
    st.query_params["session"] = build_auth_token(username)


def clear_persisted_login() -> None:
    st.session_state.is_authenticated = False
    st.session_state.auth_user = ""
    st.session_state.current_operator = "young"
    if "session" in st.query_params:
        del st.query_params["session"]


def load_user_accounts() -> dict[str, dict[str, Any]]:
    if not USER_ACCOUNT_PATH.exists():
        return {}
    try:
        accounts = json.loads(USER_ACCOUNT_PATH.read_text(encoding="utf-8"))
        changed = False
        for username, account in accounts.items():
            if "role" not in account:
                account["role"] = "superadmin" if username == "youngeun" else "user"
                changed = True
            if "managed_languages" not in account:
                if account.get("role") == "superadmin" or username == "youngeun":
                    account["managed_languages"] = list(TRANSLATION_PRESETS.keys())
                else:
                    account["managed_languages"] = []
                changed = True
        if changed:
            save_user_accounts(accounts)
        return accounts
    except Exception:
        return {}


def save_user_accounts(accounts: dict[str, dict[str, Any]]) -> None:
    USER_ACCOUNT_PATH.parent.mkdir(parents=True, exist_ok=True)
    USER_ACCOUNT_PATH.write_text(json.dumps(accounts, ensure_ascii=False, indent=2), encoding="utf-8")


def create_user_account(username: str, password: str, display_name: str = "") -> None:
    accounts = load_user_accounts()
    role = "superadmin" if username == "youngeun" else "user"
    accounts[username] = {
        "password_hash": hash_password(password),
        "display_name": display_name or username,
        "role": role,
        "managed_languages": list(TRANSLATION_PRESETS.keys()) if role == "superadmin" else [],
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    save_user_accounts(accounts)


def verify_user_login(username: str, password: str) -> bool:
    accounts = load_user_accounts()
    account = accounts.get(username)
    if not account:
        return False
    return account.get("password_hash") == hash_password(password)


def is_superadmin(username: str | None = None) -> bool:
    user = (username or st.session_state.get("auth_user", "")).strip()
    if not user:
        return False
    account = load_user_accounts().get(user, {})
    return account.get("role") == "superadmin" or user == "youngeun"


def update_user_password(username: str, new_password: str) -> None:
    accounts = load_user_accounts()
    if username not in accounts:
        return
    accounts[username]["password_hash"] = hash_password(new_password)
    accounts[username]["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_user_accounts(accounts)


def get_user_managed_languages(username: str | None = None) -> list[str]:
    user = (username or st.session_state.get("auth_user", "")).strip()
    if not user:
        return []
    if is_superadmin(user):
        return list(TRANSLATION_PRESETS.keys())
    account = load_user_accounts().get(user, {})
    languages = account.get("managed_languages", [])
    normalized = [lang for lang in languages if lang in TRANSLATION_PRESETS]
    if not normalized:
        return list(TRANSLATION_PRESETS.keys())
    return normalized


def can_manage_translation_language(language_label: str, username: str | None = None) -> bool:
    return language_label in get_user_managed_languages(username)


def update_user_managed_languages(username: str, languages: list[str]) -> None:
    accounts = load_user_accounts()
    if username not in accounts:
        return
    accounts[username]["managed_languages"] = [lang for lang in languages if lang in TRANSLATION_PRESETS]
    accounts[username]["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_user_accounts(accounts)


def render_login_screen() -> None:
    st.title("🎰 블로그 가챠 (Blog Gacha)")
    st.caption("작업자별 프리셋과 참고 파일을 분리해 관리하려면 먼저 로그인해 주세요.")
    mode = st.session_state.get("home_auth_mode", "login") or "login"
    target_page = st.session_state.get("post_login_page", "home")

    accounts = load_user_accounts()
    if not accounts:
        st.info("처음 사용하는 환경입니다. 첫 작업자 계정을 먼저 만들어 주세요.")
        init_cols = st.columns([1.15, 0.85])
        with init_cols[0]:
            username = st.text_input("작업자 아이디", key="init_username")
            display_name = st.text_input("표시 이름", key="init_display_name", placeholder="예: young, 마케팅팀A")
            password = st.text_input("비밀번호", type="password", key="init_password")
            confirm = st.text_input("비밀번호 확인", type="password", key="init_password_confirm")
            if st.button("첫 계정 만들기", type="primary", use_container_width=True):
                if not username.strip() or not password.strip():
                    st.error("아이디와 비밀번호를 입력해 주세요.")
                elif password != confirm:
                    st.error("비밀번호 확인이 일치하지 않습니다.")
                else:
                    create_user_account(username.strip(), password.strip(), display_name.strip())
                    persist_login(username.strip())
                    st.session_state.current_page = target_page
                    st.success("첫 계정을 만들었습니다. 바로 대시보드로 들어갑니다.")
                    st.rerun()
        with init_cols[1]:
            st.markdown(
                """
                <div class="app-card">
                    <div class="stage-title">왜 로그인하나요?</div>
                    <div class="stage-help">
                        작업자별로 시술 프리셋, 참고 파일, 초안 흐름이 섞이지 않도록 분리 저장합니다.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
        )
        return

    auth_cols = st.columns([1.05, 0.95])
    if mode == "login":
        with auth_cols[0]:
            username = st.text_input("작업자 아이디", key="login_username")
            password = st.text_input("비밀번호", type="password", key="login_password")
            if st.button("로그인", type="primary", use_container_width=True):
                if verify_user_login(username.strip(), password.strip()):
                    persist_login(username.strip())
                    st.session_state.current_page = target_page
                    st.success(f"{username.strip()} 계정으로 로그인했습니다.")
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호가 맞지 않습니다.")
            if st.button("회원가입으로 이동", use_container_width=True, key="go_signup_from_login"):
                st.session_state.home_auth_mode = "signup"
                st.rerun()
        with auth_cols[1]:
            account_names = ", ".join(load_user_accounts().keys())
            st.markdown(
                f"""
                <div class="app-card">
                    <div class="stage-title">사내 작업자 로그인</div>
                    <div class="stage-help">
                        로그인하면 작업자별 프리셋과 참고 자료가 자동으로 분리됩니다.<br/><br/>
                        현재 등록된 작업자: {account_names}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
    else:
        with auth_cols[0]:
            username = st.text_input("새 작업자 아이디", key="signup_username")
            display_name = st.text_input("표시 이름", key="signup_display_name")
            password = st.text_input("비밀번호", type="password", key="signup_password")
            confirm = st.text_input("비밀번호 확인", type="password", key="signup_confirm")
            if st.button("회원가입", type="primary", use_container_width=True, key="signup_submit"):
                if not username.strip() or not password.strip():
                    st.error("아이디와 비밀번호를 입력해 주세요.")
                elif password != confirm:
                    st.error("비밀번호 확인이 일치하지 않습니다.")
                elif username.strip() in accounts:
                    st.error("이미 존재하는 작업자 아이디입니다.")
                else:
                    create_user_account(username.strip(), password.strip(), display_name.strip())
                    persist_login(username.strip())
                    st.session_state.current_page = target_page
                    st.success(f"{username.strip()} 계정을 만들고 로그인했습니다.")
                    st.rerun()
            if st.button("로그인으로 이동", use_container_width=True, key="go_login_from_signup"):
                st.session_state.home_auth_mode = "login"
                st.rerun()
        with auth_cols[1]:
            st.markdown(
                """
                <div class="app-card">
                    <div class="stage-title">왜 회원가입이 필요한가요?</div>
                    <div class="stage-help">
                        작업자별로 시술 프리셋, 참고 파일, 초안 흐름을 안전하게 분리 저장합니다.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def require_auth_before_phase(page_name: str) -> bool:
    if st.session_state.get("is_authenticated", False):
        return True
    st.session_state.post_login_page = page_name
    st.session_state.home_auth_mode = "login"
    st.session_state.current_page = "auth"
    return False


def scroll_viewport_to_top() -> None:
    components.html(
        """
        <script>
        window.parent.scrollTo({top: 0, behavior: "instant"});
        </script>
        """,
        height=0,
    )


def render_phase_footer_nav(*pages: tuple[str, str]) -> None:
    nav_cols = st.columns(len(pages))
    for idx, (label, page_name) in enumerate(pages):
        if nav_cols[idx].button(label, use_container_width=True, key=f"nav_{page_name}_{label}"):
            go_to_page(page_name)
            st.rerun()


def render_phase_top_nav(*pages: tuple[str, str]) -> None:
    nav_cols = st.columns(len(pages))
    for idx, (label, page_name) in enumerate(pages):
        if nav_cols[idx].button(label, use_container_width=True, key=f"top_nav_{page_name}_{label}"):
            go_to_page(page_name)
            st.rerun()


def _apply_inline_blog_formatting(text: str) -> str:
    escaped = html.escape(text)
    escaped = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", escaped)
    escaped = re.sub(r"`([^`]+)`", r"<code>\1</code>", escaped)
    return escaped


def _is_markdown_table_separator(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("|") and all(ch in "|:- " for ch in stripped)


def _parse_markdown_table(table_lines: list[str]) -> str:
    if len(table_lines) < 2:
        return ""

    headers = [cell.strip() for cell in table_lines[0].strip().strip("|").split("|")]
    body_rows: list[list[str]] = []
    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) < len(headers):
            cells.extend([""] * (len(headers) - len(cells)))
        elif len(cells) > len(headers):
            cells = cells[: len(headers)]
        body_rows.append(cells)

    thead = "".join(f"<th>{_apply_inline_blog_formatting(header)}</th>" for header in headers)
    tbody = "".join(
        "<tr>" + "".join(f"<td>{_apply_inline_blog_formatting(cell)}</td>" for cell in row) + "</tr>"
        for row in body_rows
    )
    return f'<div class="draft-table-wrap"><table class="draft-table"><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table></div>'


def render_blog_body_preview(body_text: str) -> None:
    lines = body_text.splitlines()
    html_parts: list[str] = []
    idx = 0

    while idx < len(lines):
        line = lines[idx].rstrip()
        stripped = line.strip()

        if not stripped:
            idx += 1
            continue

        if idx + 1 < len(lines) and stripped.startswith("|") and _is_markdown_table_separator(lines[idx + 1]):
            table_lines = [line]
            idx += 1
            while idx < len(lines):
                current = lines[idx].rstrip()
                if current.strip().startswith("|"):
                    table_lines.append(current)
                    idx += 1
                    continue
                break
            html_parts.append(_parse_markdown_table(table_lines))
            continue

        if stripped.startswith("### "):
            html_parts.append(f"<h3>{_apply_inline_blog_formatting(stripped[4:])}</h3>")
            idx += 1
            continue
        if stripped.startswith("## "):
            html_parts.append(f"<h2>{_apply_inline_blog_formatting(stripped[3:])}</h2>")
            idx += 1
            continue
        if stripped.startswith("# "):
            html_parts.append(f"<h1>{_apply_inline_blog_formatting(stripped[2:])}</h1>")
            idx += 1
            continue

        if stripped.startswith(("- ", "* ")):
            items: list[str] = []
            while idx < len(lines) and lines[idx].strip().startswith(("- ", "* ")):
                items.append(lines[idx].strip()[2:])
                idx += 1
            html_parts.append(
                "<ul>"
                + "".join(f"<li>{_apply_inline_blog_formatting(item)}</li>" for item in items)
                + "</ul>"
            )
            continue

        if re.match(r"^\d+\.\s+", stripped):
            items = []
            while idx < len(lines) and re.match(r"^\d+\.\s+", lines[idx].strip()):
                items.append(re.sub(r"^\d+\.\s+", "", lines[idx].strip()))
                idx += 1
            html_parts.append(
                "<ol>"
                + "".join(f"<li>{_apply_inline_blog_formatting(item)}</li>" for item in items)
                + "</ol>"
            )
            continue

        paragraph_lines = [stripped]
        idx += 1
        while idx < len(lines):
            next_stripped = lines[idx].strip()
            if not next_stripped:
                idx += 1
                break
            if (
                next_stripped.startswith(("# ", "## ", "### ", "- ", "* ", "|"))
                or re.match(r"^\d+\.\s+", next_stripped)
            ):
                break
            paragraph_lines.append(next_stripped)
            idx += 1
        paragraph_html = "<br/>".join(_apply_inline_blog_formatting(item) for item in paragraph_lines)
        html_parts.append(f"<p>{paragraph_html}</p>")

    rendered_html = "".join(html_parts)
    st.markdown(f'<div class="draft-rendered">{rendered_html}</div>', unsafe_allow_html=True)


def render_checklist_preview(items: list[str]) -> None:
    rows: list[tuple[str, str]] = []
    for raw in items:
        text = str(raw).strip()
        if not text:
            continue
        if ":" in text:
            label, value = text.split(":", 1)
            rows.append((label.strip(), value.strip()))
        else:
            rows.append(("체크 항목", text))

    if not rows:
        return

    html_rows = "".join(
        (
            f'<div class="checklist-row">'
            f'<div class="checklist-label">{html.escape(label)}</div>'
            f'<div class="checklist-value">{html.escape(value)}</div>'
            f"</div>"
        )
        for label, value in rows
    )
    st.markdown(f'<div class="checklist-card">{html_rows}</div>', unsafe_allow_html=True)


def infer_seed_keyword_from_summary(summary: dict[str, Any], keyword_df: pd.DataFrame | None) -> str:
    if keyword_df is None or keyword_df.empty:
        return ""
    matches = keyword_df[keyword_df["키워드"] == summary.get("keyword", "")]
    if not matches.empty:
        return str(matches.iloc[0]["주력 시술"])
    return str(keyword_df.iloc[0]["주력 시술"])


def render_custom_css() -> None:
    st.markdown(
        """
        <style>
        .main {
            background:
                radial-gradient(circle at top right, rgba(44, 122, 123, 0.12), transparent 28%),
                linear-gradient(180deg, #f8faf7 0%, #f3f6f4 100%);
        }
        .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
            max-width: 1200px;
        }
        .app-card {
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 20px;
            padding: 1.2rem 1.25rem;
            box-shadow: 0 12px 40px rgba(15, 23, 42, 0.05);
            margin-bottom: 1rem;
        }
        .stage-title {
            font-size: 1.1rem;
            font-weight: 700;
            margin-bottom: 0.3rem;
            color: #12343b;
        }
        .stage-help {
            color: #51606b;
            font-size: 0.93rem;
        }
        .diagnosis-hero {
            background: linear-gradient(135deg, rgba(19, 52, 59, 0.96), rgba(44, 122, 123, 0.94));
            color: white;
            border-radius: 22px;
            padding: 1.2rem 1.3rem;
            margin: 0.4rem 0 1rem 0;
            box-shadow: 0 14px 36px rgba(18, 52, 59, 0.18);
        }
        .diagnosis-hero-title {
            font-size: 0.82rem;
            opacity: 0.82;
            margin-bottom: 0.35rem;
            letter-spacing: 0.02em;
        }
        .diagnosis-hero-main {
            font-size: 1.18rem;
            font-weight: 800;
            line-height: 1.45;
        }
        .diagnosis-card {
            background: rgba(255,255,255,0.9);
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 18px;
            padding: 1rem 1.05rem;
            min-height: 190px;
            box-shadow: 0 10px 30px rgba(15, 23, 42, 0.04);
        }
        .diagnosis-label {
            font-size: 0.82rem;
            color: #5b6b74;
            margin-bottom: 0.45rem;
            font-weight: 700;
        }
        .diagnosis-list {
            margin: 0;
            padding-left: 1rem;
            color: #23323a;
            line-height: 1.8;
            font-size: 0.96rem;
        }
        .diagnosis-badge {
            display: inline-block;
            padding: 0.18rem 0.55rem;
            border-radius: 999px;
            font-size: 0.83rem;
            font-weight: 700;
            margin-left: 0.35rem;
        }
        .diagnosis-badge.good {
            background: #e8f7ee;
            color: #18794e;
        }
        .diagnosis-badge.warn {
            background: #fff4db;
            color: #a16207;
        }
        .diagnosis-badge.cold {
            background: #eaf2ff;
            color: #1d4ed8;
        }
        .draft-rendered {
            background: rgba(255,255,255,0.9);
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 18px;
            padding: 1.15rem 1.2rem;
            box-shadow: 0 10px 30px rgba(15, 23, 42, 0.04);
        }
        .draft-rendered h1,
        .draft-rendered h2,
        .draft-rendered h3 {
            color: #12343b;
            margin: 1.1rem 0 0.65rem 0;
            line-height: 1.4;
        }
        .draft-rendered h1 { font-size: 1.35rem; }
        .draft-rendered h2 { font-size: 1.2rem; }
        .draft-rendered h3 { font-size: 1.08rem; }
        .draft-rendered p,
        .draft-rendered li {
            color: #243640;
            line-height: 1.85;
            font-size: 0.99rem;
        }
        .draft-rendered strong {
            color: #132f38;
            font-weight: 800;
        }
        .draft-rendered code {
            background: rgba(15, 23, 42, 0.06);
            padding: 0.12rem 0.32rem;
            border-radius: 6px;
            font-size: 0.92em;
        }
        .draft-rendered ul,
        .draft-rendered ol {
            padding-left: 1.2rem;
            margin: 0.5rem 0 1rem 0;
        }
        .draft-table-wrap {
            overflow-x: auto;
            margin: 0.9rem 0 1rem 0;
        }
        .draft-table {
            width: 100%;
            border-collapse: collapse;
            min-width: 620px;
            background: white;
            border-radius: 14px;
            overflow: hidden;
            border: 1px solid rgba(15, 23, 42, 0.09);
        }
        .draft-table th {
            background: #edf5ff;
            color: #12343b;
            font-weight: 800;
            text-align: left;
            padding: 0.72rem 0.8rem;
            border-bottom: 1px solid rgba(15, 23, 42, 0.09);
        }
        .draft-table td {
            padding: 0.72rem 0.8rem;
            border-bottom: 1px solid rgba(15, 23, 42, 0.07);
            color: #243640;
            vertical-align: top;
            line-height: 1.7;
        }
        .draft-table tbody tr:last-child td {
            border-bottom: none;
        }
        .checklist-card {
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 18px;
            overflow: hidden;
            box-shadow: 0 10px 30px rgba(15, 23, 42, 0.04);
            margin-bottom: 1rem;
        }
        .checklist-row {
            display: grid;
            grid-template-columns: 240px 1fr;
            gap: 1rem;
            padding: 0.9rem 1rem;
            border-bottom: 1px solid rgba(15, 23, 42, 0.07);
            align-items: start;
        }
        .checklist-row:last-child {
            border-bottom: none;
        }
        .checklist-label {
            font-weight: 800;
            color: #12343b;
            font-size: 0.95rem;
        }
        .checklist-value {
            color: #33424b;
            line-height: 1.7;
            font-size: 0.95rem;
        }
        .home-shell {
            background: linear-gradient(180deg, #151827 0%, #1d2133 100%);
            border-radius: 32px;
            padding: 1rem 1rem 1.4rem 1rem;
            box-shadow: 0 28px 70px rgba(15, 23, 42, 0.18);
            margin-bottom: 1.4rem;
        }
        .home-topbar {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 1rem;
            padding: 0.15rem 0.3rem 1rem 0.3rem;
            color: rgba(255,255,255,0.88);
            font-size: 0.92rem;
        }
        .home-topbar-brand {
            font-size: 1.05rem;
            font-weight: 800;
            letter-spacing: 0.02em;
        }
        .home-topbar-menu {
            display: flex;
            gap: 1.1rem;
            opacity: 0.84;
            flex-wrap: wrap;
            justify-content: flex-end;
            align-items: center;
        }
        .home-topbar-auth {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-width: 88px;
            padding: 0.55rem 0.9rem;
            border-radius: 999px;
            background: rgba(255,255,255,0.08);
            border: 1px solid rgba(255,255,255,0.18);
            color: white;
            font-size: 0.85rem;
            font-weight: 700;
        }
        .home-hero {
            background:
                radial-gradient(circle at 20% 20%, rgba(255,255,255,0.22), transparent 22%),
                radial-gradient(circle at 80% 25%, rgba(255,255,255,0.14), transparent 18%),
                linear-gradient(135deg, #7c3aed 0%, #a855f7 22%, #74b8ff 66%, #d9ebff 100%);
            color: white;
            border-radius: 28px;
            padding: 2rem 2rem 1.7rem 2rem;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.2);
            margin-bottom: 1.2rem;
        }
        .home-hero-grid {
            display: grid;
            grid-template-columns: 1.05fr 0.95fr;
            gap: 1.6rem;
            align-items: center;
        }
        .home-logo-badge {
            display: inline-flex;
            align-items: center;
            gap: 0.45rem;
            padding: 0.45rem 0.8rem;
            border-radius: 999px;
            background: rgba(32, 13, 69, 0.28);
            font-size: 0.84rem;
            font-weight: 700;
            margin-bottom: 1rem;
            backdrop-filter: blur(10px);
        }
        .home-hero-title {
            font-size: 3rem;
            font-weight: 900;
            margin-bottom: 0.65rem;
            line-height: 1.14;
            letter-spacing: -0.03em;
            text-shadow: 0 10px 30px rgba(68, 18, 112, 0.22);
        }
        .home-hero-sub {
            font-size: 1.02rem;
            line-height: 1.8;
            opacity: 0.95;
            max-width: 540px;
        }
        .home-hero-actions {
            display: flex;
            gap: 0.85rem;
            margin-top: 1.4rem;
            flex-wrap: wrap;
        }
        .home-action-chip {
            display: inline-flex;
            align-items: center;
            padding: 0.8rem 1.15rem;
            border-radius: 999px;
            background: rgba(255,255,255,0.16);
            border: 1px solid rgba(255,255,255,0.22);
            font-size: 0.9rem;
            font-weight: 700;
            backdrop-filter: blur(10px);
        }
        .home-hero-visual {
            min-height: 360px;
            border-radius: 28px;
            background:
                radial-gradient(circle at 20% 22%, rgba(255,255,255,0.42), transparent 10%),
                radial-gradient(circle at 72% 18%, rgba(255,255,255,0.32), transparent 8%),
                radial-gradient(circle at 80% 70%, rgba(255,255,255,0.22), transparent 10%),
                linear-gradient(180deg, rgba(255,255,255,0.32), rgba(255,255,255,0.08));
            border: 1px solid rgba(255,255,255,0.18);
            position: relative;
            overflow: hidden;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.22);
        }
        .home-visual-machine {
            position: absolute;
            right: 8%;
            top: 8%;
            width: 58%;
            height: 78%;
            border-radius: 32px;
            background: linear-gradient(180deg, #ffb6ef 0%, #ef79d1 45%, #c35cd2 100%);
            box-shadow: 0 26px 60px rgba(100, 39, 122, 0.28);
            transform: rotate(-9deg);
        }
        .home-visual-machine::before {
            content: "Blog Gacha";
            position: absolute;
            top: 7%;
            left: 50%;
            transform: translateX(-50%);
            width: 68%;
            text-align: center;
            padding: 0.8rem 0.6rem;
            border-radius: 22px;
            background: rgba(255,255,255,0.24);
            color: #6a1b9a;
            font-size: 1.35rem;
            font-weight: 900;
            letter-spacing: -0.03em;
        }
        .home-visual-machine::after {
            content: "";
            position: absolute;
            left: 18%;
            right: 18%;
            top: 26%;
            height: 36%;
            border-radius: 22px;
            background: rgba(255,255,255,0.34);
            box-shadow: inset 0 0 0 1px rgba(255,255,255,0.26);
        }
        .home-visual-ball {
            position: absolute;
            border-radius: 999px;
            box-shadow: inset -10px -10px 18px rgba(0,0,0,0.08), 0 10px 20px rgba(68, 18, 112, 0.18);
        }
        .home-visual-ball.one {
            width: 88px;
            height: 88px;
            left: 12%;
            bottom: 16%;
            background: linear-gradient(180deg, #f7f1ff 0%, #c6e0ff 100%);
        }
        .home-visual-ball.two {
            width: 76px;
            height: 76px;
            left: 30%;
            bottom: 28%;
            background: linear-gradient(180deg, #ffe7fb 0%, #ffc8e8 100%);
        }
        .home-visual-ball.three {
            width: 66px;
            height: 66px;
            right: 24%;
            bottom: 22%;
            background: linear-gradient(180deg, #eaffcc 0%, #8be26c 100%);
        }
        .home-visual-spark {
            position: absolute;
            border-radius: 999px;
            background: rgba(255,255,255,0.86);
            box-shadow: 0 0 18px rgba(255,255,255,0.8);
        }
        .home-visual-spark.a {
            width: 8px;
            height: 8px;
            left: 18%;
            top: 18%;
        }
        .home-visual-spark.b {
            width: 12px;
            height: 12px;
            right: 14%;
            top: 22%;
        }
        .home-news-title {
            color: white;
            font-size: 1.25rem;
            font-weight: 800;
            margin: 0.1rem 0 1rem 0.2rem;
        }
        .home-grid-card {
            border-radius: 24px;
            padding: 1.25rem 1.3rem 1.35rem 1.3rem;
            color: white;
            min-height: 260px;
            box-shadow: 0 16px 36px rgba(15, 23, 42, 0.18);
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            gap: 0.8rem;
            margin-bottom: 0.8rem;
            border: 1px solid rgba(255,255,255,0.08);
        }
        .home-card-link {
            display: block;
            text-decoration: none !important;
            color: inherit !important;
        }
        .home-card-link:hover .home-grid-card {
            transform: translateY(-4px);
            box-shadow: 0 22px 42px rgba(15, 23, 42, 0.24);
        }
        .home-grid-card {
            transition: transform 0.18s ease, box-shadow 0.18s ease;
            cursor: pointer;
        }
        .home-grid-card.phase1 {
            background: linear-gradient(180deg, #ff7a59 0%, #ef476f 100%);
        }
        .home-grid-card.phase2 {
            background: linear-gradient(180deg, #2563eb 0%, #1d4ed8 100%);
        }
        .home-grid-card.phase3 {
            background: linear-gradient(180deg, #0f766e 0%, #0f9f8f 100%);
        }
        .home-card-step {
            font-size: 0.84rem;
            font-weight: 700;
            opacity: 0.9;
            letter-spacing: 0.03em;
        }
        .home-card-title {
            font-size: 1.35rem;
            font-weight: 900;
            line-height: 1.3;
            margin: 0.15rem 0 0.45rem 0;
        }
        .home-card-desc {
            font-size: 0.97rem;
            line-height: 1.7;
            opacity: 0.96;
        }
        .home-card-tag {
            display: inline-block;
            padding: 0.22rem 0.55rem;
            border-radius: 999px;
            background: rgba(255,255,255,0.18);
            font-size: 0.8rem;
            font-weight: 700;
            margin-right: 0.35rem;
        }
        @media (max-width: 980px) {
            .home-hero-grid {
                grid-template-columns: 1fr;
            }
            .home-hero-title {
                font-size: 2.3rem;
            }
            .home-hero-visual {
                min-height: 280px;
            }
            .checklist-row {
                grid-template-columns: 1fr;
                gap: 0.4rem;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def main() -> None:
    st.set_page_config(page_title="네이버 블로그 키워드 분석 대시보드", page_icon="📝", layout="wide")
    ensure_session_defaults()
    ensure_translation_data_seeded()
    render_custom_css()
    session_token = st.query_params.get("session")
    if isinstance(session_token, list):
        session_token = session_token[0] if session_token else ""
    if session_token and not st.session_state.get("is_authenticated", False):
        restored_user = decode_auth_token(str(session_token))
        if restored_user:
            st.session_state.is_authenticated = True
            st.session_state.auth_user = restored_user
            st.session_state.current_operator = restored_user
        else:
            if "session" in st.query_params:
                del st.query_params["session"]

    current_page = st.session_state.current_page

    goto_param = st.query_params.get("goto")
    if isinstance(goto_param, list):
        goto_param = goto_param[0] if goto_param else ""
    if goto_param in {"phase1", "phase2", "phase3", "phase4", "phase5"}:
        if "goto" in st.query_params:
            del st.query_params["goto"]
        if st.session_state.get("is_authenticated", False):
            st.session_state.current_page = str(goto_param)
        else:
            st.session_state.post_login_page = str(goto_param)
            st.session_state.home_auth_mode = "login"
            st.session_state.current_page = "auth"
        st.rerun()

    if current_page not in {"home", "auth"} and not st.session_state.get("is_authenticated", False):
        st.session_state.post_login_page = current_page
        st.session_state.home_auth_mode = "login"
        st.session_state.current_page = "auth"
        st.rerun()

    if current_page == "auth":
        scroll_viewport_to_top()
        render_login_screen()
        st.stop()

    if current_page != "home" and not st.session_state.get("is_authenticated", False):
        render_login_screen()
        st.stop()

    if current_page == "home":
        session_qs = st.query_params.get("session", "")
        if isinstance(session_qs, list):
            session_qs = session_qs[0] if session_qs else ""

        def phase_href(phase_name: str) -> str:
            if session_qs:
                return f"?goto={phase_name}&session={session_qs}"
            return f"?goto={phase_name}"

        title_cols = st.columns([6.2, 1.1, 1.1])
        with title_cols[0]:
            st.title("🎰 블로그 가챠 (Blog Gacha)")
            st.caption("버튼 한 번에 터지는 꽝 없는 상위노출 원고! SSS급 콘텐츠를 뽑아보세요.")
        with title_cols[1]:
            login_label = "내 계정" if st.session_state.get("is_authenticated", False) else "로그인"
            if st.button(login_label, use_container_width=True, key="home_title_login_btn"):
                st.session_state.home_auth_mode = "login"
                st.session_state.post_login_page = "home"
                st.session_state.current_page = "auth"
                st.rerun()
                st.stop()
        with title_cols[2]:
            signup_label = "로그아웃" if st.session_state.get("is_authenticated", False) else "회원가입"
            if st.button(signup_label, use_container_width=True, key="home_title_signup_btn"):
                if st.session_state.get("is_authenticated", False):
                    clear_persisted_login()
                    st.session_state.home_auth_mode = ""
                else:
                    st.session_state.home_auth_mode = "signup"
                    st.session_state.post_login_page = "home"
                    st.session_state.current_page = "auth"
                st.rerun()
                st.stop()
        if st.session_state.get("is_authenticated", False):
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")

    with st.sidebar:
        seed_keywords_text = "온다리프팅, 목주름필러, 스킨부스터"
        clinic_name = "세예의원"
        challenge_level = "안전"
        max_keyword_count = 7
        min_search_volume = 50
        max_search_volume_enabled = False
        max_search_volume = None
        view_top_n = 5
        tone_style = "전문적이지만 쉬운 설명"
        include_cta = True
        selected_translation_language = list(TRANSLATION_PRESETS.keys())[0]
        selected_website_language = WEBSITE_CONTENT_LANGUAGES[0]

        if current_page == "home":
            st.header("메인 안내")
            if st.session_state.get("is_authenticated", False):
                st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            else:
                st.caption("현재 로그인: `비로그인`")
            st.caption("메인에서는 단계 선택만 가능합니다.")
            st.caption("원하는 작업을 고른 뒤 각 페이지에서 설정을 조정해 주세요. Phase 진입 시 로그인합니다.")

        elif current_page == "phase1":
            st.header("Phase 1 설정")
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            seed_keywords_text = st.text_area(
                "주력 시술 키워드",
                value="온다리프팅, 목주름필러, 스킨부스터",
                help="콤마(,)로 구분해서 입력하세요.",
                height=100,
            )
            challenge_level = st.select_slider(
                "챌린지 난이도",
                options=["안전", "균형", "도전"],
                value="안전",
                help="안전: 내부 발굴 시트 우선, 균형: 내부 시트와 API 혼합, 도전: API 추천을 더 적극 반영",
            )
            max_keyword_count = st.slider("추천 키워드 개수", min_value=5, max_value=10, value=7)
            min_search_volume = st.slider("최소 월간 검색수", min_value=10, max_value=2000, value=50, step=10)
            max_search_volume_enabled = st.checkbox("최대 월간 검색수 제한 사용", value=False)
            if max_search_volume_enabled:
                max_search_volume = st.slider("최대 월간 검색수", min_value=500, max_value=1000000, value=50000, step=500)

        elif current_page == "phase2":
            st.header("Phase 2 설정")
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            view_top_n = st.slider("분석 게시물 수", min_value=3, max_value=5, value=5)
            if not (NAVER_DATALAB_CLIENT_ID and NAVER_DATALAB_CLIENT_SECRET):
                st.caption("추가로 `NAVER_DATALAB_CLIENT_ID`, `NAVER_DATALAB_CLIENT_SECRET`를 넣으면 월별 검색 트렌드와 월말 예상 검색량까지 볼 수 있습니다.")

        elif current_page == "phase3":
            st.header("Phase 3 설정")
            st.session_state.current_operator = st.session_state.auth_user
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            clinic_name = st.text_input("병원명", value="세예의원")
            tone_style = st.selectbox("원고 톤", ["전문적이지만 쉬운 설명", "친절하고 상담형", "프리미엄 브랜드형"])
            include_cta = st.checkbox("마무리에 상담 유도 문장 넣기", value=True)
            st.caption(f"현재 작업자 저장소: `{current_operator_name()}`")

        elif current_page == "phase4":
            st.header("Phase 4 설정")
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            clinic_name = st.text_input("병원명", value="세예의원", key="phase4_clinic_name")
            selected_translation_language = st.selectbox(
                "번역 언어 프리셋",
                list(TRANSLATION_PRESETS.keys()),
                key="phase4_translation_language",
            )

        elif current_page == "phase5":
            st.header("Phase 5 설정")
            st.caption(f"현재 로그인: `{st.session_state.auth_user}`")
            clinic_name = st.text_input("병원명", value="세예의원", key="phase5_clinic_name")
            selected_website_language = st.selectbox(
                "웹사이트 원고 언어",
                WEBSITE_CONTENT_LANGUAGES,
                key="phase5_website_language",
            )
            st.caption("언어별 프롬프트와 참고 파일을 바탕으로 구글 SEO/GEO용 웹사이트 원고를 재가공합니다.")

        if st.session_state.get("is_authenticated", False):
            with st.expander("작업자 계정 관리", expanded=False):
                new_username = st.text_input("새 작업자 아이디", key="new_account_username")
                new_display_name = st.text_input("표시 이름", key="new_account_display_name")
                new_password = st.text_input("새 작업자 비밀번호", type="password", key="new_account_password")
                if st.button("새 작업자 계정 추가", use_container_width=True, key="create_new_account"):
                    accounts = load_user_accounts()
                    if not new_username.strip() or not new_password.strip():
                        st.warning("아이디와 비밀번호를 입력해 주세요.")
                    elif new_username.strip() in accounts:
                        st.warning("이미 존재하는 작업자 아이디입니다.")
                    else:
                        create_user_account(new_username.strip(), new_password.strip(), new_display_name.strip())
                        st.success(f"`{new_username.strip()}` 계정을 추가했습니다.")

                if st.button("로그아웃", use_container_width=True, key="logout_current_user"):
                    clear_persisted_login()
                    st.session_state.current_page = "home"
                    st.rerun()

            if is_superadmin():
                with st.expander("최고관리자 계정 관리", expanded=False):
                    accounts = load_user_accounts()
                    account_options = sorted(accounts.keys())
                    selected_account = st.selectbox("관리할 작업자", account_options, key="admin_selected_account")
                    selected_info = accounts.get(selected_account, {})
                    st.caption(
                        f"표시 이름: `{selected_info.get('display_name', selected_account)}` | "
                        f"권한: `{selected_info.get('role', 'user')}` | "
                        f"생성일: `{selected_info.get('created_at', '-')}`"
                    )
                    managed_languages = st.multiselect(
                        "관리 가능 언어",
                        options=list(TRANSLATION_PRESETS.keys()),
                        default=selected_info.get("managed_languages", []),
                        key="admin_managed_languages",
                    )
                    if st.button("이 작업자 언어 권한 저장", use_container_width=True, key="admin_save_languages_btn"):
                        update_user_managed_languages(selected_account, managed_languages)
                        st.success(f"`{selected_account}` 계정의 언어 권한을 저장했습니다.")
                    reset_password = st.text_input("새 비밀번호", type="password", key="admin_reset_password")
                    if st.button("이 작업자 비밀번호 재설정", use_container_width=True, key="admin_reset_password_btn"):
                        if not reset_password.strip():
                            st.warning("새 비밀번호를 입력해 주세요.")
                        else:
                            update_user_password(selected_account, reset_password.strip())
                            st.success(f"`{selected_account}` 계정 비밀번호를 재설정했습니다.")

        st.info("테스트 단계에서는 로컬/사내 네트워크에서 충분히 검증할 수 있습니다.")

    seed_keywords = normalize_seed_keywords(seed_keywords_text)

    if not seed_keywords:
        st.warning("주력 시술 키워드를 1개 이상 입력해 주세요.")
        return

    if max_search_volume is not None and min_search_volume > max_search_volume:
        st.warning("최소 월간 검색수는 최대 월간 검색수보다 클 수 없습니다. 범위를 다시 조정해 주세요.")
        return

    if current_page == "home":
        st.markdown(
            """
            <div class="home-shell">
                <div class="home-topbar">
                    <div class="home-topbar-brand">🎰 Blog Gacha</div>
                    <div class="home-topbar-menu">
                        <span>키워드 발굴</span>
                        <span>포스트 분석</span>
                        <span>원고 생성</span>
                        <span>다국어 번역</span>
                        <span>프리셋 관리</span>
                    </div>
                </div>
                <div class="home-hero">
                    <div class="home-hero-grid">
                        <div>
                            <div class="home-logo-badge">Naver Blog Marketing Dashboard</div>
                            <div class="home-hero-title">버튼 한 번에<br/>SSS급 원고 뽑기</div>
                            <div class="home-hero-sub">네이버 블로그 마케팅에 필요한 키워드 발굴, 메인 1면 상위노출 분석, 원고 초안 생성을 한 화면 흐름으로 연결했습니다. 이번에 쓸 주제를 빠르게 점검하고, 상위 콘텐츠 구조를 읽고, 바로 실무형 초안까지 이어서 만들 수 있습니다.</div>
                            <div class="home-hero-actions">
                                <div class="home-action-chip">꽝 없는 상위노출 흐름</div>
                                <div class="home-action-chip">실측 + 추정 + 추천 진단</div>
                            </div>
                        </div>
                        <div class="home-hero-visual">
                            <div class="home-visual-machine"></div>
                            <div class="home-visual-ball one"></div>
                            <div class="home-visual-ball two"></div>
                            <div class="home-visual-ball three"></div>
                            <div class="home-visual-spark a"></div>
                            <div class="home-visual-spark b"></div>
                        </div>
                    </div>
                </div>
                <div class="home-news-title">오늘 뽑을 가챠 메뉴</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        overview_cols = st.columns(3)
        overview_cols[0].markdown(
            f"""
            <a class="home-card-link" href="{phase_href('phase1')}">
                <div class="home-grid-card phase1">
                    <div>
                        <div class="home-card-step">PHASE 1</div>
                        <div class="home-card-title">추천 키워드 발굴</div>
                        <div class="home-card-desc">검색량, 경쟁도, 내부 시트 기준을 같이 봐서 이번 주에 써볼 SSS급 키워드를 좁힙니다.</div>
                    </div>
                    <div>
                        <span class="home-card-tag">새 캠페인</span>
                        <span class="home-card-tag">롱테일 발굴</span>
                        <span class="home-card-tag">빠른 검증</span>
                    </div>
                </div>
            </a>
            """,
            unsafe_allow_html=True,
        )
        overview_cols[1].markdown(
            f"""
            <a class="home-card-link" href="{phase_href('phase2')}">
                <div class="home-grid-card phase2">
                    <div>
                        <div class="home-card-step">PHASE 2</div>
                        <div class="home-card-title">상위 노출 포스트 분석</div>
                        <div class="home-card-desc">통합검색 메인 1면 기준으로 블로그, 카페, 사이트 흐름을 읽고 상위 콘텐츠 구조를 분석합니다.</div>
                    </div>
                    <div>
                        <span class="home-card-tag">경쟁도 판단</span>
                        <span class="home-card-tag">채널 진단</span>
                        <span class="home-card-tag">메인 1면 기준</span>
                    </div>
                </div>
            </a>
            """,
            unsafe_allow_html=True,
        )
        overview_cols[2].markdown(
            f"""
            <a class="home-card-link" href="{phase_href('phase3')}">
                <div class="home-grid-card phase3">
                    <div>
                        <div class="home-card-step">PHASE 3</div>
                        <div class="home-card-title">초안 자동 생성</div>
                        <div class="home-card-desc">시술 프리셋, 참고 문서, 참고 URL을 반영해 바로 붙여넣을 수 있는 원고 초안을 생성합니다.</div>
                    </div>
                    <div>
                        <span class="home-card-tag">프리셋 반영</span>
                        <span class="home-card-tag">즉시 생성</span>
                        <span class="home-card-tag">2차 수정 가능</span>
                    </div>
                </div>
            </a>
            """,
            unsafe_allow_html=True,
        )

        second_overview_cols = st.columns(2)
        second_overview_cols[0].markdown(
            f"""
            <a class="home-card-link" href="{phase_href('phase4')}">
                <div class="home-grid-card phase2">
                    <div>
                        <div class="home-card-step">PHASE 4</div>
                        <div class="home-card-title">다국어 번역</div>
                        <div class="home-card-desc">한국어 원고를 대만/중국/일본/영어 프리셋으로 번역하고, 추가 요청으로 2차 번역본까지 다듬습니다.</div>
                    </div>
                    <div>
                        <span class="home-card-tag">번역 프리셋</span>
                        <span class="home-card-tag">2차 수정 가능</span>
                        <span class="home-card-tag">표/볼드 유지</span>
                    </div>
                </div>
            </a>
            """,
            unsafe_allow_html=True,
        )
        second_overview_cols[1].markdown(
            f"""
            <a class="home-card-link" href="{phase_href('phase5')}">
                <div class="home-grid-card phase3">
                    <div>
                        <div class="home-card-step">PHASE 5</div>
                        <div class="home-card-title">웹사이트 원고 재가공</div>
                        <div class="home-card-desc">언어별 프롬프트와 참고 자료를 반영해 구글 SEO/GEO용 웹사이트 소개·상세·FAQ 원고로 재구성합니다.</div>
                    </div>
                    <div>
                        <span class="home-card-tag">언어별 프롬프트</span>
                        <span class="home-card-tag">레퍼런스 반영</span>
                        <span class="home-card-tag">2차 수정 가능</span>
                    </div>
                </div>
            </a>
            """,
            unsafe_allow_html=True,
        )
        st.info("키워드부터 시작해도 되고, 상위노출 분석이나 원고 생성으로 바로 들어가도 됩니다.")
        return

    if current_page == "phase1":
        st.subheader("Phase 1. 꿀 키워드 발굴")
        render_phase_top_nav(("← 메인으로", "home"), ("Phase 2로", "phase2"), ("Phase 3로", "phase3"), ("Phase 4로", "phase4"))
    if current_page == "phase1" and st.button("이번 주 추천 키워드 발굴하기", type="primary", use_container_width=True):
        with st.spinner("네이버 검색광고 API와 VIEW 문서 수를 기준으로 추천 키워드를 정리하는 중입니다..."):
            try:
                build_keyword_candidates.clear()
                load_keyword_workbook_reference.clear()
                fetch_searchad_metrics_for_keywords.clear()
                fetch_exact_searchad_metric.clear()
                grouped_candidates = build_keyword_candidates(
                    tuple(seed_keywords),
                    min_search_volume,
                    max_search_volume,
                    max_keyword_count,
                    challenge_level,
                )
                keyword_df = flatten_grouped_candidates(grouped_candidates)
                if keyword_df.empty:
                    st.warning("추천할 만한 파생 키워드를 찾지 못했습니다. 입력 키워드나 최소 검색수를 조정해 보세요.")
                else:
                    st.session_state.grouped_candidates = grouped_candidates
                    st.session_state.recommended_keywords_df = keyword_df
                    st.session_state.analysis_summary = None
                    st.session_state.generated_draft = None
                    st.session_state.generated_draft_context = None
                    st.session_state.phase3_direct_seed_override = ""
            except Exception as exc:
                if "잠시 제한" in str(exc) or "429" in str(exc):
                    st.error("네이버 검색광고 API 호출이 잠시 몰려 있습니다. 1~2분 뒤 다시 시도하거나 입력 키워드 수를 줄여 보세요.")
                elif "403" in str(exc):
                    st.error("네이버 검색 결과 문서 수 확인이 잠시 차단됐습니다. 지금은 광고 API 데이터 중심으로 추천되도록 보정 중이니 다시 한 번 눌러보세요.")
                else:
                    st.error(f"키워드 추천 중 오류가 발생했습니다: {exc}")

    keyword_df: pd.DataFrame | None = st.session_state.recommended_keywords_df
    grouped_candidates: dict[str, list[dict[str, Any]]] | None = st.session_state.grouped_candidates
    if current_page == "phase1":
        if keyword_df is not None and not keyword_df.empty:
            keyword_tab, insight_tab = st.tabs(["추천 키워드 표", "실무 해석"])
            with keyword_tab:
                if grouped_candidates:
                    st.caption(f"현재 난이도는 `{challenge_level}` 입니다. `안전`은 내부 시트 우선, `균형`은 혼합, `도전`은 API 추천 비중을 더 높입니다. `데이터 구분`이 `내부 시트`면 기존 실무 시트에서 가져온 후보이고, `추정 후보`면 보조 규칙으로 만든 키워드입니다.")
                    for seed_keyword, items in grouped_candidates.items():
                        st.markdown(f"**{seed_keyword} 추천 키워드**")
                        if items:
                            table_df = render_keyword_table(items)
                            st.dataframe(style_keyword_table(table_df), use_container_width=True, hide_index=True)
                        else:
                            st.caption("조건에 맞는 추천 키워드가 없습니다.")
            with insight_tab:
                st.caption(f"지역 키워드, 병원명, 너무 포괄적인 단일 시술명은 제외하고 주력 시술별로 따로 추천합니다. 현재 `{challenge_level}` 기준으로 후보 출처 우선순위를 조절하고 있습니다.")
                for seed_keyword, items in (grouped_candidates or {}).items():
                    if not items:
                        continue
                    seed_df = render_keyword_table(items)
                    top_row = seed_df.iloc[0]
                    st.success(
                        f"'{seed_keyword}' 기준 1순위는 '{top_row['키워드']}' 입니다. "
                        f"이유는 {top_row['추천 이유']} 입니다."
                    )
                    st.markdown(
                        f"- 검색량이 가장 큰 키워드: `{seed_df.sort_values('월간 검색수', ascending=False).iloc[0]['키워드']}`\n"
                        f"- 포화도가 가장 낮은 키워드: `{seed_df.sort_values('포화도', ascending=True).iloc[0]['키워드']}`\n"
                        f"- 테스트용 롱테일 후보: `{seed_df.sort_values('월간 검색수', ascending=True).iloc[0]['키워드']}`"
                    )
        else:
            st.caption("아직 추천 키워드를 발굴하지 않았습니다.")
        if keyword_df is not None and not keyword_df.empty:
            render_phase_footer_nav(("메인으로", "home"), ("Phase 2로 이동", "phase2"), ("Phase 3로 이동", "phase3"), ("Phase 4로 이동", "phase4"))

    selected_seed_keyword = ""
    if current_page == "phase2":
        st.subheader("Phase 2. 경쟁사 분석")
        render_phase_top_nav(("← 메인으로", "home"), ("Phase 1로", "phase1"), ("Phase 3로", "phase3"), ("Phase 4로", "phase4"))
        actual_keyword = ""
        if keyword_df is not None and not keyword_df.empty:
            selectable_rows = keyword_df[["주력 시술", "키워드"]].copy()
            selectable_labels = [f"{row['주력 시술']} | {row['키워드']}" for _, row in selectable_rows.iterrows()]
            selected_keyword = st.selectbox(
                "추천 결과에서 선택하기",
                selectable_labels,
                index=0,
                key="phase2_selected_keyword",
            )
            selected_seed_keyword = selected_keyword.split(" | ", 1)[0]
            actual_keyword = selected_keyword.split(" | ", 1)[1]
        else:
            st.caption("Phase 1 없이도 바로 분석할 수 있습니다.")
            selected_seed_keyword = st.text_input("주력 시술명", value="", key="phase2_manual_seed")
            actual_keyword = st.text_input("직접 분석할 키워드", value="", key="phase2_manual_keyword")
            st.caption("직접 분석은 한 번에 1개 키워드 기준으로 동작합니다. 여러 개를 입력하면 첫 번째 키워드만 사용합니다.")

        if st.button("이 키워드로 통합검색 상위 블로그 분석하기", use_container_width=True):
            with st.spinner("네이버 통합검색 렌더링 화면에서 상위 블로그를 읽고 기준선을 정리하는 중입니다..."):
                try:
                    if not actual_keyword.strip():
                        raise RuntimeError("분석할 키워드를 입력해 주세요.")
                    actual_keyword = first_keyword_from_text(actual_keyword)
                    extract_main_page_results.clear()
                    debug_extract_blog_links.clear()
                    analyze_keyword_competition.clear()
                    st.session_state.analysis_summary = analyze_keyword_competition(actual_keyword, view_top_n)
                    st.session_state.generated_draft = None
                    st.session_state.generated_draft_context = None
                    st.session_state.phase3_direct_seed_override = selected_seed_keyword or ""
                except Exception as exc:
                    st.error(f"경쟁 포스트 분석 중 오류가 발생했습니다: {exc}")
        with st.expander("디버그: 통합검색에서 잡힌 블로그 링크 보기", expanded=False):
            try:
                if actual_keyword.strip():
                    debug_links = debug_extract_blog_links(first_keyword_from_text(actual_keyword), limit=15)
                    if debug_links:
                        st.dataframe(pd.DataFrame(debug_links), use_container_width=True, hide_index=True)
                    else:
                        st.caption("잡힌 블로그 링크가 없습니다.")
                else:
                    st.caption("키워드를 입력하면 디버그 링크를 확인할 수 있습니다.")
            except Exception as exc:
                st.caption(f"디버그 링크 추출 실패: {exc}")

    summary = st.session_state.analysis_summary
    if summary and current_page == "phase2":
        render_keyword_diagnosis_section(summary)

        metric_cols = st.columns(5)
        metric_cols[0].metric("타겟 키워드", summary["keyword"])
        metric_cols[1].metric("분석 게시물 수", f'{summary["post_count"]}개')
        metric_cols[2].metric("평균 글자 수", f'{summary["avg_chars"]}자')
        metric_cols[3].metric("중앙값 글자 수", f'{summary["median_chars"]}자')
        metric_cols[4].metric("권장 이미지 수", f"{summary['recommended_images']}장")

        if summary["post_count"] == 0:
            st.info(summary.get("channel_insight", "메인 1면에 블로그 노출이 확인되지 않았습니다."))
            if summary.get("channel_counts"):
                channel_label_map = {
                    "blog": "블로그",
                    "cafe": "카페",
                    "kin": "지식인",
                    "site": "웹사이트",
                    "other": "기타",
                }
                channel_df = pd.DataFrame(
                    [
                        {"채널": channel_label_map.get(key, key), "건수": value}
                        for key, value in summary["channel_counts"].items()
                    ]
                )
                st.markdown("**메인 1면 채널 분포**")
                st.dataframe(channel_df, use_container_width=True, hide_index=True)
                st.markdown("**채널별 콘텐츠 제안**")
                for idea in build_channel_content_ideas(summary["channel_counts"]):
                    st.markdown(f"- {idea}")
            if summary.get("main_page_results"):
                visible_df = pd.DataFrame(
                    [
                        {
                            "채널": {
                                "blog": "블로그",
                                "cafe": "카페",
                                "kin": "지식인",
                                "site": "웹사이트",
                                "other": "기타",
                            }.get(row["channel"], row["channel"]),
                            "제목": row["title"],
                            "URL": row["url"],
                        }
                        for row in summary["main_page_results"]
                    ]
                )
                st.markdown("**메인 1면 인식 결과**")
                st.dataframe(visible_df, use_container_width=True, hide_index=True)
        else:
            st.markdown(
                f"""
                <div class="app-card">
                    <div class="stage-title">경쟁사 분석 요약</div>
                    <div class="stage-help">
                        권장 원고 길이는 약 <b>{summary["recommended_chars"]}자 이상</b>이고,
                        이번 키워드는 <b>{summary["writing_angle"]}</b> 흐름으로 쓰는 편이 유리합니다.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            term_col, post_col = st.columns([1, 1.4])
            with term_col:
                st.markdown("**자주 쓰인 형태소(단어)**")
                top_term_df = pd.DataFrame(summary["top_terms"], columns=["단어", "빈도"])
                st.dataframe(top_term_df, use_container_width=True, hide_index=True)
            with post_col:
                st.markdown("**상위 노출 게시물 상세**")
                st.dataframe(build_post_detail_table(summary["posts"]), use_container_width=True, hide_index=True)
                st.markdown("**원문 링크 열기**")
                for post in summary["posts"]:
                    st.markdown(f"- [{post['title']}]({post['url']})")
        render_phase_footer_nav(("메인으로", "home"), ("Phase 1로 이동", "phase1"), ("Phase 3으로 이동", "phase3"), ("Phase 4로 이동", "phase4"))

    if current_page == "phase3":
        st.subheader("Phase 3. 원고 자동 생성")
        render_phase_top_nav(("← 메인으로", "home"), ("Phase 1로", "phase1"), ("Phase 2로", "phase2"), ("Phase 4로", "phase4"))
        if not summary:
            st.info("Phase 2 없이도 바로 원고를 만들 수 있습니다. 주력 시술과 타겟 키워드를 직접 입력해 주세요.")
            direct_seed_keyword = st.text_input("주력 시술명", value="", key="phase3_direct_seed")
            direct_keyword = st.text_input("타겟 키워드", value="", key="phase3_direct_keyword")
            if st.button("이 키워드로 바로 분석 후 원고 생성 준비", use_container_width=True, key="phase3_direct_analyze"):
                with st.spinner("키워드를 바로 분석해 원고 생성 준비를 하는 중입니다..."):
                    try:
                        if not direct_keyword.strip():
                            raise RuntimeError("타겟 키워드를 입력해 주세요.")
                        analyze_keyword_competition.clear()
                        st.session_state.analysis_summary = analyze_keyword_competition(direct_keyword.strip(), view_top_n)
                        st.session_state.generated_draft = None
                        st.session_state.generated_draft_context = None
                        if direct_seed_keyword.strip():
                            st.session_state.phase3_direct_seed_override = direct_seed_keyword.strip()
                        go_to_page("phase3")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"직접 분석 중 오류가 발생했습니다: {exc}")
        else:
            selected_seed_keyword = st.session_state.get("phase3_direct_seed_override") or infer_seed_keyword_from_summary(summary, keyword_df)
            operator_name = current_operator_name()
            initial_profile = resolve_procedure_profile(selected_seed_keyword, "", operator_name)
            if initial_profile["source"] == "preset":
                st.caption(f"`{selected_seed_keyword}`는 작업자 `{operator_name}`의 저장된 시술 프리셋 `{initial_profile['profile_name']}` 기준으로 원고를 생성합니다.")
            else:
                st.caption(f"작업자 `{operator_name}` 기준 저장된 시술 프리셋이 없는 키워드입니다. 신규 시술 설명을 넣으면 그 내용을 프롬프트에 반영합니다.")

            with st.expander("시술 프리셋 관리", expanded=False):
                manage_profile_name = st.text_input(
                    "시술명",
                    value=selected_seed_keyword,
                    key=f"manage_profile_name_{selected_seed_keyword}",
                    help="이 시술명 기준으로 프리셋과 참고 파일을 저장합니다.",
                )
                user_profiles = load_user_procedure_profiles(operator_name)
                editable_profile = user_profiles.get(
                    manage_profile_name,
                    {
                        "core_description": initial_profile.get("core_description", ""),
                        "must_include": initial_profile.get("must_include", []),
                        "avoid_phrases": initial_profile.get("avoid_phrases", []),
                        "preferred_format": initial_profile.get("preferred_format", "설명형"),
                        "cta_style": initial_profile.get("cta_style", ""),
                    },
                )
                preset_core_description = st.text_area(
                    "프리셋 시술 설명",
                    value=editable_profile.get("core_description", ""),
                    height=100,
                    key=f"preset_core_description_{selected_seed_keyword}_{manage_profile_name}",
                )
                preset_must_include = st.text_input(
                    "반드시 녹일 포인트",
                    value=", ".join(editable_profile.get("must_include", [])),
                    key=f"preset_must_include_{selected_seed_keyword}_{manage_profile_name}",
                )
                preset_avoid_phrases = st.text_input(
                    "피해야 할 표현",
                    value=", ".join(editable_profile.get("avoid_phrases", [])),
                    key=f"preset_avoid_phrases_{selected_seed_keyword}_{manage_profile_name}",
                )
                preset_preferred_format = st.selectbox(
                    "선호 포맷",
                    ["설명형", "정보형", "비교형", "후기형"],
                    index=["설명형", "정보형", "비교형", "후기형"].index(editable_profile.get("preferred_format", "설명형"))
                    if editable_profile.get("preferred_format", "설명형") in ["설명형", "정보형", "비교형", "후기형"]
                    else 0,
                    key=f"preset_preferred_format_{selected_seed_keyword}_{manage_profile_name}",
                )
                preset_cta_style = st.text_area(
                    "CTA 가이드",
                    value=editable_profile.get("cta_style", ""),
                    height=80,
                    key=f"preset_cta_style_{selected_seed_keyword}_{manage_profile_name}",
                )
                if st.button("이 시술 프리셋 저장/수정", key=f"save_preset_{selected_seed_keyword}_{manage_profile_name}"):
                    save_user_procedure_profile(
                        manage_profile_name,
                        {
                            "core_description": preset_core_description.strip(),
                            "must_include": [item.strip() for item in preset_must_include.split(",") if item.strip()],
                            "avoid_phrases": [item.strip() for item in preset_avoid_phrases.split(",") if item.strip()],
                            "preferred_format": preset_preferred_format,
                            "cta_style": preset_cta_style.strip(),
                        },
                        operator_name,
                    )
                    st.success(f"작업자 `{operator_name}`의 `{manage_profile_name}` 프리셋을 저장했습니다. 다음 생성부터 이 설정을 우선 반영합니다.")

                preset_action_cols = st.columns(2)
                if preset_action_cols[0].button("사용자 프리셋 삭제", key=f"delete_preset_{selected_seed_keyword}_{manage_profile_name}"):
                    delete_user_procedure_profile(manage_profile_name, operator_name)
                    st.success(f"작업자 `{operator_name}`의 `{manage_profile_name}` 사용자 프리셋을 삭제했습니다. 기본 프리셋 또는 공통 프롬프트로 돌아갑니다.")
                    st.rerun()
                if preset_action_cols[1].button("기본 프리셋으로 되돌리기", key=f"reset_preset_{selected_seed_keyword}_{manage_profile_name}"):
                    delete_user_procedure_profile(manage_profile_name, operator_name)
                    st.success(f"작업자 `{operator_name}`의 `{manage_profile_name}` 프리셋을 기본값으로 되돌렸습니다.")
                    st.rerun()

                builtin_files = list_builtin_reference_files(manage_profile_name)
                st.markdown("**기존 기준 자료 가져오기**")
                if builtin_files:
                    selected_builtin_names = st.multiselect(
                        "원본 자료 선택",
                        options=[path.name for path in builtin_files],
                        key=f"builtin_files_{selected_seed_keyword}_{manage_profile_name}",
                        help="선택한 파일은 앱 전용 저장소로 복사됩니다. 원본 바탕화면 파일은 그대로 유지됩니다.",
                    )
                    if st.button("선택한 기준 자료 가져오기", key=f"import_builtin_{selected_seed_keyword}_{manage_profile_name}"):
                        for file_path in builtin_files:
                            if file_path.name in selected_builtin_names:
                                import_builtin_reference_file(manage_profile_name, file_path, operator_name)
                        st.success(f"작업자 `{operator_name}`의 `{manage_profile_name}`용 기준 자료를 가져왔습니다.")
                else:
                    st.caption("이 시술과 매칭되는 원본 기준 자료 폴더가 없습니다.")

                st.markdown("**참고 파일 업로드**")
                uploaded_reference_files = st.file_uploader(
                    "docx/txt 파일 업로드",
                    type=["docx", "txt"],
                    accept_multiple_files=True,
                    key=f"uploaded_reference_files_{selected_seed_keyword}_{manage_profile_name}",
                    help="같은 파일명으로 다시 업로드하면 교체됩니다.",
                )
                if st.button("이 시술 참고 파일 저장", key=f"save_reference_files_{selected_seed_keyword}_{manage_profile_name}"):
                    if uploaded_reference_files:
                        save_uploaded_reference_files(manage_profile_name, uploaded_reference_files, operator_name)
                        st.success(f"작업자 `{operator_name}`의 `{manage_profile_name}` 참고 파일을 저장했습니다.")
                    else:
                        st.warning("업로드할 파일을 먼저 선택해 주세요.")

                existing_files = list_user_reference_files(manage_profile_name, operator_name)
                if existing_files:
                    st.markdown("**기존 참고 파일**")
                    for file_path in existing_files:
                        file_cols = st.columns([6, 1])
                        file_cols[0].caption(file_path.name)
                        if file_cols[1].button("삭제", key=f"delete_ref_{manage_profile_name}_{file_path.name}"):
                            delete_user_reference_file(manage_profile_name, file_path.name, operator_name)
                            st.success(f"작업자 `{operator_name}`의 `{file_path.name}` 파일을 삭제했습니다.")
                            st.rerun()
                else:
                    st.caption(f"작업자 `{operator_name}` 기준 저장된 참고 파일이 없습니다.")

            custom_procedure_description = st.text_area(
                "시술 설명 보강/덮어쓰기",
                value="",
                height=100,
                placeholder="프리셋이 있어도 추가로 강조할 포인트가 있으면 적어 주세요. 신규 시술이면 시술 원리, 주요 적응증, 핵심 장점, 피하고 싶은 표현 등을 적어 주세요.",
                key=f"custom_procedure_description_{selected_seed_keyword}",
                help="입력하면 저장된 시술 프리셋 대신 이 설명을 우선 반영합니다.",
            )
            reference_urls_text = st.text_area(
                "참고 URL",
                value="",
                height=90,
                placeholder="재가공에 참고할 블로그나 웹사이트 URL을 줄바꿈으로 입력해 주세요. 최대 3개까지 반영합니다.",
                key=f"reference_urls_{selected_seed_keyword}",
                help="입력한 URL의 핵심 내용을 요약해 참고 자료로 활용합니다. 문장은 새롭게 재구성합니다.",
            )
            if st.button("이 데이터로 블로그 원고 작성하기", use_container_width=True):
                with st.spinner("Gemini가 제목과 본문 초안을 작성하는 중입니다..."):
                    try:
                        procedure_profile = resolve_procedure_profile(selected_seed_keyword, custom_procedure_description, operator_name)
                        st.session_state.generated_draft = generate_blog_draft_with_gemini(
                            summary=summary,
                            clinic_name=clinic_name,
                            tone_style=tone_style,
                            include_cta=include_cta,
                            seed_keyword=selected_seed_keyword,
                            procedure_profile=procedure_profile,
                            reference_urls_text=reference_urls_text,
                            operator_name=operator_name,
                        )
                        st.session_state.generated_draft_context = {
                            "summary": summary,
                            "clinic_name": clinic_name,
                            "tone_style": tone_style,
                            "include_cta": include_cta,
                            "seed_keyword": selected_seed_keyword,
                            "procedure_profile": procedure_profile,
                            "reference_urls_text": reference_urls_text,
                            "operator_name": operator_name,
                        }
                    except Exception as exc:
                        st.error(f"원고 생성 중 오류가 발생했습니다: {exc}")

    if current_page == "phase4":
        st.subheader("Phase 4. 다국어 번역")
        render_phase_top_nav(("← 메인으로", "home"), ("Phase 3로", "phase3"), ("Phase 5로", "phase5"))
        current_draft = st.session_state.get("generated_draft")
        current_context = st.session_state.get("generated_draft_context") or {}
        translation_seed_keyword = current_context.get("seed_keyword") or st.session_state.get("phase3_direct_seed_override") or seed_keywords[0]
        current_user = st.session_state.get("auth_user", "")
        translation_prompts = load_translation_prompts()
        translation_glossaries = load_translation_glossaries()

        if "phase4_source_text" not in st.session_state and current_draft:
            st.session_state.phase4_source_text = current_draft.get("body", "")

        if current_draft:
            st.caption(f"현재 Phase 3 원고를 기준으로 `{selected_translation_language}` 번역 프리셋을 적용할 수 있습니다.")
            if st.button("Phase 3 본문 불러오기", key="phase4_load_from_phase3"):
                st.session_state.phase4_source_text = current_draft.get("body", "")
                st.rerun()

        preset_info = TRANSLATION_PRESETS[selected_translation_language]
        st.markdown(
            f"""
            <div class="app-card">
                <div class="stage-title">{selected_translation_language} 번역 프리셋</div>
                <div class="stage-help">
                    {preset_info["style"]}<br/>{preset_info["notes"]}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        can_manage_selected_language = can_manage_translation_language(selected_translation_language, current_user)
        with st.expander("언어별 학습 데이터 관리", expanded=False):
            st.caption("선택 언어의 번역 프롬프트와 용어집을 저장해 두면, 번역 실행 시 자동으로 반영됩니다.")
            if not can_manage_selected_language:
                allowed_languages = ", ".join(get_user_managed_languages(current_user)) or "없음"
                st.info(f"현재 계정은 `{selected_translation_language}` 자료를 수정할 권한이 없습니다. 관리 가능 언어: {allowed_languages}")

            prompt_rows = []
            for language_label in TRANSLATION_PRESETS:
                prompt_record = translation_prompts.get(language_label, {})
                prompt_rows.append(
                    {
                        "언어": language_label,
                        "현재 등록된 프롬프트": str(prompt_record.get("text", ""))[:120] + ("..." if len(str(prompt_record.get("text", ""))) > 120 else ""),
                        "수정자": prompt_record.get("updated_by", "-"),
                    }
                )
            st.markdown("**A. 번역 프롬프트 관리**")
            st.dataframe(pd.DataFrame(prompt_rows), use_container_width=True, hide_index=True)
            prompt_manage_tab, glossary_manage_tab = st.tabs(
                [f"{selected_translation_language} 프롬프트 수정", f"{selected_translation_language} 용어집 수정"]
            )

            with prompt_manage_tab:
                prompt_record = translation_prompts.get(selected_translation_language, {})
                prompt_text = st.text_area(
                    "프롬프트 텍스트",
                    value=prompt_record.get("text", _default_translation_prompt(selected_translation_language)),
                    height=180,
                    key=f"translation_prompt_text_{selected_translation_language}",
                    disabled=not can_manage_selected_language,
                )
                prompt_file = st.file_uploader(
                    "또는 .txt / .md 파일 업로드",
                    type=["txt", "md"],
                    key=f"translation_prompt_file_{selected_translation_language}",
                    disabled=not can_manage_selected_language,
                )
                prompt_action_cols = st.columns(2)
                if prompt_action_cols[0].button("프롬프트 업로드/수정", key=f"save_translation_prompt_{selected_translation_language}", disabled=not can_manage_selected_language):
                    try:
                        final_text = prompt_text.strip()
                        if prompt_file is not None:
                            final_text = prompt_file.getvalue().decode("utf-8").strip()
                        if not final_text:
                            raise RuntimeError("저장할 프롬프트 내용을 입력하거나 파일을 업로드해 주세요.")
                        save_translation_prompt(selected_translation_language, final_text, current_user)
                        st.success(f"{selected_translation_language} 프롬프트를 저장했습니다.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"프롬프트 저장 중 오류가 발생했습니다: {exc}")
                if prompt_action_cols[1].button("프롬프트 삭제(기본값 복구)", key=f"delete_translation_prompt_{selected_translation_language}", disabled=not can_manage_selected_language):
                    delete_translation_prompt(selected_translation_language, current_user)
                    st.success(f"{selected_translation_language} 프롬프트를 기본값으로 되돌렸습니다.")
                    st.rerun()

            glossary_rows = []
            for language_label in TRANSLATION_PRESETS:
                glossary_record = translation_glossaries.get(language_label, {})
                glossary_rows.append(
                    {
                        "언어": language_label,
                        "용어집 파일명": glossary_record.get("filename", "-"),
                        "현재 등록된 용어 수": len(glossary_record.get("terms", {})),
                        "수정자": glossary_record.get("updated_by", "-"),
                    }
                )
            st.markdown("**B. 용어집(Glossary) 관리**")
            st.dataframe(pd.DataFrame(glossary_rows), use_container_width=True, hide_index=True)
            with glossary_manage_tab:
                glossary_record = translation_glossaries.get(selected_translation_language, {})
                glossary_preview = pd.DataFrame(
                    [{"원본단어": k, "번역단어": v} for k, v in list(glossary_record.get("terms", {}).items())[:20]]
                )
                if not glossary_preview.empty:
                    st.caption("현재 등록된 용어집 미리보기")
                    st.dataframe(glossary_preview, use_container_width=True, hide_index=True)
                glossary_file = st.file_uploader(
                    ".json / .csv / .xlsx 용어집 업로드",
                    type=["json", "csv", "xlsx"],
                    key=f"translation_glossary_file_{selected_translation_language}",
                    disabled=not can_manage_selected_language,
                )
                glossary_action_cols = st.columns(2)
                if glossary_action_cols[0].button("용어집 업로드/수정", key=f"save_translation_glossary_{selected_translation_language}", disabled=not can_manage_selected_language):
                    try:
                        if glossary_file is None:
                            raise RuntimeError("업로드할 용어집 파일을 선택해 주세요.")
                        terms = parse_glossary_upload(glossary_file)
                        save_translation_glossary(selected_translation_language, terms, glossary_file.name, current_user)
                        st.success(f"{selected_translation_language} 용어집을 저장했습니다. 현재 {len(terms)}개 용어가 등록되었습니다.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"용어집 저장 중 오류가 발생했습니다: {exc}")
                if glossary_action_cols[1].button("용어집 삭제(기본값 복구)", key=f"delete_translation_glossary_{selected_translation_language}", disabled=not can_manage_selected_language):
                    delete_translation_glossary(selected_translation_language, current_user)
                    st.success(f"{selected_translation_language} 용어집을 기본값으로 되돌렸습니다.")
                    st.rerun()

        source_text = st.text_area(
            "한국어 원문",
            key="phase4_source_text",
            height=320,
            placeholder="번역할 한국어 원고를 붙여넣거나, Phase 3 본문을 불러와 사용하세요.",
        )
        translation_additional_request = st.text_area(
            "추가 번역 요청",
            key="phase4_additional_request",
            height=100,
            placeholder="예: 일본어는 더 부드럽고 상담형으로, 영어는 clinic blog tone을 더 강하게 반영해 주세요.",
        )

        if st.button("이 원고를 번역하기", use_container_width=True, key="phase4_translate_btn"):
            with st.spinner(f"Gemini가 {selected_translation_language} 번역본을 만드는 중입니다..."):
                try:
                    st.session_state.translated_draft = generate_translation_with_gemini(
                        korean_text=source_text,
                        target_language_label=selected_translation_language,
                        additional_request=translation_additional_request,
                        seed_keyword=translation_seed_keyword,
                        clinic_name=clinic_name,
                    )
                    st.session_state.translated_draft_context = {
                        "seed_keyword": translation_seed_keyword,
                        "clinic_name": clinic_name,
                        "language_label": selected_translation_language,
                    }
                except Exception as exc:
                    st.error(f"번역 중 오류가 발생했습니다: {exc}")

    if current_page == "phase5":
        st.subheader("Phase 5. 구글 SEO/GEO 웹사이트 원고")
        render_phase_top_nav(("← 메인으로", "home"), ("Phase 3로", "phase3"), ("Phase 4로", "phase4"))

        current_user = current_operator_name()
        current_draft = st.session_state.get("generated_draft")
        current_translation = st.session_state.get("translated_draft")
        phase5_seed_keyword = (
            (st.session_state.get("generated_draft_context") or {}).get("seed_keyword")
            or (st.session_state.get("translated_draft_context") or {}).get("seed_keyword")
            or st.session_state.get("phase3_direct_seed_override")
            or seed_keywords[0]
        )

        if "phase5_source_text" not in st.session_state:
            st.session_state.phase5_source_text = current_draft.get("body", "") if current_draft else ""

        if current_draft or current_translation:
            load_cols = st.columns(2)
            if current_draft and load_cols[0].button("Phase 3 본문 불러오기", key="phase5_load_phase3"):
                st.session_state.phase5_source_text = current_draft.get("body", "")
                st.rerun()
            if current_translation and load_cols[1].button("Phase 4 번역본 불러오기", key="phase5_load_phase4"):
                st.session_state.phase5_source_text = current_translation.get("translated_body", "")
                st.rerun()

        website_prompts = load_user_website_prompts(current_user)
        selected_prompt_record = website_prompts.get(selected_website_language, {})
        selected_prompt = selected_prompt_record.get("text", _default_website_prompt(selected_website_language))
        reference_files = list_user_website_reference_files(selected_website_language, current_user)

        st.markdown(
            f"""
            <div class="app-card">
                <div class="stage-title">{selected_website_language} 웹사이트 원고 프롬프트</div>
                <div class="stage-help">
                    현재 작업자 `{current_user}` 기준 저장값을 사용합니다.<br/>
                    등록 프롬프트 길이: <b>{len(selected_prompt)}</b>자 · 참고 파일 수: <b>{len(reference_files)}</b>개
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.expander("언어별 웹사이트 학습 데이터 관리", expanded=False):
            st.caption("언어별 프롬프트와 레퍼런스 파일을 저장해 두면 웹사이트 원고 생성 시 자동 반영됩니다.")

            prompt_rows = []
            for language_label in WEBSITE_CONTENT_LANGUAGES:
                record = website_prompts.get(language_label, {})
                preview_text = str(record.get("text", _default_website_prompt(language_label)))
                prompt_rows.append(
                    {
                        "언어": language_label,
                        "현재 등록된 프롬프트": preview_text[:120] + ("..." if len(preview_text) > 120 else ""),
                        "수정자": record.get("updated_by", "-"),
                    }
                )
            st.markdown("**A. 웹사이트 프롬프트 관리**")
            st.dataframe(pd.DataFrame(prompt_rows), use_container_width=True, hide_index=True)

            prompt_tab, reference_tab = st.tabs(
                [f"{selected_website_language} 프롬프트 수정", f"{selected_website_language} 레퍼런스 파일 관리"]
            )
            with prompt_tab:
                website_prompt_text = st.text_area(
                    "프롬프트 텍스트",
                    value=selected_prompt,
                    height=180,
                    key=f"website_prompt_text_{selected_website_language}",
                )
                website_prompt_file = st.file_uploader(
                    "또는 .txt / .md 파일 업로드",
                    type=["txt", "md"],
                    key=f"website_prompt_file_{selected_website_language}",
                )
                prompt_action_cols = st.columns(2)
                if prompt_action_cols[0].button("프롬프트 업로드/수정", key=f"save_website_prompt_{selected_website_language}"):
                    try:
                        final_text = website_prompt_text.strip()
                        if website_prompt_file is not None:
                            final_text = website_prompt_file.getvalue().decode("utf-8").strip()
                        if not final_text:
                            raise RuntimeError("저장할 프롬프트 내용을 입력하거나 파일을 업로드해 주세요.")
                        save_user_website_prompt(selected_website_language, final_text, current_user)
                        st.success(f"{selected_website_language} 웹사이트 프롬프트를 저장했습니다.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"웹사이트 프롬프트 저장 중 오류가 발생했습니다: {exc}")
                if prompt_action_cols[1].button("프롬프트 삭제(기본값 복구)", key=f"delete_website_prompt_{selected_website_language}"):
                    delete_user_website_prompt(selected_website_language, current_user)
                    st.success(f"{selected_website_language} 웹사이트 프롬프트를 기본값으로 되돌렸습니다.")
                    st.rerun()

            with reference_tab:
                reference_rows = []
                for language_label in WEBSITE_CONTENT_LANGUAGES:
                    files = list_user_website_reference_files(language_label, current_user)
                    reference_rows.append(
                        {
                            "언어": language_label,
                            "파일 수": len(files),
                            "최근 파일": files[-1].name if files else "-",
                        }
                    )
                st.markdown("**B. 레퍼런스 파일 관리**")
                st.dataframe(pd.DataFrame(reference_rows), use_container_width=True, hide_index=True)
                uploaded_files = st.file_uploader(
                    "docx/txt/md 참고 파일 업로드",
                    type=["docx", "txt", "md"],
                    accept_multiple_files=True,
                    key=f"website_reference_upload_{selected_website_language}",
                    help="같은 파일명으로 다시 업로드하면 교체됩니다.",
                )
                action_cols = st.columns(2)
                if action_cols[0].button("레퍼런스 파일 저장", key=f"save_website_reference_{selected_website_language}"):
                    if uploaded_files:
                        save_uploaded_website_reference_files(selected_website_language, uploaded_files, current_user)
                        st.success(f"{selected_website_language} 레퍼런스 파일을 저장했습니다.")
                        st.rerun()
                    else:
                        st.warning("업로드할 파일을 먼저 선택해 주세요.")
                if action_cols[1].button("전체 레퍼런스 비우기", key=f"clear_website_reference_{selected_website_language}"):
                    for file_path in list_user_website_reference_files(selected_website_language, current_user):
                        delete_user_website_reference_file(selected_website_language, file_path.name, current_user)
                    st.success(f"{selected_website_language} 레퍼런스 파일을 모두 비웠습니다.")
                    st.rerun()
                current_reference_files = list_user_website_reference_files(selected_website_language, current_user)
                if current_reference_files:
                    st.markdown("**현재 등록된 레퍼런스 파일**")
                    for file_path in current_reference_files:
                        file_cols = st.columns([6, 1])
                        file_cols[0].caption(file_path.name)
                        if file_cols[1].button("삭제", key=f"delete_website_ref_{selected_website_language}_{file_path.name}"):
                            delete_user_website_reference_file(selected_website_language, file_path.name, current_user)
                            st.success(f"{file_path.name} 파일을 삭제했습니다.")
                            st.rerun()
                else:
                    st.caption("현재 등록된 레퍼런스 파일이 없습니다.")

        source_text = st.text_area(
            "원문 입력",
            value=st.session_state.get("phase5_source_text", ""),
            height=260,
            placeholder="웹사이트 원고로 재가공할 블로그 본문, 번역본, 직접 작성한 원문을 넣어 주세요.",
            key="phase5_source_text",
        )
        website_additional_request = st.text_area(
            "추가 수정 요청",
            value="",
            height=90,
            placeholder="예: 랜딩페이지형으로 더 간결하게, FAQ를 4개 넣고 CTA는 상담형으로 정리해 주세요.",
            key="phase5_additional_request",
        )
        if st.button("이 원문으로 웹사이트 원고 생성하기", use_container_width=True, type="primary", key="phase5_generate_btn"):
            with st.spinner("Gemini가 웹사이트 SEO/GEO용 원고를 생성하는 중입니다..."):
                try:
                    st.session_state.website_generated_draft = generate_website_copy_with_gemini(
                        source_text=source_text,
                        target_language_label=selected_website_language,
                        additional_request=website_additional_request,
                        clinic_name=clinic_name,
                        operator_name=current_user,
                        seed_keyword=phase5_seed_keyword,
                    )
                    st.session_state.website_generated_context = {
                        "target_language_label": selected_website_language,
                        "clinic_name": clinic_name,
                        "operator_name": current_user,
                        "seed_keyword": phase5_seed_keyword,
                    }
                except Exception as exc:
                    st.error(f"웹사이트 원고 생성 중 오류가 발생했습니다: {exc}")

    draft = st.session_state.generated_draft
    if draft and current_page == "phase3":
        title_col, outline_col = st.columns([1.2, 1])
        with title_col:
            st.markdown("**후킹되는 제목 후보 5개**")
            title_text = "\n".join(f"{idx}. {title}" for idx, title in enumerate(draft["titles"], start=1))
            st.markdown(title_text)
        with outline_col:
            st.markdown("**추천 소제목 구조**")
            outline_text = "\n".join(f"- {item}" for item in draft.get("outline", [])) or "- 소제목 없음"
            st.markdown(outline_text)

        st.markdown("**본문 초안**")
        preview_tab, copy_tab = st.tabs(["서식 미리보기", "복붙용 원문"])
        with preview_tab:
            render_blog_body_preview(draft["body"])
        with copy_tab:
            st.caption("우측 상단 복사 아이콘으로 바로 복사해 사용할 수 있습니다.")
            st.code(draft["body"], language="markdown")
        if draft.get("checklist"):
            st.markdown("**체크리스트**")
            render_checklist_preview(draft["checklist"])
        revision_request = st.text_area(
            "추가 수정 요청",
            value="",
            height=90,
            placeholder="예: 도입부를 더 부드럽게, 통증 설명을 줄이고 유지기간 Q&A를 더 강조해 주세요.",
            key="draft_revision_request",
        )
        if st.button("추가 요청 반영해 2차 초안 받기", use_container_width=True):
            with st.spinner("Gemini가 추가 수정 요청을 반영해 2차 초안을 작성하는 중입니다..."):
                try:
                    context = st.session_state.generated_draft_context or {}
                    st.session_state.generated_draft = revise_blog_draft_with_gemini(
                        original_draft=draft,
                        revision_request=revision_request,
                        summary=context["summary"],
                        clinic_name=context["clinic_name"],
                        tone_style=context["tone_style"],
                        include_cta=context["include_cta"],
                        seed_keyword=context["seed_keyword"],
                        procedure_profile=context["procedure_profile"],
                        reference_urls_text=context["reference_urls_text"],
                        operator_name=context["operator_name"],
                    )
                except Exception as exc:
                    st.error(f"추가 수정 반영 중 오류가 발생했습니다: {exc}")
        st.download_button(
            "본문 초안 TXT 다운로드",
            data=draft["body"],
            file_name=f"naver_blog_draft_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            use_container_width=True,
        )
        render_phase_footer_nav(("메인으로", "home"), ("Phase 1로 이동", "phase1"), ("Phase 2로 이동", "phase2"), ("Phase 4로 이동", "phase4"))

    translated_draft = st.session_state.get("translated_draft")
    if translated_draft and current_page == "phase4":
        st.markdown("**번역 결과**")
        st.caption(f"{translated_draft.get('language_label', selected_translation_language)} ({translated_draft.get('locale', '')})")
        preview_tab, copy_tab = st.tabs(["서식 미리보기", "복붙용 원문"])
        with preview_tab:
            render_blog_body_preview(translated_draft["translated_body"])
        with copy_tab:
            st.caption("우측 상단 복사 아이콘으로 바로 복사해 사용할 수 있습니다.")
            st.code(translated_draft["translated_body"], language="markdown")

        if translated_draft.get("summary_note"):
            st.caption(f"번역 메모: {translated_draft['summary_note']}")

        translation_revision_request = st.text_area(
            "추가 번역 수정 요청",
            value="",
            height=90,
            placeholder="예: 영어 표현을 더 부드럽게, 일본어는 의원 안내 느낌을 더 줄여 주세요.",
            key="phase4_revision_request",
        )
        if st.button("추가 요청 반영해 2차 번역본 받기", use_container_width=True, key="phase4_revision_btn"):
            with st.spinner("Gemini가 추가 번역 수정 요청을 반영하는 중입니다..."):
                try:
                    context = st.session_state.get("translated_draft_context") or {}
                    st.session_state.translated_draft = revise_translation_with_gemini(
                        original_translation=translated_draft["translated_body"],
                        target_language_label=context.get("language_label", selected_translation_language),
                        revision_request=translation_revision_request,
                        seed_keyword=context.get("seed_keyword", seed_keywords[0]),
                        clinic_name=context.get("clinic_name", clinic_name),
                    )
                    st.session_state.translated_draft_context = {
                        **context,
                        "language_label": context.get("language_label", selected_translation_language),
                    }
                except Exception as exc:
                    st.error(f"추가 번역 수정 중 오류가 발생했습니다: {exc}")
        st.download_button(
            "번역본 TXT 다운로드",
            data=translated_draft["translated_body"],
            file_name=f"translated_blog_draft_{translated_draft.get('locale', 'intl')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            use_container_width=True,
        )
        render_phase_footer_nav(("메인으로", "home"), ("Phase 3로 이동", "phase3"), ("Phase 5로 이동", "phase5"))

    website_draft = st.session_state.get("website_generated_draft")
    if website_draft and current_page == "phase5":
        title_col, outline_col = st.columns([1.2, 1])
        with title_col:
            st.markdown("**웹사이트 제목 후보 5개**")
            title_text = "\n".join(f"{idx}. {title}" for idx, title in enumerate(website_draft["titles"], start=1))
            st.markdown(title_text)
        with outline_col:
            st.markdown("**추천 섹션 구조**")
            outline_text = "\n".join(f"- {item}" for item in website_draft.get("outline", [])) or "- 섹션 없음"
            st.markdown(outline_text)

        meta_cols = st.columns(3)
        with meta_cols[0]:
            st.markdown("**썸네일 문구**")
            st.code(website_draft.get("thumbnail_text", ""), language="text")
        with meta_cols[1]:
            st.markdown("**Slug**")
            st.code(website_draft.get("slug", ""), language="text")
        with meta_cols[2]:
            st.markdown("**Description**")
            st.code(website_draft.get("description", ""), language="text")

        st.markdown("**웹사이트 원고 본문**")
        preview_tab, copy_tab = st.tabs(["서식 미리보기", "복붙용 원문"])
        with preview_tab:
            render_blog_body_preview(website_draft["body"])
        with copy_tab:
            st.caption("우측 상단 복사 아이콘으로 바로 복사해 사용할 수 있습니다.")
            st.code(website_draft["body"], language="markdown")
        if website_draft.get("checklist"):
            st.markdown("**체크리스트**")
            render_checklist_preview(website_draft["checklist"])

        website_revision_request = st.text_area(
            "추가 수정 요청",
            value="",
            height=90,
            placeholder="예: H2를 더 명확하게, FAQ를 줄이고 GEO 문구를 더 자연스럽게 넣어 주세요.",
            key="phase5_revision_request",
        )
        if st.button("추가 요청 반영해 2차 웹사이트 원고 받기", use_container_width=True, key="phase5_revision_btn"):
            with st.spinner("Gemini가 추가 수정 요청을 반영해 웹사이트 원고를 다듬는 중입니다..."):
                try:
                    context = st.session_state.get("website_generated_context") or {}
                    st.session_state.website_generated_draft = revise_website_copy_with_gemini(
                        original_draft=website_draft,
                        revision_request=website_revision_request,
                        target_language_label=context.get("target_language_label", selected_website_language),
                        clinic_name=context.get("clinic_name", clinic_name),
                        operator_name=context.get("operator_name", current_operator_name()),
                        seed_keyword=context.get("seed_keyword", seed_keywords[0]),
                    )
                except Exception as exc:
                    st.error(f"웹사이트 원고 수정 중 오류가 발생했습니다: {exc}")
        st.download_button(
            "웹사이트 원고 TXT 다운로드",
            data=website_draft["body"],
            file_name=f"website_copy_{website_draft.get('language_label', 'ko')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            use_container_width=True,
        )
        render_phase_footer_nav(("메인으로", "home"), ("Phase 3로 이동", "phase3"), ("Phase 4로 이동", "phase4"))


if __name__ == "__main__":
    main()
