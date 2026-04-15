from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook


HEADER_CANDIDATES = {
    "keyword",
    "keywords",
    "키워드",
    "검색어",
    "해시태그",
    "hashtag",
    "tag",
}

PREFERRED_SHEETS = [
    "키워드 확정(10.15)",
    "키워드 확정",
    "시술명 미포함",
]


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "")


def _clean_keyword(value: str) -> str:
    text = value.strip()
    if text.startswith("#"):
        text = text[1:].strip()
    return text


def _choose_sheet(wb, sheet_name: str) -> Tuple[str, object]:
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return sheet_name, wb[sheet_name]

    for preferred in PREFERRED_SHEETS:
        if preferred in wb.sheetnames:
            return preferred, wb[preferred]

    return wb.sheetnames[0], wb[wb.sheetnames[0]]


def _is_all_sheets(sheet_name: str) -> bool:
    return sheet_name.strip().lower() in {"*", "all", "전체"}


def _find_header_row(ws, scan_rows: int) -> Tuple[Optional[int], List[int]]:
    rows = ws.iter_rows(values_only=True)
    for idx in range(scan_rows):
        row = next(rows, None)
        if row is None:
            break
        normalized = [
            _normalize_header(str(v)) if v is not None else "" for v in row
        ]
        cols = [i for i, h in enumerate(normalized) if h in HEADER_CANDIDATES]
        if cols:
            return idx, cols
    return None, []


def _extract_from_sheet(
    ws,
    header_scan_rows: int,
    keywords: List[str],
    seen: set,
) -> None:
    header_row_index, keyword_cols = _find_header_row(ws, header_scan_rows)

    def add_value(value: str) -> None:
        cleaned = _clean_keyword(value)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            keywords.append(cleaned)

    rows = list(ws.iter_rows(values_only=True))
    if header_row_index is None:
        for row in rows:
            if not row:
                continue
            value = row[0] if row else None
            if value is None:
                continue
            add_value(str(value))
        return

    for row in rows[header_row_index + 1 :]:
        if not row:
            continue
        for col in keyword_cols:
            if col >= len(row):
                continue
            value = row[col]
            if value is None:
                continue
            add_value(str(value))


def load_keywords_xlsx(
    path: str,
    sheet_name: str = "",
    header_scan_rows: int = 30,
) -> Tuple[List[str], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    keywords: List[str] = []
    seen = set()

    if _is_all_sheets(sheet_name):
        for name in wb.sheetnames:
            _extract_from_sheet(wb[name], header_scan_rows, keywords, seen)
        return keywords, "ALL"

    chosen_name, ws = _choose_sheet(wb, sheet_name)
    _extract_from_sheet(ws, header_scan_rows, keywords, seen)
    return keywords, chosen_name
